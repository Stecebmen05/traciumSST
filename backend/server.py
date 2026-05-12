from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query, UploadFile, File, Header
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import io
import base64
import requests
import bcrypt
import jwt as pyjwt
import resend
import asyncio
from standards_bank import STANDARDS_BANK, DECRETO_1072_COMPONENTS, get_applicable_standards, get_total_weight, calculate_score, CLASSIFICATION_THRESHOLDS, PHVA_WEIGHTS, STANDARD_GROUPS, CRITERIA_VERIFICATION, PESV_STANDARDS, get_pesv_standards, calculate_pesv_score, PESV_PHASES

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== PASSWORD & JWT HELPERS ====================

JWT_SECRET = os.environ.get("JWT_SECRET", "fallback_secret_change_me")
JWT_ALGORITHM = "HS256"

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_session_token() -> str:
    return str(uuid.uuid4())

# ==================== EMAIL (Resend) ====================
resend.api_key = os.environ.get("RESEND_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")

async def send_email(to: str, subject: str, html: str):
    """Send email via Resend (non-blocking)"""
    if not resend.api_key:
        logger.warning("RESEND_API_KEY not set, skipping email")
        return None
    try:
        params = {"from": SENDER_EMAIL, "to": [to], "subject": subject, "html": html}
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {to}: {subject}")
        return result
    except Exception as e:
        logger.error(f"Email send failed to {to}: {e}")
        return None


async def _notify_approvers(company_id: str, subject: str, html: str, exclude_email: str = ""):
    """Notify all admin/sgsst_manager users in the company (plus the platform owner) via email."""
    if not company_id:
        return
    query = {
        "role": {"$in": ["admin", "owner", "sgsst_manager"]},
        "active": {"$ne": False},
        "$or": [
            {"company_ids": company_id},
            {"email": OWNER_EMAIL},
        ],
    }
    approvers = await db.users.find(query, {"_id": 0, "email": 1, "name": 1, "role": 1}).to_list(50)
    sent = 0
    for approver in approvers:
        email = (approver.get("email") or "").strip()
        if not email or email == exclude_email:
            continue
        result = await send_email(email, subject, html)
        if result:
            sent += 1
    logger.info(f"Approval notification: sent to {sent}/{len(approvers)} approvers for company {company_id}")
    return sent


async def _create_notification(user_id: str, ntype: str, title: str, message: str, link: str = "", related_id: str = "", company_id: str = ""):
    """Create an in-app notification document."""
    doc = {
        "notification_id": uuid.uuid4().hex,
        "user_id": user_id,
        "type": ntype,
        "title": title,
        "message": message,
        "link": link,
        "related_id": related_id,
        "company_id": company_id,
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(doc)
    return doc


async def _notify_action_plan_change(plan: dict, audit: dict, change_type: str, actor: dict, extra: str = ""):
    """Send email + in-app notifications to interested users when a plan changes.
    Recipients: admin/sgsst_manager of the company AND the responsible user (matched by name) - excluding the actor.
    change_type: created | updated | follow_up | closed
    """
    company_id = plan.get("company_id") or audit.get("company_id", "")
    if not company_id:
        return
    actor_email = (actor.get("email") or "").lower().strip()
    actor_id = actor.get("user_id", "")

    # Build recipient set: admins + sgsst_managers of company + owner + user whose name matches responsible
    query = {
        "active": {"$ne": False},
        "$or": [
            {"role": {"$in": ["admin", "owner", "sgsst_manager"]}, "company_ids": company_id},
            {"email": OWNER_EMAIL},
            {"name": plan.get("responsible", ""), "company_ids": company_id} if plan.get("responsible") else {"_unused": True},
        ],
    }
    recipients = await db.users.find(query, {"_id": 0, "email": 1, "name": 1, "user_id": 1}).to_list(50)
    seen = set()
    unique = []
    for r in recipients:
        rid = r.get("user_id", "")
        if rid and rid not in seen and rid != actor_id and (r.get("email") or "").lower() != actor_email:
            seen.add(rid)
            unique.append(r)

    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0, "name": 1}) or {}
    cname = company.get("name", "N/A")

    type_label = {"corrective": "Correctiva", "preventive": "Preventiva", "improvement": "Mejora"}.get(plan.get("action_type", ""), plan.get("action_type", ""))
    change_label = {
        "created": "creado un nuevo",
        "updated": "actualizado un",
        "follow_up": "agregado seguimiento al",
        "closed": "cerrado un",
    }.get(change_type, "modificado un")

    subject = f"TraciumSST - Cambio en plan de accion: {plan.get('action','')[:60]}"
    title_in_app = {
        "created": "Nuevo plan de accion",
        "updated": "Plan de accion actualizado",
        "follow_up": "Seguimiento agregado",
        "closed": "Plan de accion cerrado",
    }.get(change_type, "Cambio en plan de accion")

    short_action = (plan.get("action") or "")[:120]
    in_app_message = f"{actor.get('name','Un usuario')} ha {change_label} plan ({type_label}) en {audit.get('title','la auditoria')}."

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto">
      <div style="background:#0047AB;padding:20px;color:#fff;border-radius:8px 8px 0 0">
        <h2 style="margin:0">TraciumSST - Plan de Accion</h2>
        <p style="margin:4px 0 0 0;font-size:13px;opacity:0.9">{actor.get('name','Un usuario')} ha {change_label} plan</p>
      </div>
      <div style="padding:20px;background:#F8F9FA;border:1px solid #E2E8F0;border-top:0;border-radius:0 0 8px 8px">
        <table style="width:100%;border-collapse:collapse;margin:12px 0">
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Empresa</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{cname}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Auditoria</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('title','')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Tipo de Accion</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{type_label}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Responsable</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{plan.get('responsible','N/A')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Fecha Limite</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{plan.get('due_date','N/A')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Accion</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{plan.get('action','')[:300]}</td></tr>
          {f'<tr><td style="padding:8px;background:#FFF8E1;border:1px solid #E2E8F0"><b>Seguimiento</b></td><td style="padding:8px;background:#FFF8E1;border:1px solid #E2E8F0">{extra[:300]}</td></tr>' if extra else ''}
        </table>
        <p>Ingresa a <a href="{os.environ.get('FRONTEND_URL','')}/audits" style="color:#0047AB;text-decoration:none;font-weight:bold">TraciumSST - Auditorias</a> para revisar.</p>
        <p style="color:#94A3B8;font-size:11px;margin-top:16px">Este es un mensaje automatico. No respondas.</p>
      </div>
    </div>
    """

    # Send email and create in-app notification per recipient
    for r in unique:
        await _create_notification(
            user_id=r.get("user_id", ""),
            ntype=f"action_plan_{change_type}",
            title=title_in_app,
            message=in_app_message + (f" Nota: {extra[:120]}" if extra else ""),
            link=f"/audits",
            related_id=plan.get("plan_id", ""),
            company_id=company_id,
        )
        if r.get("email"):
            await send_email(r["email"], subject, html)
    logger.info(f"Action plan notification ({change_type}) sent to {len(unique)} users for plan {plan.get('plan_id')}")

# ==================== AUDITOR CREDENTIALS ====================
AUDITOR_SIGNATURE = {
    "name": "STEPHANIA CEBALLOS MENDOZA",
    "title_lines": [
        "Especialista en Seguridad y Salud en el Trabajo",
        "Licencia N. 201806023926",
        "Auditor Interno en Sistemas de Gestion Integrados HSEQ",
        "ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018",
        "Lider Implementador en Sistemas de Gestion Integrados HSEQ",
        "ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018",
    ],
    "annexes": [
        {"title": "Certificado Lider Implementador HSEQ - SGS Academy", "desc": "Lider Implementador en Sistemas de Gestion Integrados HSEQ - ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018. SGS / Positiva Compania de Seguros S.A. 80 horas. Sept-Dic 2024. Cert. 152847641/178316595."},
        {"title": "Certificado Auditor Interno HSEQ - SGS Academy", "desc": "Auditor Interno en Sistemas de Gestion Integrados HSEQ - ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018. SGS / Positiva Compania de Seguros S.A. 80 horas. Sept-Dic 2024. Cert. 152847446/178316605."},
        {"title": "Licencia SST - Secretaria Seccional de Salud de Antioquia", "desc": "Resolucion S 2018060239266 del 27/08/2018. Psicologa, Especialista en Gerencia de la Salud Ocupacional. Licencia para ofertar servicios de SST a nivel nacional. Vigencia: 10 anos. Areas: Diseno, Administracion y Ejecucion del SG-SST, Investigacion de AT, Educacion y Capacitacion."},
    ]
}

# ==================== OBJECT STORAGE ====================

STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "safeguard-sst"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    try:
        resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
        resp.raise_for_status()
        storage_key = resp.json()["storage_key"]
        logger.info("Object storage initialized")
        return storage_key
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
        return None

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    if not key:
        raise HTTPException(status_code=503, detail="Storage not available")
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    if not key:
        raise HTTPException(status_code=503, detail="Storage not available")
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ==================== RBAC ====================

ROLE_HIERARCHY = {
    "owner": 200,
    "admin": 100,
    "sgsst_manager": 80,
    "auditor": 60,
    "area_leader": 40,
    "collaborator": 20
}

WRITE_ROLES = {"owner", "admin", "sgsst_manager"}
AUDIT_WRITE_ROLES = {"owner", "admin", "sgsst_manager", "auditor"}
INCIDENT_REPORT_ROLES = {"owner", "admin", "sgsst_manager", "area_leader", "collaborator"}
READ_ROLES = {"owner", "admin", "sgsst_manager", "auditor", "area_leader", "collaborator"}

# Owner email - has full access to all companies
OWNER_EMAIL = "stephaniaceballosmendoza@gmail.com"

def is_owner(user):
    return user.get("role") == "owner" or user.get("email") == OWNER_EMAIL

def require_role(*allowed_roles):
    async def checker(request: Request):
        user = await get_current_user(request)
        if is_owner(user):
            return user
        if user.get("role", "collaborator") not in allowed_roles:
            raise HTTPException(status_code=403, detail="Permiso insuficiente para esta accion")
        return user
    return checker

def can_write(user):
    return is_owner(user) or user.get("role", "collaborator") in WRITE_ROLES

def can_audit_write(user):
    return is_owner(user) or user.get("role", "collaborator") in AUDIT_WRITE_ROLES

# ==================== AUTH HELPERS ====================

EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

async def get_current_user(request: Request):
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session_doc = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    if not user_doc.get("active", True):
        raise HTTPException(status_code=403, detail="Cuenta inhabilitada")
    return user_doc

# Multi-company helper: get company_id from user
def get_company_id(user):
    return user.get("active_company_id", "default")

# ==================== PYDANTIC MODELS ====================

class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "collaborator"
    created_at: Optional[datetime] = None

class DocumentCreate(BaseModel):
    title: str
    category: str
    description: Optional[str] = ""
    version: str = "1.0"

class DocumentOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    doc_id: str
    title: str
    category: str
    description: str = ""
    version: str = "1.0"
    status: str = "active"
    created_by: str = ""
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    approval_status: Optional[str] = None
    approval_history: Optional[List[Dict[str, Any]]] = None
    submitted_by: Optional[str] = None
    submitted_at: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    rejection_reason: Optional[str] = None
    rejected_by: Optional[str] = None
    rejected_at: Optional[str] = None

class HazardCreate(BaseModel):
    area: str
    hazard_type: str
    description: str
    risk_source: str
    probability: int = Field(ge=1, le=5)
    severity: int = Field(ge=1, le=5)
    existing_controls: str = ""
    proposed_controls: str = ""

class HazardOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    hazard_id: str
    area: str
    hazard_type: str
    description: str
    risk_source: str
    probability: int
    severity: int
    risk_level: int
    risk_category: str
    existing_controls: str = ""
    proposed_controls: str = ""
    status: str = "active"
    created_at: Optional[str] = None

class IncidentCreate(BaseModel):
    incident_type: str
    date: str
    location: str
    description: str
    affected_person: str = ""
    severity: str = "minor"
    immediate_actions: str = ""

class IncidentOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    incident_id: str
    incident_type: str
    date: str
    location: str
    description: str
    affected_person: str = ""
    severity: str = "minor"
    immediate_actions: str = ""
    root_cause: str = ""
    corrective_actions: str = ""
    status: str = "open"
    created_by: str = ""
    created_at: Optional[str] = None

class TrainingCreate(BaseModel):
    title: str
    description: str = ""
    trainer: str = ""
    scheduled_date: str
    duration_hours: float = 1
    max_participants: int = 30

class TrainingOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    training_id: str
    title: str
    description: str = ""
    trainer: str = ""
    scheduled_date: str
    duration_hours: float = 1
    max_participants: int = 30
    attendees: List[str] = []
    status: str = "scheduled"
    effectiveness_score: Optional[float] = None
    created_at: Optional[str] = None

class AuditCreate(BaseModel):
    title: str
    audit_type: str
    scheduled_date: str
    auditor: str = ""
    scope: str = ""

class AuditOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    audit_id: str
    title: str
    audit_type: str
    scheduled_date: str
    auditor: str = ""
    scope: str = ""
    status: str = "planned"
    findings_count: int = 0
    created_at: Optional[str] = None

class FindingCreate(BaseModel):
    audit_id: str
    finding_type: str
    description: str
    area: str = ""
    corrective_action: str = ""
    responsible: str = ""
    due_date: str = ""

class FindingOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    finding_id: str
    audit_id: str
    finding_type: str
    description: str
    area: str = ""
    corrective_action: str = ""
    responsible: str = ""
    due_date: str = ""
    status: str = "open"
    created_at: Optional[str] = None

class ActivityCreate(BaseModel):
    title: str
    description: str = ""
    responsible: str = ""
    due_date: str
    category: str = "general"
    priority: str = "medium"

class ActivityOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    activity_id: str
    title: str
    description: str = ""
    responsible: str = ""
    due_date: str
    category: str = "general"
    priority: str = "medium"
    status: str = "pending"
    completion_percentage: int = 0
    created_at: Optional[str] = None

class ChecklistItemCreate(BaseModel):
    standard: str
    description: str
    compliant: bool = False
    evidence: str = ""
    observations: str = ""

class ChecklistItemOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    item_id: str
    standard: str
    description: str
    compliant: bool = False
    evidence: str = ""
    observations: str = ""
    updated_at: Optional[str] = None

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/session")
async def exchange_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        async with httpx.AsyncClient(timeout=15.0) as http_client:
            resp = await http_client.get(EMERGENT_AUTH_URL, headers={"X-Session-ID": session_id})
    except httpx.TimeoutException:
        logger.error("Timeout connecting to Emergent Auth")
        raise HTTPException(status_code=504, detail="Auth service timeout")
    except Exception as e:
        logger.error(f"Error connecting to Emergent Auth: {e}")
        raise HTTPException(status_code=502, detail="Auth service unavailable")
    if resp.status_code != 200:
        logger.warning(f"Emergent Auth returned {resp.status_code} for session exchange")
        raise HTTPException(status_code=401, detail="Sesion invalida o expirada. Intenta iniciar sesion de nuevo.")
    data = resp.json()
    email = data.get("email")
    name = data.get("name", "")
    picture = data.get("picture", "")
    session_token = data.get("session_token", str(uuid.uuid4()))
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    if existing_user:
        user_id = existing_user["user_id"]
        update_data = {"name": name, "picture": picture, "auth_type": "google"}
        # Ensure owner email always gets owner role
        if email == OWNER_EMAIL:
            update_data["role"] = "owner"
        await db.users.update_one({"email": email}, {"$set": update_data})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        role = "owner" if email == OWNER_EMAIL else "collaborator"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": name,
            "picture": picture,
            "role": role,
            "auth_type": "google",
            "company_ids": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 3600
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return user_doc

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/", samesite="none", secure=True)
    return {"message": "Logged out"}


@api_router.post("/auth/login-email")
async def login_email(request: Request, response: Response):
    """Email/password login"""
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email y contraseña son obligatorios")
    # Brute force check
    ip = request.client.host if request.client else "unknown"
    ident = f"{ip}:{email}"
    attempts = await db.login_attempts.find_one({"identifier": ident}, {"_id": 0})
    if attempts and attempts.get("count", 0) >= 5:
        locked_until = attempts.get("locked_until", "")
        if locked_until:
            if isinstance(locked_until, str):
                locked_until = datetime.fromisoformat(locked_until)
            if locked_until.tzinfo is None:
                locked_until = locked_until.replace(tzinfo=timezone.utc)
            if locked_until > datetime.now(timezone.utc):
                raise HTTPException(status_code=429, detail="Demasiados intentos fallidos. Intenta en 15 minutos.")
            else:
                await db.login_attempts.delete_one({"identifier": ident})
    # Find user
    user_doc = await db.users.find_one({"email": email}, {"_id": 0})
    if not user_doc or not user_doc.get("password_hash"):
        # Increment failed attempts
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1}, "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    if not verify_password(password, user_doc["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1}, "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    # Check if user is disabled
    if not user_doc.get("active", True):
        raise HTTPException(status_code=403, detail="Tu cuenta esta inhabilitada. Contacta al administrador.")
    # Check if demo user expired
    if user_doc.get("is_demo") and user_doc.get("demo_expires_at"):
        exp = user_doc["demo_expires_at"]
        if isinstance(exp, str):
            exp_dt = datetime.fromisoformat(exp)
        else:
            exp_dt = exp
        if exp_dt.tzinfo is None:
            exp_dt = exp_dt.replace(tzinfo=timezone.utc)
        if exp_dt < datetime.now(timezone.utc):
            raise HTTPException(status_code=403, detail="Tu cuenta de prueba ha expirado. Contacta al administrador.")
    # Success - clear attempts
    await db.login_attempts.delete_one({"identifier": ident})
    # Create session
    session_token = create_session_token()
    await db.user_sessions.insert_one({
        "user_id": user_doc["user_id"],
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    response.set_cookie(key="session_token", value=session_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    safe_user = {k: v for k, v in user_doc.items() if k != "password_hash"}
    return safe_user


@api_router.post("/auth/create-user")
async def create_user_with_password(request: Request, user=Depends(require_role("admin"))):
    """Admin creates a user with email/password"""
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    name = body.get("name") or ""
    role = body.get("role", "collaborator")
    company_id = body.get("company_id", "")
    if not email or not password or not name:
        raise HTTPException(status_code=400, detail="Email, contraseña y nombre son obligatorios")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    if role not in ROLE_HIERARCHY:
        raise HTTPException(status_code=400, detail="Rol invalido")
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=409, detail="Ya existe un usuario con ese email")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    # Strict company isolation: non-admin users get EXACTLY [company_id] (no defaults)
    if role in ("admin", "owner"):
        company_ids_list = [company_id] if company_id else user.get("company_ids", [])
        active_cid = company_id or user.get("active_company_id", "")
    else:
        if not company_id:
            raise HTTPException(status_code=400, detail="Debes asignar una empresa al crear un usuario con rol no-admin")
        company_ids_list = [company_id]
        active_cid = company_id
    new_user = {
        "user_id": user_id,
        "email": email,
        "name": name,
        "password_hash": hash_password(password),
        "role": role,
        "active_company_id": active_cid,
        "company_ids": company_ids_list,
        "auth_type": "email",
        "picture": "",
        "created_by": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_user)
    new_user.pop("_id", None)
    new_user.pop("password_hash", None)
    return new_user


@api_router.put("/users/{user_id}/password")
async def change_user_password(user_id: str, request: Request, admin=Depends(require_role("admin"))):
    """Admin changes a user's password"""
    body = await request.json()
    new_password = body.get("password", "")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    result = await db.users.update_one({"user_id": user_id}, {"$set": {"password_hash": hash_password(new_password)}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Contraseña actualizada"}


@api_router.put("/users/{user_id}/toggle-status")
async def toggle_user_status(user_id: str, admin=Depends(require_role("admin"))):
    """Admin enables/disables a user"""
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if user_doc.get("email") == OWNER_EMAIL:
        raise HTTPException(status_code=400, detail="No se puede inhabilitar al propietario")
    new_status = not user_doc.get("active", True)
    await db.users.update_one({"user_id": user_id}, {"$set": {"active": new_status}})
    # If disabling, kill all sessions
    if not new_status:
        await db.user_sessions.delete_many({"user_id": user_id})
    label = "habilitado" if new_status else "inhabilitado"
    return {"message": f"Usuario {label}", "active": new_status}


@api_router.post("/users/create-demo")
async def create_demo_user(request: Request, admin=Depends(require_role("admin"))):
    """Create a demo/trial user with limited access and expiration"""
    body = await request.json()
    name = body.get("name", "Usuario Demo")
    company_id = body.get("company_id", "")
    days = body.get("days", 7)
    role = body.get("role", "collaborator")
    demo_id = uuid.uuid4().hex[:6]
    email = f"demo_{demo_id}@traciumsst.com"
    password = f"Demo{demo_id}!"
    expires_at = (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    # Strict company isolation for demo users
    if role in ("admin", "owner"):
        company_ids_list = [company_id] if company_id else admin.get("company_ids", [])
        active_cid = company_id or admin.get("active_company_id", "")
    else:
        active_cid = company_id or admin.get("active_company_id", "")
        if not active_cid:
            raise HTTPException(status_code=400, detail="Empresa requerida para demo con rol no-admin")
        company_ids_list = [active_cid]
    new_user = {
        "user_id": user_id, "email": email, "name": name,
        "password_hash": hash_password(password), "role": role,
        "active_company_id": active_cid,
        "company_ids": company_ids_list,
        "auth_type": "email", "is_demo": True, "active": True,
        "demo_expires_at": expires_at, "picture": "",
        "created_by": admin.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_user)
    new_user.pop("_id", None)
    new_user.pop("password_hash", None)
    new_user["demo_password"] = password
    return new_user


@api_router.put("/users/{user_id}/company")
async def assign_user_company(user_id: str, request: Request, admin=Depends(require_role("admin"))):
    """Admin assigns a company to a user. REPLACES any previous assignment (strict isolation)."""
    body = await request.json()
    company_id = body.get("company_id", "")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id es obligatorio")
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    # For non-admin/non-owner users, STRICTLY replace company_ids with the single assignment.
    # For admin/owner users, allow multiple companies (addToSet).
    if target.get("role") in ("admin", "owner") or target.get("email") == OWNER_EMAIL:
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"active_company_id": company_id}, "$addToSet": {"company_ids": company_id}}
        )
    else:
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"active_company_id": company_id, "company_ids": [company_id]}}
        )
    return {"message": "Empresa asignada"}


# ==================== DASHBOARD ====================

@api_router.get("/dashboard")
async def get_dashboard(user=Depends(get_current_user)):
    cid = get_company_id(user)
    q = {"company_id": cid}
    total_docs = await db.documents.count_documents(q)
    total_hazards = await db.hazards.count_documents(q)
    high_risks = await db.hazards.count_documents({**q, "risk_category": {"$in": ["high", "critical"]}})
    total_incidents = await db.incidents.count_documents(q)
    open_incidents = await db.incidents.count_documents({**q, "status": "open"})
    total_trainings = await db.trainings.count_documents(q)
    completed_trainings = await db.trainings.count_documents({**q, "status": "completed"})
    total_audits = await db.audits.count_documents(q)
    total_findings = await db.findings.count_documents({**q, "status": {"$ne": "resolved_by_compliance"}})
    open_findings = await db.findings.count_documents({**q, "status": "open"})
    total_activities = await db.activities.count_documents(q)
    completed_activities = await db.activities.count_documents({**q, "status": "completed"})
    plan_progress = round((completed_activities / total_activities * 100) if total_activities > 0 else 0, 1)
    recent_incidents = await db.incidents.find(q, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_findings = await db.findings.find({**q, "status": {"$ne": "resolved_by_compliance"}}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    # Action plans summary
    total_plans = await db.action_plans.count_documents({**q, "status": {"$in": ["open", "in_progress"]}})
    overdue_plans = 0
    open_plans = await db.action_plans.find({**q, "status": {"$in": ["open", "in_progress"]}, "due_date": {"$ne": ""}}, {"_id": 0}).to_list(200)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for p in open_plans:
        if p.get("due_date", "") < today and p.get("due_date", ""):
            overdue_plans += 1
    # Latest audit score
    latest_audit = await db.audits.find(q, {"_id": 0}).sort("created_at", -1).limit(1).to_list(1)
    audit_score = None
    if latest_audit and latest_audit[0].get("score_result"):
        audit_score = latest_audit[0]["score_result"]
        audit_score["audit_title"] = latest_audit[0].get("title", "")
        audit_score["audit_type"] = latest_audit[0].get("audit_type", "")
    return {
        "kpis": {
            "total_documents": total_docs, "total_hazards": total_hazards,
            "high_risk_hazards": high_risks, "total_incidents": total_incidents,
            "open_incidents": open_incidents, "total_trainings": total_trainings,
            "completed_trainings": completed_trainings, "total_audits": total_audits,
            "total_findings": total_findings, "open_findings": open_findings,
            "plan_progress": plan_progress, "total_activities": total_activities,
            "completed_activities": completed_activities,
            "total_plans": total_plans, "overdue_plans": overdue_plans,
        },
        "recent_incidents": recent_incidents,
        "recent_findings": recent_findings,
        "audit_score": audit_score,
    }

# ==================== CONSULTANT DASHBOARD ====================

@api_router.get("/consultant/dashboard")
async def get_consultant_dashboard(user=Depends(get_current_user)):
    """Consolidated dashboard across all companies"""
    companies = await db.companies.find({}, {"_id": 0}).to_list(100)
    result = []
    for comp in companies:
        cid = comp["company_id"]
        std_items = await db.standards_compliance.find({"company_id": cid, "applicable": True}, {"_id": 0}).to_list(200)
        compliant = [i for i in std_items if i.get("compliant")]
        total_w = sum(i.get("weight", 0) for i in std_items)
        comp_w = sum(i.get("weight", 0) for i in compliant)
        score = round((comp_w / total_w * 100) if total_w > 0 else 0, 1)
        audits = await db.audits.count_documents({"company_id": cid})
        findings_open = await db.findings.count_documents({"company_id": cid, "status": "open"})
        incidents = await db.incidents.count_documents({"company_id": cid})
        result.append({
            "company_id": cid,
            "name": comp.get("name", ""),
            "workers_count": comp.get("workers_count", 0),
            "risk_level": comp.get("risk_level", 1),
            "compliance_score": score,
            "total_standards": len(std_items),
            "compliant_standards": len(compliant),
            "total_audits": audits,
            "open_findings": findings_open,
            "total_incidents": incidents,
        })
    return result

# ==================== IMPLEMENTATION MODULE ====================

@api_router.get("/checklist", response_model=List[ChecklistItemOut])
async def get_checklist(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.checklist.find({"company_id": cid}, {"_id": 0}).to_list(200)
    return items

@api_router.post("/checklist", response_model=ChecklistItemOut)
async def create_checklist_item(item: ChecklistItemCreate, user=Depends(get_current_user)):
    doc = item.model_dump()
    doc["item_id"] = f"chk_{uuid.uuid4().hex[:8]}"
    doc["company_id"] = get_company_id(user)
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.checklist.insert_one(doc)
    return ChecklistItemOut(**doc)

@api_router.put("/checklist/{item_id}")
async def update_checklist_item(item_id: str, item: ChecklistItemCreate, user=Depends(get_current_user)):
    update_data = item.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.checklist.update_one({"item_id": item_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    updated = await db.checklist.find_one({"item_id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/checklist/{item_id}")
async def delete_checklist_item(item_id: str, user=Depends(get_current_user)):
    result = await db.checklist.delete_one({"item_id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Deleted"}

# Activities (Annual Plan)
@api_router.get("/activities", response_model=List[ActivityOut])
async def get_activities(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.activities.find({"company_id": cid}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/activities", response_model=ActivityOut)
async def create_activity(item: ActivityCreate, user=Depends(get_current_user)):
    doc = item.model_dump()
    doc["activity_id"] = f"act_{uuid.uuid4().hex[:8]}"
    doc["status"] = "pending"
    doc["completion_percentage"] = 0
    doc["company_id"] = get_company_id(user)
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.activities.insert_one(doc)
    return ActivityOut(**doc)

@api_router.put("/activities/{activity_id}")
async def update_activity(activity_id: str, updates: Dict[str, Any], user=Depends(get_current_user)):
    updates.pop("_id", None)
    updates.pop("activity_id", None)
    result = await db.activities.update_one({"activity_id": activity_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    updated = await db.activities.find_one({"activity_id": activity_id}, {"_id": 0})
    return updated

@api_router.delete("/activities/{activity_id}")
async def delete_activity(activity_id: str, user=Depends(get_current_user)):
    result = await db.activities.delete_one({"activity_id": activity_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    return {"message": "Deleted"}

# ==================== DOCUMENTS MODULE ====================

@api_router.get("/documents", response_model=List[DocumentOut])
async def get_documents(user=Depends(get_current_user)):
    cid = get_company_id(user)
    docs = await db.documents.find({"company_id": cid}, {"_id": 0}).to_list(500)
    return docs

@api_router.post("/documents", response_model=DocumentOut)
async def create_document(doc: DocumentCreate, user=Depends(require_role("admin", "sgsst_manager"))):
    d = doc.model_dump()
    d["doc_id"] = f"doc_{uuid.uuid4().hex[:8]}"
    d["status"] = "active"
    d["created_by"] = user.get("name", "")
    d["company_id"] = get_company_id(user)
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    d["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.documents.insert_one(d)
    return DocumentOut(**d)

@api_router.put("/documents/{doc_id}")
async def update_document(doc_id: str, updates: Dict[str, Any], user=Depends(require_role("admin", "sgsst_manager"))):
    updates.pop("_id", None)
    updates.pop("doc_id", None)
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.documents.update_one({"doc_id": doc_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    updated = await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})
    return updated

@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, user=Depends(require_role("admin", "sgsst_manager"))):
    result = await db.documents.delete_one({"doc_id": doc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Deleted"}

# ---- Document Approval Workflow ----

# ---- AI Document Templates Generator ----
from document_templates import list_templates as _list_doc_templates, get_template as _get_doc_template, CATEGORIES as DOC_TEMPLATE_CATEGORIES


@api_router.get("/documents/templates")
async def list_document_templates(category: str = "", user=Depends(get_current_user)):
    """Return the catalog of AI-generated document templates."""
    items = _list_doc_templates(category)
    # Strip the system_prompt; the user does not need it
    public = [{k: v for k, v in t.items() if k != "system_prompt"} for t in items]
    return {"items": public, "categories": DOC_TEMPLATE_CATEGORIES}


@api_router.post("/documents/generate-ai")
async def generate_document_with_ai(request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    """Generate a full document using AI based on a template. Persists as a draft document."""
    body = await request.json()
    template_id = body.get("template_id", "")
    customizations = (body.get("customizations") or "").strip()
    save_as_document = bool(body.get("save", True))
    requested_cid = (body.get("company_id") or "").strip()
    template = _get_doc_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")

    # Resolve target company: explicit body.company_id (validate access) or active company
    if requested_cid:
        if is_owner(user) or user.get("role") == "admin":
            cid = requested_cid
        elif requested_cid in (user.get("company_ids", []) or []):
            cid = requested_cid
        else:
            raise HTTPException(status_code=403, detail="No tienes acceso a esa empresa")
    else:
        cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}

    company_block = (
        f"DATOS DE LA EMPRESA (usalos en el documento):\n"
        f"- Razon social: {company.get('name', 'N/A')}\n"
        f"- NIT: {company.get('nit', 'N/A')}\n"
        f"- Ciudad: {company.get('city', 'N/A')}\n"
        f"- Actividad economica: {company.get('economic_activity', 'N/A')}\n"
        f"- Trabajadores: {company.get('workers_count', 'N/A')}\n"
        f"- Nivel de riesgo (Decreto 1607/2002): {company.get('risk_level', 'N/A')}\n"
        f"- Sedes: {', '.join(company.get('sedes', []) or []) or 'Sede principal'}\n"
        f"- Procesos: {', '.join(company.get('processes', []) or []) or 'N/A'}\n"
    )
    if customizations:
        company_block += f"\nINSTRUCCIONES ADICIONALES DEL USUARIO:\n{customizations}\n"
    company_block += (
        "\nINSTRUCCIONES DE FORMATO:\n"
        "- Usa formato Markdown limpio (encabezados con #, ##, ###, listas con -, negritas con **).\n"
        "- Encabezado del documento con: Logo (placeholder [LOGO EMPRESA]), Razon social, NIT, ciudad y fecha actual.\n"
        "- Numera articulos cuando aplique.\n"
        "- Cierre con linea de firma del Representante Legal y otros firmantes que correspondan.\n"
        "- Espanol formal colombiano sin emojis.\n"
    )

    system_prompt = template["system_prompt"] + "\n\n" + company_block

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")
        chat = LlmChat(
            api_key=api_key,
            session_id=f"doc_gen_{uuid.uuid4().hex[:10]}",
            system_message=system_prompt,
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text=f"Genera el documento '{template['title']}' completo en formato Markdown listo para revisar.")
        content = await chat.send_message(msg)
    except Exception as e:
        logger.error(f"AI document generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Error al generar el documento: {str(e)}")

    if save_as_document:
        doc_id = f"doc_{uuid.uuid4().hex[:10]}"
        d = {
            "doc_id": doc_id,
            "title": template["title"],
            "category": template["category"],
            "description": f"Generado con IA desde la plantilla {template['template_id']}",
            "version": "1.0",
            "status": "active",
            "approval_status": None,
            "ai_generated_content": content,
            "template_id": template_id,
            "company_id": cid,
            "created_by": user.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.documents.insert_one(d)
        d.pop("_id", None)
        return {"document": d, "content": content}
    return {"content": content, "template_id": template_id}


@api_router.get("/documents/{doc_id}/ai-content")
async def get_document_ai_content(doc_id: str, user=Depends(get_current_user)):
    doc = await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"content": doc.get("ai_generated_content", ""), "title": doc.get("title", "")}


@api_router.put("/documents/{doc_id}/ai-content")
async def update_document_ai_content(doc_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    new_content = body.get("content", "")
    result = await db.documents.update_one(
        {"doc_id": doc_id},
        {"$set": {"ai_generated_content": new_content, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"ok": True}


# ---- ARL Indicators (Resolucion 0312/2019) ----

def _calc_indicator(numerator: float, denominator: float, multiplier: float = 1.0):
    if not denominator:
        return 0.0
    return round((numerator / denominator) * multiplier, 2)


@api_router.get("/indicators/arl")
async def get_arl_indicators(year: int, month: int = 0, user=Depends(get_current_user)):
    """Return ARL indicators for a company based on stored data.
    If month=0, returns the year aggregate. Otherwise the specific month.
    """
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}
    workers = max(1, int(company.get("workers_count", 1) or 1))
    # Estimated worked hours per period (8h x 22 days x months)
    months = 12 if month == 0 else 1
    horas_hombre = workers * 8 * 22 * months

    # Build date range
    if month == 0:
        date_start = f"{year}-01-01"
        date_end = f"{year}-12-31"
    else:
        date_start = f"{year}-{month:02d}-01"
        # Last day of month using calendar
        import calendar
        last = calendar.monthrange(year, month)[1]
        date_end = f"{year}-{month:02d}-{last:02d}"
    period_q = {"date": {"$gte": date_start, "$lte": date_end}}

    # --- Resultado: AT, EL, ausentismo ---
    incidents = await db.incidents.find({"company_id": cid, **period_q}, {"_id": 0}).to_list(1000)
    accidents = [i for i in incidents if i.get("incident_type") in ("accident", "minor_injury", "major_injury", "fatality")]
    at_count = len(accidents)
    fatalities = sum(1 for i in accidents if i.get("severity") == "fatal" or i.get("incident_type") == "fatality")
    days_lost = sum(int(i.get("days_lost", 0) or 0) for i in accidents)
    occupational_diseases = [i for i in incidents if i.get("incident_type") == "occupational_disease"]
    el_count = len(occupational_diseases)
    el_new = sum(1 for i in occupational_diseases if i.get("is_new", True))

    # --- Estructura: cumplimiento Res 0312, EPP, capacitaciones ---
    # Get the most recent SG-SST audit score for the company
    audits = await db.audits.find(
        {"company_id": cid, "audit_type": {"$in": ["internal", "external"]}},
        {"_id": 0, "score_result": 1, "created_at": 1},
    ).sort("created_at", -1).limit(1).to_list(1)
    res0312_score = (audits[0].get("score_result", {}) or {}).get("percentage", 0) if audits else 0

    trainings_total = await db.trainings.count_documents({"company_id": cid, "scheduled_date": {"$gte": date_start, "$lte": date_end}})
    trainings_done = await db.trainings.count_documents({"company_id": cid, "scheduled_date": {"$gte": date_start, "$lte": date_end}, "status": "completed"})
    pct_capacitaciones = _calc_indicator(trainings_done, trainings_total, 100)

    # --- Proceso: planes accion + inspecciones ---
    plans_total = await db.action_plans.count_documents({"company_id": cid})
    plans_closed = await db.action_plans.count_documents({"company_id": cid, "status": "closed"})
    pct_planes_cerrados = _calc_indicator(plans_closed, plans_total, 100)
    findings_total = await db.findings.count_documents({"company_id": cid})
    findings_closed = await db.findings.count_documents({"company_id": cid, "status": "closed"})
    pct_findings_cerrados = _calc_indicator(findings_closed, findings_total, 100)

    # --- Indicadores principales por fórmula Res. 1111/2017 ---
    frecuencia_at = _calc_indicator(at_count, horas_hombre, 200000)
    severidad_at = _calc_indicator(days_lost, horas_hombre, 200000)
    mortalidad_at = _calc_indicator(fatalities, workers, 100000)
    prevalencia_el = _calc_indicator(el_count, workers, 100000)
    incidencia_el = _calc_indicator(el_new, workers, 100000)
    # Ausentismo (eventos ausentismo registrados / dias laborales programados)
    days_scheduled = workers * 22 * months
    ausentismo_dias = sum(int(i.get("days_lost", 0) or 0) for i in incidents)
    ausentismo_pct = _calc_indicator(ausentismo_dias, days_scheduled, 100)

    return {
        "period": {"year": year, "month": month, "label": f"{year}" if month == 0 else f"{year}-{month:02d}"},
        "company": {"name": company.get("name", ""), "nit": company.get("nit", ""), "workers": workers, "risk_level": company.get("risk_level", 2)},
        "estructura": {
            "cumplimiento_resolucion_0312": res0312_score,
            "porcentaje_capacitaciones_ejecutadas": pct_capacitaciones,
            "trainings_planeadas": trainings_total,
            "trainings_realizadas": trainings_done,
        },
        "proceso": {
            "porcentaje_planes_cerrados": pct_planes_cerrados,
            "porcentaje_hallazgos_cerrados": pct_findings_cerrados,
            "planes_total": plans_total,
            "planes_cerrados": plans_closed,
            "hallazgos_total": findings_total,
            "hallazgos_cerrados": findings_closed,
        },
        "resultado": {
            "frecuencia_at": frecuencia_at,
            "severidad_at": severidad_at,
            "mortalidad_at": mortalidad_at,
            "prevalencia_el": prevalencia_el,
            "incidencia_el": incidencia_el,
            "ausentismo_porcentaje": ausentismo_pct,
            "accidentes_total": at_count,
            "muertes": fatalities,
            "dias_perdidos": days_lost,
            "enfermedades_laborales": el_count,
            "horas_hombre_trabajadas": horas_hombre,
        },
    }


@api_router.get("/indicators/arl/excel")
async def export_arl_indicators_excel(year: int, month: int = 0, user=Depends(get_current_user)):
    data = await get_arl_indicators(year, month, user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicadores ARL"
    bold = Font(bold=True, color="FFFFFFFF")
    fill = PatternFill("solid", fgColor="FF0047AB")
    sub_fill = PatternFill("solid", fgColor="FFE2E8F0")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = f"INDICADORES SST - {data['company']['name']}"
    ws["A1"].font = Font(size=14, bold=True, color="FF0047AB")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Periodo: {data['period']['label']}   |   NIT: {data['company']['nit']}   |   Trabajadores: {data['company']['workers']}"
    ws.merge_cells("A2:D2")

    row = 4
    for section_key, section_label in [("estructura", "INDICADORES DE ESTRUCTURA"), ("proceso", "INDICADORES DE PROCESO"), ("resultado", "INDICADORES DE RESULTADO")]:
        ws.cell(row=row, column=1, value=section_label).font = bold
        ws.cell(row=row, column=1).fill = fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value="Indicador").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Valor").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = sub_fill
        ws.cell(row=row, column=2).fill = sub_fill
        row += 1
        for k, v in data[section_key].items():
            ws.cell(row=row, column=1, value=k.replace("_", " ").title())
            ws.cell(row=row, column=2, value=v)
            ws.cell(row=row, column=2).alignment = center
            row += 1
        row += 1

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 32

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"Indicadores_ARL_{data['period']['label']}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={fname}"})


@api_router.get("/indicators/arl/pdf")
async def export_arl_indicators_pdf(year: int, month: int = 0, user=Depends(get_current_user)):
    data = await get_arl_indicators(year, month, user)
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}

    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0047AB"), spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#1F3C5E"), spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10)

    _logo = _company_logo_flowable(company, max_w=120, max_h=60)
    if _logo:
        _logo.hAlign = "LEFT"
        elements.append(_logo)
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#0047AB")))
    elements.append(Paragraph("REPORTE DE INDICADORES SST - ARL", title_st))
    elements.append(Paragraph(f"<b>Empresa:</b> {data['company']['name']} &nbsp;&nbsp; <b>NIT:</b> {data['company']['nit']}", body))
    elements.append(Paragraph(f"<b>Periodo:</b> {data['period']['label']} &nbsp;&nbsp; <b>Trabajadores:</b> {data['company']['workers']} &nbsp;&nbsp; <b>Nivel de riesgo:</b> {data['company']['risk_level']}", body))
    elements.append(Spacer(1, 16))

    def make_table(section_dict, title):
        elements.append(Paragraph(title, h2))
        rows = [["Indicador", "Valor"]]
        for k, v in section_dict.items():
            rows.append([Paragraph(k.replace("_", " ").title(), body), Paragraph(str(v), body)])
        t = Table(rows, colWidths=[3.5 * inch, 2.0 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0047AB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 16))

    make_table(data["estructura"], "Indicadores de Estructura")
    make_table(data["proceso"], "Indicadores de Proceso")
    make_table(data["resultado"], "Indicadores de Resultado")

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<i>Indicadores calculados conforme a la Resolucion 1111/2017 (deroga 0312/2019).</i>", body))

    doc.build(elements)
    bio.seek(0)
    fname = f"Indicadores_ARL_{data['period']['label']}.pdf"
    return StreamingResponse(bio, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})


# ---- Document Approval Workflow ----

def _record_approval(entity: dict, action: str, user: dict, comment: str = "") -> List[dict]:
    history = list(entity.get("approval_history", []))
    history.append({
        "action": action,
        "by": user.get("user_id", ""),
        "by_name": user.get("name", ""),
        "at": datetime.now(timezone.utc).isoformat(),
        "comment": comment or ""
    })
    return history

@api_router.post("/documents/{doc_id}/submit-approval")
async def submit_document_approval(doc_id: str, user=Depends(require_role("admin", "sgsst_manager"))):
    doc = await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.get("approval_status") == "pending":
        raise HTTPException(status_code=400, detail="Documento ya esta en revision")
    history = _record_approval(doc, "submitted", user)
    await db.documents.update_one({"doc_id": doc_id}, {"$set": {
        "approval_status": "pending",
        "approval_history": history,
        "submitted_by": user.get("name", ""),
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    # Notify approvers via email (non-blocking)
    company = await db.companies.find_one({"company_id": doc.get("company_id", "")}, {"_id": 0, "name": 1})
    cname = (company or {}).get("name", "N/A")
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto">
      <div style="background:#0047AB;padding:20px;color:#fff;border-radius:8px 8px 0 0">
        <h2 style="margin:0">TraciumSST - Documento pendiente de aprobacion</h2>
      </div>
      <div style="padding:20px;background:#F8F9FA;border:1px solid #E2E8F0;border-top:0;border-radius:0 0 8px 8px">
        <p>Un documento del SG-SST ha sido enviado para tu aprobacion.</p>
        <table style="width:100%;border-collapse:collapse;margin:12px 0">
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Empresa</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{cname}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Documento</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{doc.get('title','')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Categoria</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{doc.get('category','')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Version</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{doc.get('version','1.0')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Enviado por</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{user.get('name','')}</td></tr>
        </table>
        <p>Ingresa a <a href="{os.environ.get('FRONTEND_URL','')}/approvals" style="color:#0047AB;text-decoration:none;font-weight:bold">TraciumSST - Aprobaciones</a> para revisarlo.</p>
        <p style="color:#94A3B8;font-size:11px;margin-top:16px">Este es un mensaje automatico. No respondas.</p>
      </div>
    </div>
    """
    asyncio.create_task(_notify_approvers(doc.get("company_id", ""), f"Aprobacion pendiente: {doc.get('title','Documento')}", html, exclude_email=user.get("email", "")))
    return await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})

@api_router.post("/documents/{doc_id}/approve")
async def approve_document(doc_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json() if request.headers.get("content-length") and int(request.headers["content-length"]) > 0 else {}
    comment = (body or {}).get("comment", "")
    doc = await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Solo se aprueban documentos en revision")
    history = _record_approval(doc, "approved", user, comment)
    await db.documents.update_one({"doc_id": doc_id}, {"$set": {
        "approval_status": "approved",
        "approval_history": history,
        "approved_by": user.get("name", ""),
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    return await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})

@api_router.post("/documents/{doc_id}/reject")
async def reject_document(doc_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Motivo de rechazo es obligatorio")
    doc = await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Solo se rechazan documentos en revision")
    history = _record_approval(doc, "rejected", user, reason)
    await db.documents.update_one({"doc_id": doc_id}, {"$set": {
        "approval_status": "rejected",
        "approval_history": history,
        "rejection_reason": reason,
        "rejected_by": user.get("name", ""),
        "rejected_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    return await db.documents.find_one({"doc_id": doc_id}, {"_id": 0})

# ==================== HAZARDS (IPER) MODULE ====================

def calc_risk_category(level: int) -> str:
    if level >= 20:
        return "critical"
    elif level >= 12:
        return "high"
    elif level >= 6:
        return "medium"
    return "low"

@api_router.get("/hazards", response_model=List[HazardOut])
async def get_hazards(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.hazards.find({"company_id": cid}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/hazards", response_model=HazardOut)
async def create_hazard(h: HazardCreate, user=Depends(get_current_user)):
    d = h.model_dump()
    d["hazard_id"] = f"hzd_{uuid.uuid4().hex[:8]}"
    d["risk_level"] = d["probability"] * d["severity"]
    d["risk_category"] = calc_risk_category(d["risk_level"])
    d["status"] = "active"
    d["company_id"] = get_company_id(user)
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hazards.insert_one(d)
    return HazardOut(**d)

@api_router.put("/hazards/{hazard_id}")
async def update_hazard(hazard_id: str, updates: Dict[str, Any], user=Depends(get_current_user)):
    updates.pop("_id", None)
    updates.pop("hazard_id", None)
    if "probability" in updates and "severity" in updates:
        updates["risk_level"] = updates["probability"] * updates["severity"]
        updates["risk_category"] = calc_risk_category(updates["risk_level"])
    result = await db.hazards.update_one({"hazard_id": hazard_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Hazard not found")
    updated = await db.hazards.find_one({"hazard_id": hazard_id}, {"_id": 0})
    return updated

@api_router.delete("/hazards/{hazard_id}")
async def delete_hazard(hazard_id: str, user=Depends(get_current_user)):
    result = await db.hazards.delete_one({"hazard_id": hazard_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Hazard not found")
    return {"message": "Deleted"}

# ==================== INCIDENTS MODULE ====================

@api_router.get("/incidents", response_model=List[IncidentOut])
async def get_incidents(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.incidents.find({"company_id": cid}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/incidents", response_model=IncidentOut)
async def create_incident(inc: IncidentCreate, user=Depends(get_current_user)):
    if user.get("role", "collaborator") not in INCIDENT_REPORT_ROLES and not is_owner(user):
        raise HTTPException(status_code=403, detail="Rol sin permiso para registrar incidentes")
    d = inc.model_dump()
    d["incident_id"] = f"inc_{uuid.uuid4().hex[:8]}"
    d["root_cause"] = ""
    d["corrective_actions"] = ""
    d["status"] = "open"
    d["created_by"] = user.get("name", "")
    d["company_id"] = get_company_id(user)
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.incidents.insert_one(d)
    return IncidentOut(**d)

@api_router.put("/incidents/{incident_id}")
async def update_incident(incident_id: str, updates: Dict[str, Any], user=Depends(get_current_user)):
    updates.pop("_id", None)
    updates.pop("incident_id", None)
    result = await db.incidents.update_one({"incident_id": incident_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incident not found")
    updated = await db.incidents.find_one({"incident_id": incident_id}, {"_id": 0})
    return updated

@api_router.delete("/incidents/{incident_id}")
async def delete_incident(incident_id: str, user=Depends(get_current_user)):
    result = await db.incidents.delete_one({"incident_id": incident_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Incident not found")
    return {"message": "Deleted"}

# ==================== TRAINING MODULE ====================

@api_router.get("/trainings", response_model=List[TrainingOut])
async def get_trainings(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.trainings.find({"company_id": cid}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/trainings", response_model=TrainingOut)
async def create_training(t: TrainingCreate, user=Depends(get_current_user)):
    d = t.model_dump()
    d["training_id"] = f"trn_{uuid.uuid4().hex[:8]}"
    d["attendees"] = []
    d["status"] = "scheduled"
    d["effectiveness_score"] = None
    d["company_id"] = get_company_id(user)
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.trainings.insert_one(d)
    return TrainingOut(**d)

@api_router.put("/trainings/{training_id}")
async def update_training(training_id: str, updates: Dict[str, Any], user=Depends(get_current_user)):
    updates.pop("_id", None)
    updates.pop("training_id", None)
    result = await db.trainings.update_one({"training_id": training_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Training not found")
    updated = await db.trainings.find_one({"training_id": training_id}, {"_id": 0})
    return updated

@api_router.delete("/trainings/{training_id}")
async def delete_training(training_id: str, user=Depends(get_current_user)):
    result = await db.trainings.delete_one({"training_id": training_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Training not found")
    return {"message": "Deleted"}

# ==================== AUDIT MODULE (COMPLETE FLOW) ====================

@api_router.get("/audits")
async def get_audits(user=Depends(get_current_user)):
    cid = get_company_id(user)
    items = await db.audits.find({"company_id": cid}, {"_id": 0}).to_list(500)
    for a in items:
        a["findings_count"] = await db.findings.count_documents({"audit_id": a["audit_id"], "status": {"$ne": "resolved_by_compliance"}})
        a["checklist_count"] = await db.audit_checklist.count_documents({"audit_id": a["audit_id"]})
        a["checklist_completed"] = await db.audit_checklist.count_documents({"audit_id": a["audit_id"], "result": {"$in": ["cumple", "no_cumple", "parcial", "no_aplica"]}})
        a["action_plans_count"] = await db.action_plans.count_documents({"audit_id": a["audit_id"]})
        a["action_plans_closed"] = await db.action_plans.count_documents({"audit_id": a["audit_id"], "status": "closed"})
    return items

@api_router.post("/audits")
async def create_audit(request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    cid = get_company_id(user)
    d = {
        "audit_id": f"aud_{uuid.uuid4().hex[:8]}",
        "title": body.get("title", ""),
        "audit_type": body.get("audit_type", "internal"),
        "pesv_level": body.get("pesv_level", "avanzado"),
        "scheduled_date": body.get("scheduled_date", ""),
        "start_time": body.get("start_time", ""),
        "end_date": body.get("end_date", ""),
        "end_time": body.get("end_time", ""),
        "auditor": body.get("auditor", ""),
        "auditor_id": body.get("auditor_id", ""),
        "additional_auditors": body.get("additional_auditors", []),
        "process_responsibles": body.get("process_responsibles", []),
        "copasst_member": body.get("copasst_member", {}),
        "scope": body.get("scope", ""),
        "criteria": body.get("criteria", "Resolucion 0312 de 2019, Decreto 1072 de 2015"),
        "objective": body.get("objective", ""),
        "status": "planned",
        "findings_count": 0,
        "company_id": cid,
        "created_by": user.get("name", ""),
        "management_review": None,
        "executive_summary": "",
        "ai_redacted_summary": "",
        "ai_redacted_findings": "",
        "ai_redacted_recommendations": "",
        "ai_redacted_conclusions": "",
        "score_result": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.audits.insert_one(d)
    d.pop("_id", None)
    return d

@api_router.get("/audits/{audit_id}")
async def get_audit_detail(audit_id: str, user=Depends(get_current_user)):
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    audit["findings"] = await db.findings.find({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}}, {"_id": 0}).to_list(100)
    audit["checklist"] = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    audit["action_plans"] = await db.action_plans.find({"audit_id": audit_id, "status": {"$ne": "closed"}}, {"_id": 0}).to_list(100)
    audit["closed_plans"] = await db.action_plans.count_documents({"audit_id": audit_id, "status": "closed"})
    audit["findings_count"] = len(audit["findings"])
    checklist = audit["checklist"]
    audit["checklist_count"] = len(checklist)
    audit["checklist_completed"] = sum(1 for c in checklist if c.get("result") in ("cumple", "no_cumple", "parcial", "no_aplica"))
    # Calculate score (SG-SST or PESV)
    if checklist:
        if audit.get("audit_type") == "pesv":
            audit["score_result"] = calculate_pesv_score(checklist)
        else:
            audit["score_result"] = calculate_score(checklist)
    # Preserve cascade flags from DB
    db_audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0, "report_stale": 1, "last_execution_change": 1, "change_history": 1})
    if db_audit:
        audit["report_stale"] = db_audit.get("report_stale", False)
        audit["last_execution_change"] = db_audit.get("last_execution_change", "")
        audit["change_history"] = db_audit.get("change_history", [])
    return audit

@api_router.put("/audits/{audit_id}")
async def update_audit(audit_id: str, request: Request, user=Depends(require_role("admin", "auditor"))):
    body = await request.json()
    body.pop("_id", None)
    body.pop("audit_id", None)
    new_status = body.get("status")

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")

    # Block editing if audit is closed (except reopening by admin)
    if audit.get("status") in ("closed", "reviewed") and new_status not in ("in_progress", "follow_up"):
        raise HTTPException(status_code=400, detail="Auditoria cerrada. No se puede editar. Reabra la auditoria para modificarla.")

    # Business rules for closing
    if new_status in ("closed", "reviewed"):
        end_date = body.get("end_date", audit.get("end_date", ""))
        end_time = body.get("end_time", audit.get("end_time", ""))
        auditor = body.get("auditor", audit.get("auditor", ""))
        copasst = body.get("copasst_member", audit.get("copasst_member", {}))
        process_resp = body.get("process_responsibles", audit.get("process_responsibles", []))
        errors = []
        if not auditor:
            errors.append("Auditor lider es obligatorio")
        if not end_date:
            errors.append("Fecha de cierre es obligatoria")
        if not end_time:
            errors.append("Hora de cierre es obligatoria")
        if not copasst or not copasst.get("name"):
            errors.append("Miembro COPASST es obligatorio")
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))

    # Track programming changes for traceability
    tracked_fields = ["scheduled_date", "start_time", "end_date", "end_time", "auditor",
                      "additional_auditors", "process_responsibles", "copasst_member",
                      "scope", "criteria", "objective", "title"]
    changes = []
    for field in tracked_fields:
        if field in body and str(body[field]) != str(audit.get(field, "")):
            changes.append({"field": field, "old": audit.get(field, ""), "new": body[field],
                          "by": user.get("name", ""), "at": datetime.now(timezone.utc).isoformat()})
    if changes:
        body.setdefault("change_history", audit.get("change_history", []))
        body["change_history"] = audit.get("change_history", []) + changes

    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.audits.update_one({"audit_id": audit_id}, {"$set": body})
    updated = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    return updated


@api_router.get("/audits/{audit_id}/score")
async def get_audit_score(audit_id: str, user=Depends(get_current_user)):
    """Calculate compliance score for an audit (SG-SST or PESV)"""
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    checklist = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    if not checklist:
        return {"message": "No hay checklist generado", "percentage": 0}
    if audit and audit.get("audit_type") == "pesv":
        score = calculate_pesv_score(checklist)
    else:
        score = calculate_score(checklist)
    await db.audits.update_one({"audit_id": audit_id}, {"$set": {"score_result": score}})
    return score

@api_router.delete("/audits/{audit_id}")
async def delete_audit(audit_id: str, user=Depends(require_role("admin"))):
    result = await db.audits.delete_one({"audit_id": audit_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    await db.findings.delete_many({"audit_id": audit_id})
    await db.audit_checklist.delete_many({"audit_id": audit_id})
    await db.action_plans.delete_many({"audit_id": audit_id})
    return {"message": "Eliminada"}

# ---- Audit Closure Approval Workflow ----

@api_router.post("/audits/{audit_id}/submit-closure")
async def submit_audit_closure(audit_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager", "auditor"))):
    body = await request.json()
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    if audit.get("status") in ("closed", "reviewed"):
        raise HTTPException(status_code=400, detail="La auditoria ya esta cerrada")
    if audit.get("closure_approval_status") == "pending":
        raise HTTPException(status_code=400, detail="Cierre ya enviado a aprobacion")

    # Validate closure data
    errors = []
    end_date = body.get("end_date", audit.get("end_date", ""))
    end_time = body.get("end_time", audit.get("end_time", ""))
    auditor = audit.get("auditor", "")
    copasst = audit.get("copasst_member", {})
    if not auditor: errors.append("Auditor lider requerido")
    if not end_date: errors.append("Fecha de cierre requerida")
    if not end_time: errors.append("Hora de cierre requerida")
    if not copasst or not copasst.get("name"): errors.append("COPASST requerido")
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))

    history = _record_approval(audit, "closure_submitted", user, body.get("comment", ""))
    await db.audits.update_one({"audit_id": audit_id}, {"$set": {
        "closure_approval_status": "pending",
        "closure_approval_history": history,
        "closure_submitted_by": user.get("name", ""),
        "closure_submitted_at": datetime.now(timezone.utc).isoformat(),
        "pending_closure_data": {"end_date": end_date, "end_time": end_time},
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    # Notify approvers via email
    company = await db.companies.find_one({"company_id": audit.get("company_id", "")}, {"_id": 0, "name": 1})
    cname = (company or {}).get("name", "N/A")
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto">
      <div style="background:#1F3C5E;padding:20px;color:#fff;border-radius:8px 8px 0 0">
        <h2 style="margin:0">TraciumSST - Cierre de Auditoria pendiente</h2>
      </div>
      <div style="padding:20px;background:#F8F9FA;border:1px solid #E2E8F0;border-top:0;border-radius:0 0 8px 8px">
        <p>Una auditoria ha sido enviada para aprobacion de cierre.</p>
        <table style="width:100%;border-collapse:collapse;margin:12px 0">
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Empresa</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{cname}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Auditoria</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('title','')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Tipo</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('audit_type','')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Auditor</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{auditor}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Fecha y Hora de Cierre</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{end_date} {end_time}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Enviado por</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{user.get('name','')}</td></tr>
        </table>
        <p>Ingresa a <a href="{os.environ.get('FRONTEND_URL','')}/approvals" style="color:#0047AB;text-decoration:none;font-weight:bold">TraciumSST - Aprobaciones</a> para revisarlo.</p>
        <p style="color:#94A3B8;font-size:11px;margin-top:16px">Este es un mensaje automatico. No respondas.</p>
      </div>
    </div>
    """
    asyncio.create_task(_notify_approvers(audit.get("company_id", ""), f"Cierre pendiente: {audit.get('title','Auditoria')}", html, exclude_email=user.get("email", "")))
    return await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})

@api_router.post("/audits/{audit_id}/approve-closure")
async def approve_audit_closure(audit_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = {}
    try: body = await request.json()
    except Exception: pass
    comment = (body or {}).get("comment", "")
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    if audit.get("closure_approval_status") != "pending":
        raise HTTPException(status_code=400, detail="El cierre no esta en revision")

    pending = audit.get("pending_closure_data", {})
    history = _record_approval(audit, "closure_approved", user, comment)
    update = {
        "status": "closed",
        "closure_approval_status": "approved",
        "closure_approval_history": history,
        "closure_approved_by": user.get("name", ""),
        "closure_approved_at": datetime.now(timezone.utc).isoformat(),
        "end_date": pending.get("end_date", audit.get("end_date", "")),
        "end_time": pending.get("end_time", audit.get("end_time", "")),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.audits.update_one({"audit_id": audit_id}, {"$set": update, "$unset": {"pending_closure_data": ""}})
    return await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})

@api_router.post("/audits/{audit_id}/reject-closure")
async def reject_audit_closure(audit_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Motivo de rechazo es obligatorio")
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    if audit.get("closure_approval_status") != "pending":
        raise HTTPException(status_code=400, detail="El cierre no esta en revision")
    history = _record_approval(audit, "closure_rejected", user, reason)
    await db.audits.update_one({"audit_id": audit_id}, {"$set": {
        "closure_approval_status": "rejected",
        "closure_approval_history": history,
        "closure_rejection_reason": reason,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }, "$unset": {"pending_closure_data": ""}})
    return await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})

@api_router.get("/approvals/pending")
async def list_pending_approvals(user=Depends(require_role("admin", "sgsst_manager"))):
    """List all items pending approval for the current user's scope."""
    cid = get_company_id(user)
    company_filter = {"company_id": cid} if not is_owner_or_admin(user) else {}
    docs = await db.documents.find({**company_filter, "approval_status": "pending"}, {"_id": 0}).to_list(200)
    audits = await db.audits.find({**company_filter, "closure_approval_status": "pending"}, {"_id": 0}).to_list(200)
    return {
        "documents": [{
            "doc_id": d.get("doc_id"), "title": d.get("title"),
            "category": d.get("category"), "version": d.get("version"),
            "submitted_by": d.get("submitted_by", ""),
            "submitted_at": d.get("submitted_at", ""),
        } for d in docs],
        "audits": [{
            "audit_id": a.get("audit_id"), "title": a.get("title"),
            "audit_type": a.get("audit_type"),
            "submitted_by": a.get("closure_submitted_by", ""),
            "submitted_at": a.get("closure_submitted_at", ""),
        } for a in audits],
        "total": len(docs) + len(audits),
    }

# Helper for RBAC
def is_owner_or_admin(user: dict) -> bool:
    return user.get("role") in ("owner", "admin")

# Audit Checklist
@api_router.post("/audits/{audit_id}/checklist/generate")
async def generate_audit_checklist(audit_id: str, user=Depends(require_role("admin", "auditor"))):
    """Generate checklist from applicable standards (SG-SST or PESV)"""
    cid = get_company_id(user)
    existing = await db.audit_checklist.count_documents({"audit_id": audit_id})
    if existing > 0:
        return {"message": f"Checklist ya existe ({existing} items)"}

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    audit_type = audit.get("audit_type", "internal") if audit else "internal"

    if audit_type == "pesv":
        # PESV checklist
        pesv_level = audit.get("pesv_level", "avanzado") if audit else "avanzado"
        standards = get_pesv_standards(pesv_level)
        docs = []
        for s in standards:
            docs.append({
                "item_id": f"aci_{uuid.uuid4().hex[:8]}",
                "audit_id": audit_id,
                "code": s["code"],
                "standard": s.get("paso", ""),
                "description": s.get("description", ""),
                "evidence_required": s.get("evidence", ""),
                "criterio": s.get("description", ""),
                "modo_verificacion": s.get("evidence", ""),
                "phva": s.get("fase", ""),
                "fase": s.get("fase", ""),
                "paso": s.get("paso", ""),
                "paso_num": s.get("paso_num", ""),
                "weight": s.get("weight", 1),
                "checked": False,
                "result": "",
                "observations": "",
                "evidence_files": [],
                "audit_framework": "pesv",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    else:
        # SG-SST checklist (existing logic)
        std_count = await db.standards_compliance.count_documents({"company_id": cid})
        if std_count == 0:
            await seed_standards(user=user)
        standards = await db.standards_compliance.find({"company_id": cid, "applicable": True}, {"_id": 0}).to_list(200)
        if not standards:
            company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
            workers = company.get("workers_count", 25) if company else 25
            risk = company.get("risk_level", 2) if company else 2
            applicable_codes = get_applicable_standards(workers, risk)
            standards = [s for s in STANDARDS_BANK if s["code"] in applicable_codes]
        docs = []
        for s in standards:
            code = s.get("code", "")
            cv = CRITERIA_VERIFICATION.get(code, {})
            docs.append({
                "item_id": f"aci_{uuid.uuid4().hex[:8]}",
                "audit_id": audit_id,
                "code": code,
                "standard": s.get("standard", ""),
                "description": s.get("description", ""),
                "evidence_required": s.get("evidence", s.get("evidence_required", "")),
                "criterio": cv.get("criterio", ""),
                "modo_verificacion": cv.get("modo_verificacion", ""),
                "phva": s.get("phva", ""),
                "weight": s.get("weight", 0),
                "checked": False,
                "result": "",
                "observations": "",
                "evidence_files": [],
                "audit_framework": "sgsst",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    if docs:
        await db.audit_checklist.insert_many(docs)
    framework = "PESV" if audit_type == "pesv" else "SG-SST"
    return {"message": f"Checklist {framework} generado con {len(docs)} items"}

@api_router.get("/audits/{audit_id}/checklist")
async def get_audit_checklist(audit_id: str, user=Depends(get_current_user)):
    items = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    # Enrich items with criteria if not present
    enriched = False
    for item in items:
        if not item.get("criterio") and item.get("code"):
            cv = CRITERIA_VERIFICATION.get(item["code"], {})
            if cv:
                item["criterio"] = cv.get("criterio", "")
                item["modo_verificacion"] = cv.get("modo_verificacion", "")
                await db.audit_checklist.update_one(
                    {"item_id": item["item_id"]},
                    {"$set": {"criterio": item["criterio"], "modo_verificacion": item["modo_verificacion"]}}
                )
                enriched = True
    return items

@api_router.put("/audits/{audit_id}/checklist/{item_id}")
async def update_checklist_item(audit_id: str, item_id: str, request: Request, user=Depends(require_role("admin", "auditor"))):
    body = await request.json()
    body.pop("_id", None)
    body.pop("item_id", None)
    old_item = await db.audit_checklist.find_one({"item_id": item_id, "audit_id": audit_id}, {"_id": 0})
    if not old_item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    old_result = old_item.get("result", "")
    new_result = body.get("result", old_result)
    # Auto-set checked when result is set
    if "result" in body and body["result"]:
        body["checked"] = True
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.audit_checklist.update_one({"item_id": item_id, "audit_id": audit_id}, {"$set": body})
    updated = await db.audit_checklist.find_one({"item_id": item_id}, {"_id": 0})

    # --- FASE 3: Dynamic finding sync ---
    if old_result != new_result:
        cid = get_company_id(user)
        existing = await db.findings.find_one({"audit_id": audit_id, "source_item_id": item_id}, {"_id": 0})
        if new_result in ("no_cumple", "parcial"):
            finding_type = "no_conformity" if new_result == "no_cumple" else "observation"
            desc = f"Estandar {updated.get('code', '')}: {updated.get('description', '')}."
            if body.get("observations") or updated.get("observations"):
                desc += f" Observacion: {body.get('observations', updated.get('observations', ''))}"
            if existing:
                await db.findings.update_one(
                    {"finding_id": existing["finding_id"]},
                    {"$set": {"finding_type": finding_type, "description": desc, "status": "open", "updated_at": datetime.now(timezone.utc).isoformat()},
                     "$push": {"change_log": {"from": old_result, "to": new_result, "by": user.get("name", ""), "at": datetime.now(timezone.utc).isoformat()}}}
                )
                # Check if action plan exists, create if not
                existing_plan = await db.action_plans.find_one({"finding_id": existing["finding_id"], "audit_id": audit_id}, {"_id": 0})
                if existing_plan:
                    # Update linked action plan
                    await db.action_plans.update_many(
                        {"finding_id": existing["finding_id"], "audit_id": audit_id},
                        {"$set": {"status": "open", "updated_at": datetime.now(timezone.utc).isoformat(),
                                  "action": f"{'Accion correctiva' if finding_type == 'no_conformity' else 'Accion preventiva'} para: {desc[:150]}"}}
                    )
                else:
                    # Create action plan if missing
                    action_type = "corrective" if finding_type == "no_conformity" else "preventive"
                    plan = {
                        "plan_id": f"ap_{uuid.uuid4().hex[:8]}",
                        "audit_id": audit_id, "finding_id": existing["finding_id"],
                        "action": f"{'Accion correctiva' if action_type == 'corrective' else 'Accion preventiva'} para: {desc[:150]}",
                        "action_type": action_type,
                        "responsible": "", "due_date": "", "status": "open", "progress": 0,
                        "follow_up_notes": [], "company_id": cid,
                        "source_item_id": item_id,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.action_plans.insert_one(plan)
                    plan.pop("_id", None)
            else:
                fid = f"fnd_{uuid.uuid4().hex[:8]}"
                new_finding = {
                    "finding_id": fid,
                    "audit_id": audit_id, "finding_type": finding_type, "description": desc,
                    "area": updated.get("phva", "General"), "standard_ref": updated.get("code", ""),
                    "corrective_action": "", "responsible": "", "due_date": "", "evidence_files": [],
                    "status": "open", "company_id": cid, "source_item_id": item_id,
                    "created_by": user.get("name", ""), "change_log": [],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.findings.insert_one(new_finding)
                new_finding.pop("_id", None)
                # Auto-create action plan for this finding
                action_type = "corrective" if finding_type == "no_conformity" else "preventive"
                plan = {
                    "plan_id": f"ap_{uuid.uuid4().hex[:8]}",
                    "audit_id": audit_id, "finding_id": fid,
                    "action": f"{'Accion correctiva' if action_type == 'corrective' else 'Accion preventiva'} para: {desc[:150]}",
                    "action_type": action_type,
                    "responsible": "", "due_date": "", "status": "open", "progress": 0,
                    "follow_up_notes": [], "company_id": cid,
                    "source_item_id": item_id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.action_plans.insert_one(plan)
                plan.pop("_id", None)
        elif new_result in ("cumple", "no_aplica") and existing:
            await db.findings.update_one(
                {"finding_id": existing["finding_id"]},
                {"$set": {"status": "resolved_by_compliance", "updated_at": datetime.now(timezone.utc).isoformat()},
                 "$push": {"change_log": {"from": old_result, "to": new_result, "by": user.get("name", ""), "at": datetime.now(timezone.utc).isoformat()}}}
            )
            # Auto-close linked action plans
            await db.action_plans.update_many(
                {"finding_id": existing["finding_id"], "audit_id": audit_id},
                {"$set": {"status": "closed", "progress": 100, "updated_at": datetime.now(timezone.utc).isoformat(),
                          "closure_note": f"Cerrado automaticamente: item cambio a {new_result}"}}
            )

        # --- CASCADE: Recalculate score, update audit summary, mark stale ---
        all_checklist = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
        audit_doc = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
        if audit_doc and audit_doc.get("audit_type") == "pesv":
            new_score = calculate_pesv_score(all_checklist)
        else:
            new_score = calculate_score(all_checklist)
        all_findings = await db.findings.find({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}}, {"_id": 0}).to_list(200)
        all_plans = await db.action_plans.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
        await db.audits.update_one(
            {"audit_id": audit_id},
            {"$set": {
                "score_result": new_score,
                "findings_count": len(all_findings),
                "report_stale": True,
                "last_execution_change": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    return updated


@api_router.post("/audits/{audit_id}/findings/generate-from-checklist")
async def generate_findings_from_checklist(audit_id: str, user=Depends(require_role("admin", "auditor"))):
    """Auto-generate/update findings AND action plans from checklist execution"""
    cid = get_company_id(user)
    all_checklist = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    nc_items = [c for c in all_checklist if c.get("result") in ("no_cumple", "parcial")]
    resolved_items = [c for c in all_checklist if c.get("result") in ("cumple", "no_aplica")]

    existing_findings = await db.findings.find(
        {"audit_id": audit_id, "source_item_id": {"$exists": True}}, {"_id": 0}
    ).to_list(500)
    existing_map = {f["source_item_id"]: f for f in existing_findings}

    created_findings = 0
    updated_findings = 0
    created_plans = 0
    resolved_findings = 0

    # Process non-compliant items: create or update findings + action plans
    for item in nc_items:
        finding_type = "no_conformity" if item["result"] == "no_cumple" else "observation"
        desc = f"Estandar {item.get('code', '')}: {item.get('description', '')}."
        if item.get("observations"):
            desc += f" Observacion: {item['observations']}"
        existing = existing_map.get(item["item_id"])
        if existing:
            if existing.get("description") != desc or existing.get("finding_type") != finding_type or existing.get("status") == "resolved_by_compliance":
                await db.findings.update_one(
                    {"finding_id": existing["finding_id"]},
                    {"$set": {"finding_type": finding_type, "description": desc, "status": "open",
                              "updated_at": datetime.now(timezone.utc).isoformat()},
                     "$push": {"change_log": {"action": "updated_from_execution", "by": user.get("name", ""),
                               "at": datetime.now(timezone.utc).isoformat()}}}
                )
                updated_findings += 1
            # Check if action plan exists for this finding, create if not
            existing_plan = await db.action_plans.find_one({"finding_id": existing["finding_id"], "audit_id": audit_id}, {"_id": 0})
            if existing_plan:
                # Update linked action plans
                await db.action_plans.update_many(
                    {"finding_id": existing["finding_id"], "audit_id": audit_id},
                    {"$set": {"status": "open", "updated_at": datetime.now(timezone.utc).isoformat(),
                              "action": f"{'Accion correctiva' if finding_type == 'no_conformity' else 'Accion preventiva'}: {desc[:150]}"}}
                )
            else:
                # Create action plan if missing
                action_type = "corrective" if finding_type == "no_conformity" else "preventive"
                plan = {
                    "plan_id": f"ap_{uuid.uuid4().hex[:8]}",
                    "audit_id": audit_id, "finding_id": existing["finding_id"],
                    "action": f"{'Accion correctiva' if action_type == 'corrective' else 'Accion preventiva'}: {desc[:150]}",
                    "action_type": action_type,
                    "responsible": "", "due_date": "", "status": "open", "progress": 0,
                    "follow_up_notes": [], "company_id": cid,
                    "source_item_id": item["item_id"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.action_plans.insert_one(plan)
                plan.pop("_id", None)
                created_plans += 1
        else:
            fid = f"fnd_{uuid.uuid4().hex[:8]}"
            new_f = {
                "finding_id": fid,
                "audit_id": audit_id, "finding_type": finding_type, "description": desc,
                "area": item.get("phva", "General"), "standard_ref": item.get("code", ""),
                "corrective_action": "", "responsible": "", "due_date": "", "evidence_files": [],
                "status": "open", "company_id": cid, "source_item_id": item["item_id"],
                "created_by": user.get("name", ""), "change_log": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.findings.insert_one(new_f)
            new_f.pop("_id", None)
            created_findings += 1
            # Auto-create action plan
            action_type = "corrective" if finding_type == "no_conformity" else "preventive"
            plan = {
                "plan_id": f"ap_{uuid.uuid4().hex[:8]}",
                "audit_id": audit_id, "finding_id": fid,
                "action": f"{'Accion correctiva' if action_type == 'corrective' else 'Accion preventiva'}: {desc[:150]}",
                "action_type": action_type,
                "responsible": "", "due_date": "", "status": "open", "progress": 0,
                "follow_up_notes": [], "company_id": cid,
                "source_item_id": item["item_id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            existing_plan = await db.action_plans.find_one({"audit_id": audit_id, "finding_id": fid}, {"_id": 0})
            if not existing_plan:
                await db.action_plans.insert_one(plan)
                plan.pop("_id", None)
                created_plans += 1

    # Resolve findings for items that now comply
    for item in resolved_items:
        existing = existing_map.get(item["item_id"])
        if existing and existing.get("status") != "resolved_by_compliance":
            await db.findings.update_one(
                {"finding_id": existing["finding_id"]},
                {"$set": {"status": "resolved_by_compliance", "updated_at": datetime.now(timezone.utc).isoformat()},
                 "$push": {"change_log": {"action": "resolved_by_execution_change", "by": user.get("name", ""),
                           "at": datetime.now(timezone.utc).isoformat()}}}
            )
            await db.action_plans.update_many(
                {"finding_id": existing["finding_id"], "audit_id": audit_id},
                {"$set": {"status": "closed", "progress": 100, "updated_at": datetime.now(timezone.utc).isoformat(),
                          "closure_note": f"Cerrado: item cambio a {item.get('result')}"}}
            )
            resolved_findings += 1

    # Recalculate score and clear stale
    new_score = calculate_score(all_checklist)
    active_findings = await db.findings.count_documents({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}})
    await db.audits.update_one(
        {"audit_id": audit_id},
        {"$set": {"score_result": new_score, "findings_count": active_findings, "report_stale": True,
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    msg = f"Hallazgos: {created_findings} nuevos, {updated_findings} actualizados, {resolved_findings} resueltos. Planes de accion: {created_plans} creados."
    return {"message": msg, "created": created_findings, "updated": updated_findings, "resolved": resolved_findings, "plans_created": created_plans}


# Findings (enhanced)
async def _recount_audit_findings(audit_id: str):
    """Resync audit.findings_count with the actual DB count (active findings only)."""
    if not audit_id:
        return
    count = await db.findings.count_documents({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}})
    await db.audits.update_one({"audit_id": audit_id}, {"$set": {"findings_count": count}})


@api_router.get("/findings")
async def get_findings(audit_id: Optional[str] = None, user=Depends(get_current_user)):
    cid = get_company_id(user)
    query = {"company_id": cid}
    if audit_id:
        query["audit_id"] = audit_id
    items = await db.findings.find(query, {"_id": 0}).to_list(500)
    return items

@api_router.post("/findings")
async def create_finding(request: Request, user=Depends(require_role("admin", "auditor"))):
    body = await request.json()
    d = {
        "finding_id": f"fnd_{uuid.uuid4().hex[:8]}",
        "audit_id": body.get("audit_id", ""),
        "finding_type": body.get("finding_type", "no_conformity"),
        "description": body.get("description", ""),
        "area": body.get("area", ""),
        "standard_ref": body.get("standard_ref", ""),
        "corrective_action": body.get("corrective_action", ""),
        "responsible": body.get("responsible", ""),
        "due_date": body.get("due_date", ""),
        "evidence_files": body.get("evidence_files", []),
        "status": "open",
        "company_id": get_company_id(user),
        "created_by": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.findings.insert_one(d)
    d.pop("_id", None)
    await _recount_audit_findings(d["audit_id"])
    return d

@api_router.put("/findings/{finding_id}")
async def update_finding(finding_id: str, request: Request, user=Depends(require_role("admin", "auditor"))):
    body = await request.json()
    body.pop("_id", None)
    body.pop("finding_id", None)
    existing = await db.findings.find_one({"finding_id": finding_id}, {"_id": 0, "audit_id": 1})
    result = await db.findings.update_one({"finding_id": finding_id}, {"$set": body})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Hallazgo no encontrado")
    updated = await db.findings.find_one({"finding_id": finding_id}, {"_id": 0})
    if existing:
        await _recount_audit_findings(existing.get("audit_id", ""))
    return updated

@api_router.delete("/findings/{finding_id}")
async def delete_finding(finding_id: str, user=Depends(require_role("admin", "sgsst_manager"))):
    existing = await db.findings.find_one({"finding_id": finding_id}, {"_id": 0, "audit_id": 1})
    result = await db.findings.delete_one({"finding_id": finding_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Hallazgo no encontrado")
    if existing:
        await _recount_audit_findings(existing.get("audit_id", ""))
    return {"message": "Eliminado"}

# Action Plans
@api_router.get("/action-plans")
async def get_action_plans(audit_id: Optional[str] = None, user=Depends(get_current_user)):
    cid = get_company_id(user)
    query = {"company_id": cid}
    if audit_id:
        query["audit_id"] = audit_id
    items = await db.action_plans.find(query, {"_id": 0}).to_list(500)
    return items

@api_router.post("/action-plans")
async def create_action_plan(request: Request, user=Depends(require_role("admin", "sgsst_manager", "auditor"))):
    body = await request.json()
    d = {
        "plan_id": f"ap_{uuid.uuid4().hex[:8]}",
        "audit_id": body.get("audit_id", ""),
        "finding_id": body.get("finding_id", ""),
        "action": body.get("action", ""),
        "action_type": body.get("action_type", "corrective"),
        "responsible": body.get("responsible", ""),
        "start_date": body.get("start_date", ""),
        "due_date": body.get("due_date", ""),
        "resources": body.get("resources", ""),
        "evidence": body.get("evidence", ""),
        "status": "open",
        "progress": 0,
        "follow_up_notes": [],
        "company_id": get_company_id(user),
        "created_by": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.action_plans.insert_one(d)
    d.pop("_id", None)
    # Notify (async, non-blocking)
    audit = await db.audits.find_one({"audit_id": d["audit_id"]}, {"_id": 0}) or {}
    asyncio.create_task(_notify_action_plan_change(d, audit, "created", user))
    return d

@api_router.put("/action-plans/{plan_id}")
async def update_action_plan(plan_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager", "auditor"))):
    body = await request.json()
    body.pop("_id", None)
    body.pop("plan_id", None)
    body["updated_by"] = user.get("name", "")
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.action_plans.update_one({"plan_id": plan_id}, {"$set": body})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    updated = await db.action_plans.find_one({"plan_id": plan_id}, {"_id": 0})
    audit = await db.audits.find_one({"audit_id": updated.get("audit_id", "")}, {"_id": 0}) or {}
    change_type = "closed" if body.get("status") == "closed" else "updated"
    asyncio.create_task(_notify_action_plan_change(updated, audit, change_type, user))
    return updated

@api_router.post("/action-plans/{plan_id}/follow-up")
async def add_follow_up(plan_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager", "auditor"))):
    body = await request.json()
    note_text = body.get("note", "")
    note = {
        "note": note_text,
        "by": user.get("name", ""),
        "date": datetime.now(timezone.utc).isoformat()
    }
    await db.action_plans.update_one({"plan_id": plan_id}, {"$push": {"follow_up_notes": note}})
    updated = await db.action_plans.find_one({"plan_id": plan_id}, {"_id": 0})
    if updated:
        audit = await db.audits.find_one({"audit_id": updated.get("audit_id", "")}, {"_id": 0}) or {}
        asyncio.create_task(_notify_action_plan_change(updated, audit, "follow_up", user, extra=note_text))
    return updated


# ==================== IN-APP NOTIFICATIONS ====================

@api_router.get("/notifications")
async def list_notifications(user=Depends(get_current_user), only_unread: bool = False, limit: int = 50):
    """Return notifications for the current user, latest first."""
    q = {"user_id": user.get("user_id", "")}
    if only_unread:
        q["read"] = False
    items = await db.notifications.find(q, {"_id": 0}).sort("created_at", -1).limit(min(200, limit)).to_list(min(200, limit))
    unread_count = await db.notifications.count_documents({"user_id": user.get("user_id", ""), "read": False})
    return {"items": items, "unread_count": unread_count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user=Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"notification_id": notification_id, "user_id": user.get("user_id", "")},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notificacion no encontrada")
    return {"ok": True}

@api_router.post("/notifications/mark-all-read")
async def mark_all_notifications_read(user=Depends(get_current_user)):
    result = await db.notifications.update_many(
        {"user_id": user.get("user_id", ""), "read": False},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"ok": True, "updated": result.modified_count}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user=Depends(get_current_user)):
    result = await db.notifications.delete_one({"notification_id": notification_id, "user_id": user.get("user_id", "")})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notificacion no encontrada")
    return {"ok": True}

# Management Review
@api_router.post("/audits/{audit_id}/management-review")
async def save_management_review(audit_id: str, request: Request, user=Depends(require_role("admin"))):
    body = await request.json()
    review = {
        "reviewer": user.get("name", ""),
        "date": datetime.now(timezone.utc).isoformat(),
        "conclusions": body.get("conclusions", ""),
        "decisions": body.get("decisions", ""),
        "resources_needed": body.get("resources_needed", ""),
        "next_steps": body.get("next_steps", ""),
    }
    await db.audits.update_one({"audit_id": audit_id}, {"$set": {"management_review": review, "status": "reviewed"}})
    return review

# Opening Meeting Minutes (Acta de Apertura) - Formal structure per reference PDF

def _company_logo_flowable(company: dict, max_w: float = 80, max_h: float = 40):
    """Return a ReportLab Image flowable from company.logo_data_url, or None if absent/invalid."""
    if not company:
        return None
    data_url = company.get("logo_data_url", "")
    if not data_url or "base64," not in data_url:
        return None
    try:
        from reportlab.platypus import Image as RLImage
        b64 = data_url.split("base64,", 1)[1]
        raw = base64.b64decode(b64)
        bio = io.BytesIO(raw)
        img = RLImage(bio, width=max_w, height=max_h, kind='proportional')
        return img
    except Exception as e:
        logger.warning(f"Could not render company logo: {e}")
        return None


def _audit_is_closed(audit: dict) -> bool:
    return (audit or {}).get("status") in ("closed", "reviewed")

async def _require_pdf_download_access(audit_id: str, user: dict) -> dict:
    """Role-based PDF download gate. sgsst_manager: only if audit closed."""
    role = user.get("role", "")
    is_priv = role in ("admin", "owner", "auditor") or is_owner(user)
    is_mgr = role == "sgsst_manager"
    if not (is_priv or is_mgr):
        raise HTTPException(status_code=403, detail="Rol sin permiso para descargar documentos")
    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    # sgsst_manager can only download closed/reviewed audits
    if is_mgr and not _audit_is_closed(audit):
        raise HTTPException(status_code=403, detail="El Responsable de SST solo puede descargar documentos de auditorias cerradas")
    return audit


@api_router.get("/audits/{audit_id}/plan/pdf")
async def generate_audit_plan_pdf(audit_id: str, user=Depends(get_current_user)):
    """Generate the Audit Plan PDF (ISO 19011 / Decreto 1072) - available before opening."""
    # Plan de Auditoria es documento pre-auditoria: admin/owner/auditor/sgsst_manager pueden descargarlo en cualquier estado
    role = user.get("role", "")
    if role not in ("admin", "owner", "auditor", "sgsst_manager"):
        raise HTTPException(status_code=403, detail="Rol sin permiso para descargar el Plan de Auditoria")
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    cid = audit.get("company_id", get_company_id(user))
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}
    checklist_count = await db.audit_checklist.count_documents({"audit_id": audit_id})

    DARK = colors.HexColor("#1F3C5E")
    CORAL = colors.HexColor("#F2A292")
    GB = colors.HexColor("#94A3B8")
    LB = colors.HexColor("#F1F5F9")
    W = colors.white

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    st = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=14, textColor=DARK, alignment=TA_CENTER, spaceAfter=10)
    sh = ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=11, textColor=DARK, spaceBefore=8, spaceAfter=4)
    sb = ParagraphStyle('b', fontName='Helvetica', fontSize=9, alignment=TA_JUSTIFY, spaceAfter=4, leading=12)
    sbl = ParagraphStyle('bl', fontName='Helvetica', fontSize=9, leftIndent=14, spaceAfter=2)
    ths = TableStyle([('BACKGROUND', (0, 0), (-1, 0), DARK), ('TEXTCOLOR', (0, 0), (-1, 0), W), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.4, GB), ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')])

    cn = company.get("name", "N/A")
    nit = company.get("nit", "N/A")
    atype = {"internal": "INTERNA", "external": "EXTERNA", "pesv": "PESV"}.get(audit.get("audit_type", "internal"), "INTERNA")
    dt = audit.get("scheduled_date", "Por confirmar")
    end_dt = audit.get("end_date", "Por confirmar")
    aud = audit.get("auditor", "")
    copasst = audit.get("copasst_member", {}) or {}

    el = []
    # Header
    el.append(HRFlowable(width="100%", thickness=3, color=CORAL))
    el.append(Spacer(1, 4))
    _logo = _company_logo_flowable(company, max_w=70, max_h=35)
    _name_para = Paragraph(f"<b>{cn}</b><br/><font size=8>NIT: {nit}</font>", ParagraphStyle('', fontName='Helvetica-Bold', fontSize=10, textColor=DARK))
    if _logo:
        _left_cell = Table([[_logo], [_name_para]], colWidths=[170])
        _left_cell.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0, 0), (-1, -1), 2), ('RIGHTPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2)]))
    else:
        _left_cell = _name_para
    ht = Table([[_left_cell, Paragraph(f"<b>PLAN DE AUDITORIA</b><br/><font size=8>AUDITORIA {atype} AL SG-SST</font>", ParagraphStyle('', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER, textColor=DARK)), Paragraph(f"<b>Version:</b> 01<br/><font size=8><b>Fecha:</b> {dt}</font>", ParagraphStyle('', fontName='Helvetica', fontSize=8, alignment=TA_CENTER))]], colWidths=[180, 200, 100])
    ht.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, DARK), ('INNERGRID', (0, 0), (-1, -1), 0.5, GB), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    el.append(ht)
    el.append(Spacer(1, 10))
    el.append(Paragraph("PLAN DE AUDITORIA AL SG-SST", st))
    el.append(Paragraph(f"<b>Auditoria:</b> {audit.get('title', '')}", sb))
    el.append(Spacer(1, 8))

    # 1. INFO GENERAL
    el.append(Paragraph("1. INFORMACION GENERAL", sh))
    info_rows = [
        ["Empresa Auditada", cn],
        ["NIT", nit],
        ["Tipo de Auditoria", atype],
        ["Fecha de Inicio", dt],
        ["Hora de Inicio", audit.get("start_time", "Por confirmar")],
        ["Fecha de Cierre Prevista", end_dt],
        ["Hora de Cierre Prevista", audit.get("end_time", "Por confirmar")],
        ["Numero de Estandares a Evaluar", str(checklist_count) if checklist_count else "Por generar"],
    ]
    if audit.get("audit_type") == "pesv":
        info_rows.append(["Nivel PESV", audit.get("pesv_level", "Avanzado").upper()])
    info_t = Table(info_rows, colWidths=[200, 320])
    info_t.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 9), ('GRID', (0, 0), (-1, -1), 0.4, GB), ('BACKGROUND', (0, 0), (0, -1), LB), ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    el.append(info_t)

    # 2. OBJETIVO
    el.append(Paragraph("2. OBJETIVO", sh))
    objective = audit.get("objective", "") or f"Verificar el cumplimiento de los estandares minimos del SG-SST de {cn} conforme a la Resolucion 0312 de 2019 y el Decreto 1072 de 2015, identificando hallazgos y oportunidades de mejora."
    el.append(Paragraph(objective, sb))

    # 3. ALCANCE
    el.append(Paragraph("3. ALCANCE", sh))
    scope = audit.get("scope", "") or "Procesos, sedes y actividades que se desarrollan en la organizacion en el marco del Sistema de Gestion de Seguridad y Salud en el Trabajo."
    el.append(Paragraph(scope, sb))

    # 4. CRITERIOS
    el.append(Paragraph("4. CRITERIOS DE AUDITORIA", sh))
    criteria = audit.get("criteria", "Resolucion 0312 de 2019, Decreto 1072 de 2015")
    el.append(Paragraph(criteria, sb))
    el.append(Paragraph("Documentos internos: Politica SG-SST, Procedimientos, Matriz de Peligros, Plan de Trabajo Anual, Programas de Vigilancia Epidemiologica.", sb))

    # 5. EQUIPO AUDITOR
    el.append(Paragraph("5. EQUIPO AUDITOR", sh))
    team_rows = [["Nombre", "Rol", "Idoneidad"]]
    if aud:
        team_rows.append([aud, "Auditor Lider", "Certificado SGI / Licencia SST"])
    for aa in audit.get("additional_auditors", []) or []:
        if aa:
            team_rows.append([aa, "Auditor de Apoyo", "Profesional SST"])
    while len(team_rows) < 3:
        team_rows.append(["_______________", "_______________", "_______________"])
    team_t = Table(team_rows, colWidths=[180, 160, 180])
    team_t.setStyle(ths)
    el.append(team_t)

    # 6. AUDITADOS
    el.append(Paragraph("6. RESPONSABLES DE LOS PROCESOS A AUDITAR", sh))
    aud_rows = [["Nombre", "Cargo / Rol"]]
    for pr in audit.get("process_responsibles", []) or []:
        if pr:
            aud_rows.append([pr, "Responsable SST / Lider de Proceso"])
    if copasst.get("name"):
        aud_rows.append([copasst.get("name", ""), f"COPASST - {copasst.get('role', '')}"])
    while len(aud_rows) < 3:
        aud_rows.append(["_______________", "_______________"])
    aud_t = Table(aud_rows, colWidths=[260, 260])
    aud_t.setStyle(ths)
    el.append(aud_t)

    # 7. METODOLOGIA
    el.append(Paragraph("7. METODOLOGIA", sh))
    el.append(Paragraph("La auditoria se desarrollara mediante:", sb))
    for x in [
        "Revision documental (politicas, procedimientos, registros, evidencias).",
        "Entrevistas con personal clave y muestreo aleatorio de trabajadores.",
        "Inspeccion fisica de areas, equipos y condiciones de trabajo.",
        "Verificacion del cumplimiento de los estandares minimos en sitio.",
        "Aplicacion del checklist conforme a la Res. 0312 de 2019 y la matriz de evaluacion.",
    ]:
        el.append(Paragraph(f"- {x}", sbl))

    # 8. CRONOGRAMA
    el.append(Paragraph("8. CRONOGRAMA DE LA AUDITORIA", sh))
    cron_rows = [
        ["Etapa", "Fecha", "Hora", "Responsable"],
        ["Reunion de Apertura", dt, audit.get("start_time", "08:00"), aud or "Auditor Lider"],
        ["Revision documental y entrevistas", dt, "09:00 - 12:00", aud or "Equipo Auditor"],
        ["Inspeccion fisica en sitio", dt, "13:30 - 16:00", aud or "Equipo Auditor"],
        ["Consolidacion de hallazgos", end_dt if end_dt != "Por confirmar" else dt, "16:00 - 17:00", aud or "Auditor Lider"],
        ["Reunion de Cierre", end_dt, audit.get("end_time", "17:00"), aud or "Auditor Lider"],
    ]
    cron_t = Table(cron_rows, colWidths=[210, 100, 100, 110])
    cron_t.setStyle(ths)
    el.append(cron_t)

    # 9. RECURSOS
    el.append(Paragraph("9. RECURSOS NECESARIOS", sh))
    for x in [
        "Sala de reuniones con proyector y conexion a internet.",
        "Acceso a la documentacion del SG-SST (fisica o digital).",
        "Disponibilidad de los responsables de los procesos y miembros del COPASST.",
        "Permisos para inspeccion en las areas operativas.",
        "Elementos de Proteccion Personal (EPP) para los auditores.",
    ]:
        el.append(Paragraph(f"- {x}", sbl))

    # 10. CONFIDENCIALIDAD
    el.append(Paragraph("10. CONFIDENCIALIDAD", sh))
    el.append(Paragraph("El equipo auditor mantendra estricta confidencialidad sobre la informacion a la que tenga acceso durante la auditoria. Los resultados unicamente se compartiran con las personas autorizadas por la alta direccion.", sb))

    # 11. FIRMAS
    el.append(Spacer(1, 16))
    el.append(Paragraph("12. FIRMAS DE APROBACION DEL PLAN", sh))
    fr = [["Auditor Lider", "Responsable SG-SST / Auditado", "Representante Legal"], ["_______________________", "_______________________", "_______________________"], [aud or "Nombre y firma", (copasst.get("name") or (audit.get("process_responsibles") or [""])[0] or "Nombre y firma"), "Nombre y firma"]]
    fr_t = Table(fr, colWidths=[173, 173, 174])
    fr_t.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 9), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 15), ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
    el.append(fr_t)
    el.append(Spacer(1, 6))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Paragraph(f"<i>Documento generado por TraciumSST - {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC</i>", ParagraphStyle('foot', fontName='Helvetica-Oblique', fontSize=7, textColor=GB, alignment=TA_CENTER)))

    doc.build(el)
    bio.seek(0)
    fname = f"Plan_Auditoria_{(cn or 'Empresa').replace(' ', '_')}_{dt}.pdf"
    return StreamingResponse(bio, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})


@api_router.post("/audits/{audit_id}/plan/send-email")
async def send_audit_plan_email(audit_id: str, request: Request, user=Depends(require_role("admin", "owner", "auditor", "sgsst_manager"))):
    """Send the Audit Plan PDF by email to the SG-SST manager and team."""
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    explicit_recipients = body.get("recipients", []) or []
    comment_raw = (body.get("comment") or "").strip()
    import html as _html
    comment = _html.escape(comment_raw)[:1000]

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    cid = audit.get("company_id", get_company_id(user))
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}

    # Build recipient list: explicit + all sgsst_managers of company + owner + auditors registered
    recipients = set()
    for e in explicit_recipients:
        if e and "@" in e:
            recipients.add(e.strip().lower())
    sgsst_users = await db.users.find({"role": {"$in": ["admin", "owner", "sgsst_manager"]}, "active": {"$ne": False}, "$or": [{"company_ids": cid}, {"email": OWNER_EMAIL}]}, {"_id": 0, "email": 1, "name": 1}).to_list(50)
    for u in sgsst_users:
        if u.get("email"):
            recipients.add(u["email"].strip().lower())
    if not recipients:
        raise HTTPException(status_code=400, detail="No hay destinatarios configurados")

    # Generate PDF in memory
    from fastapi.responses import StreamingResponse  # noqa: F401  (already imported globally)
    plan_resp = await generate_audit_plan_pdf(audit_id, user)
    # Drain the StreamingResponse body
    pdf_bytes = b""
    async for chunk in plan_resp.body_iterator:
        pdf_bytes += chunk if isinstance(chunk, bytes) else chunk.encode()

    pdf_b64 = base64.b64encode(pdf_bytes).decode("ascii")
    cn = company.get("name", "N/A")
    dt = audit.get("scheduled_date", "Por confirmar")
    fname = f"Plan_Auditoria_{cn.replace(' ', '_')}_{dt}.pdf"

    subject = f"TraciumSST - Plan de Auditoria SG-SST: {audit.get('title', '')}"
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto">
      <div style="background:#1F3C5E;padding:20px;color:#fff;border-radius:8px 8px 0 0">
        <h2 style="margin:0">Plan de Auditoria SG-SST</h2>
        <p style="margin:4px 0 0 0;font-size:13px;opacity:0.9">{cn}</p>
      </div>
      <div style="padding:20px;background:#F8F9FA;border:1px solid #E2E8F0;border-top:0;border-radius:0 0 8px 8px">
        <p>Estimado(a) Responsable de SG-SST,</p>
        <p>Adjunto encontrara el <b>Plan de Auditoria</b> correspondiente a:</p>
        <table style="width:100%;border-collapse:collapse;margin:12px 0">
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Auditoria</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('title', '')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Tipo</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('audit_type', '').upper()}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Fecha de Inicio</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{dt} {audit.get('start_time', '')}</td></tr>
          <tr><td style="padding:8px;background:#fff;border:1px solid #E2E8F0"><b>Auditor Lider</b></td><td style="padding:8px;background:#fff;border:1px solid #E2E8F0">{audit.get('auditor', 'N/A')}</td></tr>
        </table>
        {f'<p><b>Nota:</b> {comment}</p>' if comment else ''}
        <p>Por favor revise el plan y confirme la disponibilidad de recursos y personal indicados.</p>
        <p style="color:#94A3B8;font-size:11px;margin-top:16px">Enviado por: {user.get('name', '')} · {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC</p>
      </div>
    </div>
    """

    # Send with attachment using Resend Python SDK
    sent_count = 0
    failed = []
    if resend.api_key:
        for to in recipients:
            try:
                params = {
                    "from": SENDER_EMAIL,
                    "to": [to],
                    "subject": subject,
                    "html": html,
                    "attachments": [{"filename": fname, "content": pdf_b64}],
                }
                result = await asyncio.to_thread(resend.Emails.send, params)
                if result:
                    sent_count += 1
                else:
                    failed.append(to)
            except Exception as e:
                logger.error(f"Plan email send failed to {to}: {e}")
                failed.append(to)
    else:
        logger.warning("RESEND_API_KEY not configured")
    return {"sent": sent_count, "total": len(recipients), "recipients": list(recipients), "failed": failed}


@api_router.get("/audits/{audit_id}/opening-minutes/pdf")
async def generate_opening_minutes_pdf(audit_id: str, user=Depends(get_current_user)):
    await _require_pdf_download_access(audit_id, user)
    """Generate formal opening meeting minutes PDF matching reference structure"""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    cid = audit.get("company_id", get_company_id(user))
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    checklist_count = await db.audit_checklist.count_documents({"audit_id": audit_id})

    CORAL = colors.HexColor("#F2A292")
    DARK = colors.HexColor("#1F3C5E")
    BLUE = colors.HexColor("#0047AB")
    LBG = colors.HexColor("#F8F9FA")
    W = colors.white
    GB = colors.HexColor("#CBD5E1")

    cn = company.get("name", "N/A") if company else "N/A"
    nit = company.get("nit", "N/A") if company else "N/A"
    city = company.get("city", "") if company else ""
    aud = audit.get("auditor", "")
    dt = audit.get("scheduled_date", datetime.now().strftime("%Y-%m-%d"))
    atype = "Interna" if audit.get("audit_type") == "internal" else "Externa"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.7*inch, bottomMargin=0.7*inch, leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles = getSampleStyleSheet()
    st = ParagraphStyle('T', fontName='Helvetica-Bold', fontSize=14, textColor=DARK, alignment=TA_CENTER, spaceAfter=4)
    ss = ParagraphStyle('S', fontName='Helvetica-Bold', fontSize=10, textColor=DARK, alignment=TA_CENTER, spaceAfter=8)
    sh = ParagraphStyle('H', fontName='Helvetica-Bold', fontSize=11, textColor=DARK, spaceBefore=12, spaceAfter=5)
    ssh = ParagraphStyle('SH', fontName='Helvetica-Bold', fontSize=10, textColor=BLUE, spaceBefore=8, spaceAfter=4)
    sb = ParagraphStyle('B', fontName='Helvetica', fontSize=9, leading=13, alignment=TA_JUSTIFY, spaceAfter=3)
    sbb = ParagraphStyle('BB', fontName='Helvetica-Bold', fontSize=9, leading=13, alignment=TA_JUSTIFY, spaceAfter=3)
    sbl = ParagraphStyle('BL', fontName='Helvetica', fontSize=9, leading=12, leftIndent=18, spaceAfter=2)
    sm = ParagraphStyle('SM', fontName='Helvetica', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

    ths = TableStyle([('BACKGROUND',(0,0),(-1,0),DARK),('TEXTCOLOR',(0,0),(-1,0),W),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'TOP')])
    tis = TableStyle([('BACKGROUND',(0,0),(0,-1),LBG),('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6)])

    el = []
    # HEADER
    el.append(HRFlowable(width="100%", thickness=3, color=CORAL))
    el.append(Spacer(1,4))
    _logo = _company_logo_flowable(company, max_w=70, max_h=35)
    _name_para = Paragraph(f"<b>{cn}</b><br/><font size=8>NIT: {nit}</font>", ParagraphStyle('',fontName='Helvetica-Bold',fontSize=10,textColor=DARK))
    if _logo:
        _left_cell = Table([[_logo], [_name_para]], colWidths=[170])
        _left_cell.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    else:
        _left_cell = _name_para
    ht = Table([[_left_cell, Paragraph("<b>ACTA DE REUNION DE APERTURA</b><br/><font size=8>AUDITORIA SG-SST</font>", ParagraphStyle('',fontName='Helvetica-Bold',fontSize=10,alignment=TA_CENTER,textColor=DARK)), Paragraph(f"<b>Version:</b> 01<br/><font size=8><b>Fecha:</b> {dt}</font>", ParagraphStyle('',fontName='Helvetica',fontSize=8,alignment=TA_CENTER))]], colWidths=[180,200,100])
    ht.setStyle(TableStyle([('BOX',(0,0),(-1,-1),1,DARK),('INNERGRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    el.append(ht)
    el.append(Spacer(1,10))
    el.append(Paragraph("ACTA DE REUNION DE APERTURA", st))
    el.append(Paragraph(f"AUDITORIA {atype.upper()} AL SISTEMA DE GESTION DE SEGURIDAD Y SALUD EN EL TRABAJO (SG-SST)", ss))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Spacer(1,8))
    # Custom narrative if auditor has filled it via AI
    _narrative = audit.get("narrative_opening", "").strip() if audit.get("narrative_opening") else ""
    if _narrative:
        for para in _narrative.split("\n"):
            p = para.strip()
            if p:
                el.append(Paragraph(p, sb))
                el.append(Spacer(1,4))
    # Meeting details
    start_time = audit.get("start_time", "")
    t = Table([["Fecha:", dt],["Hora:", start_time or "____:____ a.m./p.m."],["Lugar:", city or "_______________"]], colWidths=[80,400])
    t.setStyle(tis)
    el.append(t)
    el.append(Spacer(1,6))
    el.append(Paragraph(f"<b>ASUNTO:</b> Reunion de apertura auditoria {atype.lower()} al Sistema de Gestion de Seguridad y Salud en el Trabajo (SG-SST) de {cn}.", sb))
    # 1. INVITADOS
    el.append(Paragraph("1. INVITADOS", sh))
    el.append(Paragraph("EQUIPO AUDITOR:", ssh))
    ed = [["Nombre","Cargo"],[aud or "_______________","Auditor Lider"]]
    for aa in audit.get("additional_auditors", []):
        ed.append([aa, "Auditor de Apoyo"])
    if len(ed) < 3:
        ed.append(["_______________","Auditor de Apoyo"])
    et = Table(ed, colWidths=[240,240]); et.setStyle(ths); el.append(et)
    el.append(Spacer(1,6))
    el.append(Paragraph("RESPONSABLES DE LOS PROCESOS Y/O FUNCIONES (SG-SST):", ssh))
    rd = [["Nombre","Cargo / Rol"]]
    for pr in audit.get("process_responsibles", []):
        rd.append([pr, "Responsable SST"])
    copasst = audit.get("copasst_member", {})
    if copasst and copasst.get("name"):
        rd.append([copasst["name"], f"COPASST - {copasst.get('role', '')}"])
    while len(rd) < 3: rd.append(["_______________","_______________"])
    rt = Table(rd, colWidths=[240,240]); rt.setStyle(ths); el.append(rt)
    # 2. REVISION DEL PLAN
    el.append(Paragraph("2. REVISION DEL PLAN DE AUDITORIA", sh))
    el.append(Paragraph("OBJETIVO DE LA AUDITORIA:", ssh))
    el.append(Paragraph(audit.get("objective","Verificar el cumplimiento del SG-SST, la eficacia de los procesos, identificar oportunidades de mejora, factores de riesgo, fortalezas, controles aplicados, analisis de datos, sostenimiento y mejora del sistema conforme a los Estandares Minimos de la Resolucion 0312 de 2019 con el fin de notificar a la alta gerencia y elaborar el respectivo plan de accion."), sb))
    el.append(Paragraph("OBJETIVOS ESPECIFICOS:", ssh))
    for o in ["Verificar la normatividad y procedimientos implementados en el cumplimiento de los objetivos y compromisos adquiridos en el control del SG-SST.","Identificar oportunidades de mejora a traves de la auditoria y determinar responsables.","Presentar un informe a la alta direccion con informacion para la toma de decisiones sobre la mejora del SG-SST."]:
        el.append(Paragraph(f"- {o}", sbl))
    el.append(Paragraph("ALCANCE:", ssh))
    el.append(Paragraph(f"Este proceso de auditoria tendra como base los lineamientos definidos por el Ministerio del Trabajo a traves del Decreto 1072 (Articulo 2.2.4.6.30 Alcance de la auditoria de cumplimiento del SG-SST) y la Resolucion 0312 (Estandares Minimos del SG-SST), abarcando entre otros:", sb))
    for i,item in enumerate(["El cumplimiento de la politica de seguridad y salud en el trabajo;","Los indicadores de estructura, proceso y resultado (a evaluar durante la auditoria);","La participacion de los trabajadores;","El desarrollo de la responsabilidad y la obligacion de rendir cuentas;","El mecanismo de comunicacion de los contenidos del SG-SST a los trabajadores;","La planificacion, desarrollo y aplicacion del SG-SST;","La gestion del cambio;","La consideracion de la seguridad y salud en el trabajo en las nuevas adquisiciones;","El alcance y aplicacion del SG-SST frente a los proveedores y contratistas;","La supervision y medicion de los resultados;","El proceso de investigacion de incidentes, accidentes de trabajo y enfermedades laborales, y su efecto sobre el mejoramiento de la SST en la empresa;","El desarrollo del proceso de auditoria; y","La evaluacion por parte de la alta direccion."],1):
        el.append(Paragraph(f"<b>{i}.</b> {item}", sbl))
    # 3. PERIODICIDAD
    el.append(Paragraph("3. PERIODICIDAD", sh))
    el.append(Paragraph(f"La auditoria {atype.lower()} al SG-SST se realizara anualmente, conforme al programa de auditoria de {cn}.", sb))
    # 4. IDONEIDAD DEL AUDITOR
    el.append(Paragraph("4. IDONEIDAD DEL AUDITOR", sh))
    el.append(Paragraph("Los auditores son independientes del proceso auditado, lo que permite estar libres de sesgo y conflicto de intereses. Los auditores deben mantener una actitud objetiva a lo largo del proceso para asegurarse de que los hallazgos y conclusiones estaran basados solo en la evidencia de la auditoria.", sb))
    el.append(Paragraph("Conforme a la Norma Tecnica Colombiana ISO 19011, Numeral 7.2, un auditor debe ser:", sb))
    for l,a,d in [("a","Etico","imparcial, sincero, honesto y discreto"),("b","De mentalidad abierta","dispuesto a considerar ideas o puntos de vista alternativos"),("c","Diplomatico","con tacto en las relaciones con las personas"),("d","Observador","activamente consciente del entorno fisico y las actividades"),("e","Perceptivo","instintivamente consciente y capaz de entender las situaciones"),("f","Versatil","se adapta facilmente a diferentes situaciones"),("g","Tenaz","persistente, orientado hacia el logro de los objetivos"),("h","Decidido","alcanza conclusiones oportunas basadas en el analisis y razonamiento logicos"),("i","Seguro de si mismo","actua y funciona de forma independiente a la vez que se relaciona eficazmente con otros")]:
        el.append(Paragraph(f"<b>{l})</b> <b>{a}:</b> {d};", sbl))
    # 5. PROCESO AUDITADO
    el.append(Paragraph("5. PROCESO AUDITADO", sh))
    el.append(Paragraph(f"Sistema de Gestion de Seguridad y Salud en el Trabajo (SG-SST) de {cn}. Estandares minimos aplicables: {checklist_count or 'por generar'}.", sb))
    # 6. MARCO LEGAL
    el.append(Paragraph("6. MARCO LEGAL", sh))
    for ml in ["Decreto 1072 de 2015, Articulo 2.2.4.6.29 y Articulo 2.2.4.6.30","Resolucion 0312 de 2019 - Estandares Minimos del SG-SST","NTC ISO 19011 - Directrices para la auditoria de sistemas de gestion"]:
        el.append(Paragraph(f"- {ml}", sbl))
    # 7. AGENDA
    el.append(Paragraph("7. AGENDA DE LA REUNION", sh))
    el.append(Paragraph("ORDEN DEL DIA:", ssh))
    ag = Table([["No.","Actividad"],["1","Llamado a lista."],["2","Presentacion, propositos y objetivos de la auditoria."],["3","Proposiciones, compromisos y tareas."],["4","Cierre."]], colWidths=[40,440]); ag.setStyle(ths); el.append(ag)
    el.append(Spacer(1,8))
    el.append(Paragraph("DESARROLLO DEL ORDEN DEL DIA:", ssh))
    el.append(Paragraph("<b>1. Llamado a lista.</b>", sbb))
    el.append(Paragraph("La reunion de apertura conto con los siguientes servidores:", sb))
    ad = [["No.","Nombre y Apellido","Cargo / Rol","Firma"]]
    r = 1
    # Auditor lider
    ad.append([str(r), aud or "", "Auditor Lider", ""]); r += 1
    # Auditores de apoyo
    for aa in audit.get("additional_auditors", []):
        if aa:
            ad.append([str(r), aa, "Auditor de Apoyo", ""]); r += 1
    # Responsables de los procesos
    for pr in audit.get("process_responsibles", []):
        if pr:
            ad.append([str(r), pr, "Responsable SST", ""]); r += 1
    # COPASST
    if copasst and copasst.get("name"):
        ad.append([str(r), copasst["name"], f"COPASST - {copasst.get('role', '')}", ""]); r += 1
    # Otros asistentes registrados en el campo libre attendees
    for at_str in audit.get("attendees", []):
        if at_str and at_str not in [row[1] for row in ad[1:]]:
            ad.append([str(r), at_str, "Asistente", ""]); r += 1
    # Filas vacias para firmas adicionales presenciales (minimo 5 filas)
    target_rows = max(5, len(ad))
    while len(ad) < target_rows: ad.append([str(r), "", "", ""]); r += 1
    at = Table(ad, colWidths=[30,180,150,120]); at.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),DARK),('TEXTCOLOR',(0,0),(-1,0),W),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('ROWHEIGHTS',(0,1),(-1,-1),24),('ALIGN',(0,0),(0,-1),'CENTER')])); el.append(at)
    el.append(Spacer(1,8))
    el.append(Paragraph("<b>2. Presentacion de la Auditoria.</b>", sbb))
    el.append(Paragraph(f"A la hora senalada se inicio la reunion de apertura de la Auditoria al SG-SST procediendo a hacer la presentacion formal del Equipo Auditor indicando las funciones de cada uno. De igual forma se llevo a cabo la presentacion de cada uno de los servidores y contratistas que participaran o atenderan como designados para las visitas de auditoria, como facilitadores y articuladores de la informacion requerida para su desarrollo.", sb))
    el.append(Paragraph("Luego se explico por parte del equipo auditor el contenido del plan de auditoria, en cuanto a metodologia de seguimiento y el termino previsto para la ejecucion.", sb))
    el.append(Paragraph(f"Se dieron a conocer los objetivos, el alcance, los criterios y el cronograma de auditoria plasmados en el plan de auditoria que se entrego a los responsables del SG-SST de {cn}.", sb))
    for p in ["Se dio a conocer la programacion de la ejecucion de las actividades de auditoria, se confirma la fecha y hora del cierre de la auditoria.","Se confirmaron los canales de comunicacion a ser empleados durante el proceso auditor.","Se confirmaron asuntos relacionados con la confidencialidad y reserva de la informacion.","Se confirmaron las personas que eventualmente atenderan la auditoria, acorde a las actividades del proceso que desarrolle cada uno.","Se aclaro que la labor de auditoria en ningun momento afecta o entorpece el normal desarrollo de las actividades propias del proceso auditado.","Se aclaro que es responsabilidad de la dependencia o proceso auditado el contenido en calidad y cantidad de la informacion suministrada, asi como el cumplimiento de las normas que le son aplicables; senalando que es obligacion del equipo auditor expresar con independencia una conclusion sobre el cumplimiento de las disposiciones aplicables, fundamentada en los resultados de la auditoria."]:
        el.append(Paragraph(p, sb))
    el.append(Paragraph("<b>2.1. Anexos:</b>", sbb))
    el.append(Paragraph("- Diapositivas de Apertura Auditoria al SG-SST.", sbl))
    el.append(Paragraph("- Plan de Auditoria.", sbl))
    el.append(Spacer(1,6))
    el.append(Paragraph("<b>3. Proposiciones, compromisos y tareas.</b>", sbb))
    el.append(Paragraph(f"Los asistentes a la reunion se comprometen a entregar informacion veridica y oportuna en los tiempos establecidos dentro del cronograma del Plan de Auditoria al SG-SST de {cn}.", sb))
    el.append(Spacer(1,6))
    el.append(Paragraph("<b>4. Cierre.</b>", sbb))
    el.append(Paragraph(f"Siendo las ____:____ a.m./p.m. se da por terminada la reunion donde se apertura el proceso de auditoria al SG-SST de {cn} y se aprueba el plan de auditoria por parte del equipo auditor y los responsables del proceso y/o funciones del SG-SST, firman los que a ella asistieron.", sb))
    el.append(Spacer(1,10))
    el.append(Paragraph("OBSERVACIONES:", sh))
    for _ in range(3): el.append(Paragraph("_" * 95, sb))
    # FIRMAS
    el.append(Spacer(1,16))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Spacer(1,8))
    sig = AUDITOR_SIGNATURE
    sig_name = aud or sig["name"]
    from reportlab.lib.enums import TA_CENTER as _TA_C
    _sc = ParagraphStyle('sc', fontName='Helvetica', fontSize=8, alignment=_TA_C, textColor=colors.HexColor("#475569"), leading=10)
    _sn = ParagraphStyle('sn', fontName='Helvetica-Bold', fontSize=9, alignment=_TA_C, textColor=DARK)
    el.append(Paragraph("_________________________________", _sc))
    el.append(Paragraph(f"<b>{sig_name}</b>", _sn))
    for line in sig["title_lines"]:
        el.append(Paragraph(line, _sc))
    el.append(Spacer(1,16))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Paragraph(f"{sig_name} | (+57) 321 620 8039 | stephania.ceballos@laofi.onmicrosoft.com", sm))
    el.append(Paragraph("Ciudadela Complex, Llanogrande Lote 57-58 Rionegro - Antioquia", sm))
    doc.build(el)
    buf.seek(0)
    fn = f"Acta_Apertura_Auditoria_{cn.replace(' ','_')}_{dt}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@api_router.get("/audits/{audit_id}/closing-minutes/pdf")
async def generate_closing_minutes_pdf(audit_id: str, user=Depends(get_current_user)):
    await _require_pdf_download_access(audit_id, user)
    """Generate formal closing meeting minutes PDF for audit"""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    cid = audit.get("company_id", get_company_id(user))
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    checklist = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    findings = await db.findings.find({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}}, {"_id": 0}).to_list(100)
    action_plans = await db.action_plans.find({"audit_id": audit_id}, {"_id": 0}).to_list(100)

    CORAL = colors.HexColor("#F2A292")
    DARK = colors.HexColor("#1F3C5E")
    BLUE = colors.HexColor("#0047AB")
    LBG = colors.HexColor("#F8F9FA")
    W = colors.white
    GB = colors.HexColor("#CBD5E1")
    RED = colors.HexColor("#D90429")
    GREEN = colors.HexColor("#2A9D8F")
    YELLOW = colors.HexColor("#FFC300")

    cn = company.get("name", "N/A") if company else "N/A"
    nit = company.get("nit", "N/A") if company else "N/A"
    city = company.get("city", "") if company else ""
    aud = audit.get("auditor", "")
    dt = audit.get("scheduled_date", datetime.now().strftime("%Y-%m-%d"))
    end_dt = audit.get("end_date", datetime.now().strftime("%Y-%m-%d"))
    atype = "Interna" if audit.get("audit_type") == "internal" else "Externa"

    total_ck = len(checklist)
    ck_na = len([c for c in checklist if c.get("result") == "no_aplica"])
    ck_evaluable = total_ck - ck_na
    ck_ok = len([c for c in checklist if c.get("result") == "cumple"])
    ck_nc = len([c for c in checklist if c.get("result") == "no_cumple"])
    ck_partial = len([c for c in checklist if c.get("result") == "parcial"])
    ck_pending = total_ck - ck_ok - ck_nc - ck_partial - ck_na
    compliance_pct = round((ck_ok / ck_evaluable * 100) if ck_evaluable > 0 else 0, 1)

    nc_findings = [f for f in findings if f.get("finding_type") == "no_conformity"]
    obs_findings = [f for f in findings if f.get("finding_type") == "observation"]
    opp_findings = [f for f in findings if f.get("finding_type") == "improvement"]
    closed_plans = [p for p in action_plans if p.get("status") == "closed"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.7*inch, bottomMargin=0.7*inch, leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles = getSampleStyleSheet()
    st = ParagraphStyle('T', fontName='Helvetica-Bold', fontSize=14, textColor=DARK, alignment=TA_CENTER, spaceAfter=4)
    ss = ParagraphStyle('S', fontName='Helvetica-Bold', fontSize=10, textColor=DARK, alignment=TA_CENTER, spaceAfter=8)
    sh = ParagraphStyle('H', fontName='Helvetica-Bold', fontSize=11, textColor=DARK, spaceBefore=12, spaceAfter=5)
    ssh = ParagraphStyle('SH', fontName='Helvetica-Bold', fontSize=10, textColor=BLUE, spaceBefore=8, spaceAfter=4)
    sb = ParagraphStyle('B', fontName='Helvetica', fontSize=9, leading=13, alignment=TA_JUSTIFY, spaceAfter=3)
    sbb = ParagraphStyle('BB', fontName='Helvetica-Bold', fontSize=9, leading=13, alignment=TA_JUSTIFY, spaceAfter=3)
    sbl = ParagraphStyle('BL', fontName='Helvetica', fontSize=9, leading=12, leftIndent=18, spaceAfter=2)
    sm = ParagraphStyle('SM', fontName='Helvetica', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

    ths = TableStyle([('BACKGROUND',(0,0),(-1,0),DARK),('TEXTCOLOR',(0,0),(-1,0),W),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'TOP')])
    tis = TableStyle([('BACKGROUND',(0,0),(0,-1),LBG),('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),6)])

    el = []
    # HEADER
    el.append(HRFlowable(width="100%", thickness=3, color=CORAL))
    el.append(Spacer(1, 4))
    _logo = _company_logo_flowable(company, max_w=70, max_h=35)
    _name_para = Paragraph(f"<b>{cn}</b><br/><font size=8>NIT: {nit}</font>", ParagraphStyle('', fontName='Helvetica-Bold', fontSize=10, textColor=DARK))
    if _logo:
        _left_cell = Table([[_logo], [_name_para]], colWidths=[170])
        _left_cell.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2),('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    else:
        _left_cell = _name_para
    ht = Table([[_left_cell, Paragraph("<b>ACTA DE REUNION DE CIERRE</b><br/><font size=8>AUDITORIA SG-SST</font>", ParagraphStyle('', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER, textColor=DARK)), Paragraph(f"<b>Version:</b> 01<br/><font size=8><b>Fecha:</b> {end_dt}</font>", ParagraphStyle('', fontName='Helvetica', fontSize=8, alignment=TA_CENTER))]], colWidths=[180, 200, 100])
    ht.setStyle(TableStyle([('BOX',(0,0),(-1,-1),1,DARK),('INNERGRID',(0,0),(-1,-1),0.5,GB),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    el.append(ht)
    el.append(Spacer(1, 10))
    el.append(Paragraph("ACTA DE REUNION DE CIERRE", st))
    el.append(Paragraph(f"AUDITORIA {atype.upper()} AL SISTEMA DE GESTION DE SEGURIDAD Y SALUD EN EL TRABAJO (SG-SST)", ss))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Spacer(1, 8))

    # Meeting details
    t = Table([["Fecha:", end_dt], ["Hora:", "____:____ a.m./p.m."], ["Lugar:", city or "_______________"]], colWidths=[80, 400])
    t.setStyle(tis)
    el.append(t)
    el.append(Spacer(1, 6))
    el.append(Paragraph(f"<b>ASUNTO:</b> Reunion de cierre de auditoria {atype.lower()} al Sistema de Gestion de Seguridad y Salud en el Trabajo (SG-SST) de {cn}.", sb))

    # Custom narrative if auditor has filled it via AI
    _narrative_close = audit.get("narrative_closing", "").strip() if audit.get("narrative_closing") else ""
    if _narrative_close:
        el.append(Spacer(1, 4))
        for para in _narrative_close.split("\n"):
            p = para.strip()
            if p:
                el.append(Paragraph(p, sb))
                el.append(Spacer(1, 4))

    # 1. ASISTENTES
    el.append(Paragraph("1. ASISTENTES", sh))
    el.append(Paragraph("EQUIPO AUDITOR:", ssh))
    ed = [["Nombre", "Cargo"], [aud or "_______________", "Auditor Lider"]]
    for aa in audit.get("additional_auditors", []):
        ed.append([aa, "Auditor de Apoyo"])
    if len(ed) < 3:
        ed.append(["_______________", "Auditor de Apoyo"])
    et = Table(ed, colWidths=[240, 240]); et.setStyle(ths); el.append(et)
    el.append(Spacer(1, 6))
    el.append(Paragraph("RESPONSABLES DE LOS PROCESOS Y/O FUNCIONES (SG-SST):", ssh))
    rd = [["Nombre", "Cargo / Rol"]]
    for pr in audit.get("process_responsibles", []):
        rd.append([pr, "Responsable SST"])
    copasst = audit.get("copasst_member", {})
    if copasst and copasst.get("name"):
        rd.append([copasst["name"], f"COPASST - {copasst.get('role', '')}"])
    while len(rd) < 3:
        rd.append(["_______________", "_______________"])
    rt = Table(rd, colWidths=[240, 240]); rt.setStyle(ths); el.append(rt)

    # 2. RESUMEN DE LA AUDITORIA
    el.append(Paragraph("2. RESUMEN DE LA AUDITORIA", sh))
    el.append(Paragraph("DATOS GENERALES:", ssh))
    info_data = [
        [Paragraph("<b>Tipo de Auditoria:</b>", sbb), Paragraph(atype, sb)],
        [Paragraph("<b>Periodo de Ejecucion:</b>", sbb), Paragraph(f"{dt} al {end_dt}", sb)],
        [Paragraph("<b>Criterios:</b>", sbb), Paragraph(audit.get("criteria", "Resolucion 0312 de 2019, Decreto 1072 de 2015"), sb)],
        [Paragraph("<b>Alcance:</b>", sbb), Paragraph(audit.get("scope", f"SG-SST de {cn}"), sb)],
    ]
    info_t = Table(info_data, colWidths=[140, 340])
    info_t.setStyle(tis)
    el.append(info_t)
    el.append(Spacer(1, 6))
    el.append(Paragraph("OBJETIVO:", ssh))
    el.append(Paragraph(audit.get("objective", "Verificar el cumplimiento del SG-SST conforme a los Estandares Minimos de la Resolucion 0312 de 2019 y el Decreto 1072 de 2015."), sb))

    # 3. RESULTADOS DE LA AUDITORIA
    el.append(Paragraph("3. RESULTADOS DE LA AUDITORIA", sh))
    el.append(Paragraph("CUMPLIMIENTO DE ESTANDARES:", ssh))
    res_data = [
        ["Indicador", "Cantidad", "Porcentaje"],
        [Paragraph("<b>Estandares Totales</b>", sbb), str(total_ck), ""],
        [Paragraph("<font color='#94A3B8'><b>No Aplica</b></font>", sb), str(ck_na), ""],
        [Paragraph("<b>Estandares Evaluables</b>", sbb), str(ck_evaluable), "100%"],
        [Paragraph("<font color='#2A9D8F'><b>Cumple</b></font>", sb), str(ck_ok), f"{compliance_pct}%"],
        [Paragraph("<font color='#D90429'><b>No Cumple</b></font>", sb), str(ck_nc), f"{round(ck_nc/ck_evaluable*100,1) if ck_evaluable else 0}%"],
        [Paragraph("<font color='#FFC300'><b>Cumple Parcialmente</b></font>", sb), str(ck_partial), f"{round(ck_partial/ck_evaluable*100,1) if ck_evaluable else 0}%"],
        ["Pendiente de evaluar", str(ck_pending), f"{round(ck_pending/ck_evaluable*100,1) if ck_evaluable else 0}%"],
    ]
    res_t = Table(res_data, colWidths=[220, 100, 100])
    res_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK), ('TEXTCOLOR', (0, 0), (-1, 0), W),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, GB), ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]))
    el.append(res_t)
    el.append(Spacer(1, 8))

    el.append(Paragraph("HALLAZGOS IDENTIFICADOS:", ssh))
    hall_data = [
        ["Tipo de Hallazgo", "Cantidad"],
        ["No Conformidades", str(len(nc_findings))],
        ["Observaciones", str(len(obs_findings))],
        ["Oportunidades de Mejora", str(len(opp_findings))],
        [Paragraph("<b>Total Hallazgos</b>", sbb), str(len(findings))],
    ]
    hall_t = Table(hall_data, colWidths=[300, 100])
    hall_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK), ('TEXTCOLOR', (0, 0), (-1, 0), W),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, GB), ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    el.append(hall_t)
    el.append(Spacer(1, 6))

    # 4. DETALLE DE HALLAZGOS
    if findings:
        el.append(Paragraph("4. DETALLE DE HALLAZGOS", sh))
        for i, f in enumerate(findings, 1):
            ftype_labels = {"no_conformity": "No Conformidad", "observation": "Observacion", "improvement": "Oportunidad de Mejora"}
            ftype = ftype_labels.get(f.get("finding_type", ""), f.get("finding_type", ""))
            el.append(Paragraph(f"<b>Hallazgo {i}:</b> {ftype}", sbb))
            el.append(Paragraph(f"<b>Area:</b> {f.get('area', 'N/A')} | <b>Estandar:</b> {f.get('standard_ref', 'N/A')}", sb))
            el.append(Paragraph(f"{f.get('description', 'Sin descripcion')}", sb))
            if f.get("recommendation"):
                el.append(Paragraph(f"<b>Recomendacion:</b> {f.get('recommendation')}", sb))
            el.append(Spacer(1, 4))
        next_section = 5
    else:
        el.append(Paragraph("4. DETALLE DE HALLAZGOS", sh))
        el.append(Paragraph("No se identificaron hallazgos durante el proceso de auditoria.", sb))
        next_section = 5

    # 5. PLANES DE ACCION
    el.append(Paragraph(f"{next_section}. PLANES DE ACCION", sh))
    if action_plans:
        el.append(Paragraph(f"Se definieron <b>{len(action_plans)}</b> planes de accion correctiva y/o preventiva. Estado actual:", sb))
        plan_data = [["No.", "Accion", "Responsable", "Fecha Limite", "Estado"]]
        for i, p in enumerate(action_plans, 1):
            status_labels = {"open": "Abierto", "in_progress": "En Progreso", "closed": "Cerrado", "overdue": "Vencido"}
            plan_data.append([
                str(i),
                Paragraph(p.get("action", "N/A")[:80], sb),
                p.get("responsible", "N/A"),
                p.get("due_date", "N/A"),
                status_labels.get(p.get("status", ""), p.get("status", ""))
            ])
        plan_t = Table(plan_data, colWidths=[30, 190, 90, 80, 70])
        plan_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DARK), ('TEXTCOLOR', (0, 0), (-1, 0), W),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, GB), ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4), ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ]))
        el.append(plan_t)
    else:
        el.append(Paragraph("Los planes de accion correctivos y/o preventivos se definiran en los proximos 30 dias habiles, con responsables y fechas concretas, y se documentaran en el Informe Final de Auditoria, conforme al Decreto 1072 de 2015.", sb))
    el.append(Spacer(1, 6))

    # 6. CONCLUSIONES
    next_section += 1
    el.append(Paragraph(f"{next_section}. CONCLUSIONES Y RECOMENDACIONES", sh))
    exec_summary = audit.get("executive_summary", "")
    if exec_summary:
        el.append(Paragraph(exec_summary, sb))
    else:
        el.append(Paragraph(f"Con base en los resultados de la auditoria {atype.lower()} al SG-SST de {cn}, se obtiene un nivel de cumplimiento del <b>{compliance_pct}%</b> sobre los estandares minimos evaluados conforme a la Resolucion 0312 de 2019.", sb))
        if nc_findings:
            el.append(Paragraph(f"Se identificaron <b>{len(nc_findings)}</b> no conformidades que requieren atencion prioritaria mediante los planes de accion definidos.", sb))
        if obs_findings:
            el.append(Paragraph(f"Se registraron <b>{len(obs_findings)}</b> observaciones que, si bien no constituyen incumplimiento, representan oportunidades de fortalecimiento del sistema.", sb))
    el.append(Spacer(1, 4))
    el.append(Paragraph("Se recomienda:", sbb))
    el.append(Paragraph("- Implementar los planes de accion en los plazos establecidos.", sbl))
    el.append(Paragraph("- Realizar seguimiento periodico al cumplimiento de las acciones correctivas.", sbl))
    el.append(Paragraph("- Programar auditoria de seguimiento para verificar la eficacia de las acciones implementadas.", sbl))
    el.append(Paragraph("- Socializar los resultados de la auditoria con todos los niveles de la organizacion.", sbl))

    # 7. COMPROMISOS
    next_section += 1
    el.append(Paragraph(f"{next_section}. COMPROMISOS Y ACUERDOS", sh))
    el.append(Paragraph("Los asistentes a la reunion de cierre acuerdan:", sb))
    el.append(Paragraph("- Dar cumplimiento a los planes de accion establecidos dentro de los plazos definidos.", sbl))
    el.append(Paragraph("- Reportar el avance de las acciones correctivas al responsable del SG-SST.", sbl))
    el.append(Paragraph("- Mantener la evidencia documental de la implementacion de las acciones.", sbl))
    el.append(Paragraph(f"- Presentar informe de seguimiento a la alta direccion de {cn}.", sbl))

    # 8. CIERRE
    next_section += 1
    el.append(Paragraph(f"{next_section}. CIERRE", sh))
    el.append(Paragraph(f"Siendo las ____:____ a.m./p.m. se da por terminada la reunion de cierre de la Auditoria {atype} al SG-SST de {cn}. Se presenta el consolidado de resultados y se socializan los hallazgos identificados, planes de accion y compromisos adquiridos.", sb))
    el.append(Paragraph("Firman los que a ella asistieron en constancia de lo anterior:", sb))

    # LISTA DE ASISTENCIA CON FIRMA
    el.append(Spacer(1, 8))
    ad = [["No.", "Nombre y Apellido", "Cargo / Rol", "Firma"]]
    r = 1
    ad.append([str(r), aud or "", "Auditor Lider", ""]); r += 1
    for aa in audit.get("additional_auditors", []):
        if aa:
            ad.append([str(r), aa, "Auditor de Apoyo", ""]); r += 1
    for pr in audit.get("process_responsibles", []):
        if pr:
            ad.append([str(r), pr, "Responsable SST", ""]); r += 1
    if copasst and copasst.get("name"):
        ad.append([str(r), copasst["name"], f"COPASST - {copasst.get('role', '')}", ""]); r += 1
    for at_str in audit.get("attendees", []):
        if at_str and at_str not in [row[1] for row in ad[1:]]:
            ad.append([str(r), at_str, "Asistente", ""]); r += 1
    target_rows = max(5, len(ad))
    while len(ad) < target_rows:
        ad.append([str(r), "", "", ""]); r += 1
    at = Table(ad, colWidths=[30, 180, 150, 120])
    at.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK), ('TEXTCOLOR', (0, 0), (-1, 0), W),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, GB), ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5), ('ROWHEIGHTS', (0, 1), (-1, -1), 24),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
    ]))
    el.append(at)

    el.append(Spacer(1, 10))
    el.append(Paragraph("OBSERVACIONES:", sh))
    for _ in range(3):
        el.append(Paragraph("_" * 95, sb))

    # FIRMAS
    el.append(Spacer(1, 16))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Spacer(1, 8))
    sig = AUDITOR_SIGNATURE
    sig_name = aud or sig["name"]
    _sc = ParagraphStyle('sc2', fontName='Helvetica', fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor("#475569"), leading=10)
    _sn = ParagraphStyle('sn2', fontName='Helvetica-Bold', fontSize=9, alignment=TA_CENTER, textColor=DARK)
    el.append(Paragraph("_________________________________", _sc))
    el.append(Paragraph(f"<b>{sig_name}</b>", _sn))
    for line in sig["title_lines"]:
        el.append(Paragraph(line, _sc))
    el.append(Spacer(1, 16))
    el.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    el.append(Paragraph(f"{sig_name} | (+57) 321 620 8039 | stephania.ceballos@laofi.onmicrosoft.com", sm))

    doc.build(el)
    buf.seek(0)
    fn = f"Acta_Cierre_Auditoria_{cn.replace(' ', '_')}_{end_dt}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# Historical comparison
@api_router.get("/audits/history/comparison")
async def get_audit_comparison(user=Depends(get_current_user)):
    cid = get_company_id(user)
    audits = await db.audits.find({"company_id": cid, "status": {"$in": ["closed", "reviewed"]}}, {"_id": 0}).sort("scheduled_date", -1).to_list(20)
    history = []
    for a in audits:
        findings_count = await db.findings.count_documents({"audit_id": a["audit_id"], "status": {"$ne": "resolved_by_compliance"}})
        nc_count = await db.findings.count_documents({"audit_id": a["audit_id"], "finding_type": "no_conformity", "status": {"$ne": "resolved_by_compliance"}})
        obs_count = await db.findings.count_documents({"audit_id": a["audit_id"], "finding_type": "observation", "status": {"$ne": "resolved_by_compliance"}})
        closed_findings = await db.findings.count_documents({"audit_id": a["audit_id"], "status": "closed"})
        checklist_total = await db.audit_checklist.count_documents({"audit_id": a["audit_id"]})
        checklist_na = await db.audit_checklist.count_documents({"audit_id": a["audit_id"], "result": "no_aplica"})
        checklist_evaluable = checklist_total - checklist_na
        checklist_ok = await db.audit_checklist.count_documents({"audit_id": a["audit_id"], "result": "cumple"})
        history.append({
            "audit_id": a["audit_id"],
            "title": a["title"],
            "date": a.get("scheduled_date", ""),
            "audit_type": a.get("audit_type", ""),
            "status": a["status"],
            "total_findings": findings_count,
            "no_conformities": nc_count,
            "observations": obs_count,
            "closed_findings": closed_findings,
            "closure_rate": round((closed_findings / findings_count * 100) if findings_count > 0 else 0, 1),
            "checklist_total": checklist_evaluable,
            "checklist_compliant": checklist_ok,
            "compliance_rate": round((checklist_ok / checklist_evaluable * 100) if checklist_evaluable > 0 else 0, 1),
        })
    return history

# AI-assisted writing for audits
@api_router.post("/audits/ai/assist")
async def ai_audit_assist(request: Request, user=Depends(get_current_user)):
    body = await request.json()
    assist_type = body.get("type", "finding")
    context = body.get("context", "")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")
        system_msgs = {
            "finding": "Eres un auditor experto en SG-SST colombiano. Redacta hallazgos de auditoria claros, concisos y accionables basados en la Resolucion 0312/2019 y el Decreto 1072/2015. Responde SOLO con el texto del hallazgo, sin encabezados ni explicaciones adicionales.",
            "action_plan": "Eres un experto en SG-SST colombiano. Genera planes de accion correctiva especificos, medibles y con plazos claros. Responde SOLO con las acciones recomendadas en formato de lista.",
            "executive_summary": "Eres un consultor senior de SG-SST en Colombia. Genera un resumen ejecutivo de auditoria para la alta direccion. Debe ser fiel a los datos proporcionados, sin inventar informacion. Incluye: estado general del SG-SST, nivel de cumplimiento, hallazgos clave, riesgos principales, recomendaciones prioritarias y conclusion. Formato profesional, tecnico y conciso.",
            "checklist_observation": "Eres un auditor de SG-SST. Redacta una observacion tecnica clara para un item de checklist de auditoria. Responde SOLO con la observacion.",
            "management_review": "Eres un asesor de la alta direccion en SG-SST colombiano. Redacta conclusiones y decisiones para la revision por la alta direccion basadas en los resultados de la auditoria. Formato profesional ejecutivo.",
            "strengths": "Eres un consultor senior de SG-SST. Con base en los resultados de la auditoria, identifica y redacta las FORTALEZAS del sistema. Solo menciona aspectos que cumplen segun los datos. No inventes. Formato de lista con viñetas, tono profesional.",
            "findings_report": "Eres un auditor lider de SG-SST. Redacta la seccion de hallazgos del informe consolidado. Agrupa por tipo (No Conformidades, Observaciones, Oportunidades de Mejora). Mejora la claridad y redaccion de cada hallazgo sin alterar el contenido tecnico ni inventar informacion nueva. Tono formal y tecnico.",
            "recommendations": "Eres un consultor experto de SG-SST en Colombia. Redacta recomendaciones especificas basadas en los hallazgos de la auditoria. Cada recomendacion debe ser accionable, medible y con un plazo sugerido. No inventes hallazgos que no existan. Formato de lista numerada.",
            "conclusions": "Eres un auditor lider de SG-SST. Redacta las conclusiones del informe final de auditoria. Incluye: valoracion general del SG-SST, nivel de cumplimiento con la Res. 0312/2019, principales brechas, y perspectiva de mejora. Basate estrictamente en los datos proporcionados. Tono ejecutivo y profesional.",
            "opening_narrative": "Eres un auditor lider de SG-SST colombiano. Redacta la parte narrativa del ACTA DE APERTURA de auditoria: introduccion, objetivo de la reunion, presentacion del equipo, metodologia y agenda. Formato formal y protocolar, 3-5 parrafos. Usa la informacion proporcionada sin inventar datos.",
            "closing_narrative": "Eres un auditor lider de SG-SST colombiano. Redacta la parte narrativa del ACTA DE CIERRE de auditoria: introduccion, sintesis de hallazgos presentados, compromisos adquiridos por la organizacion, proximos pasos, agradecimientos. Formato formal y protocolar, 3-5 parrafos. Usa la informacion proporcionada sin inventar datos.",
            "report_narrative": "Eres un auditor lider SG-SST. Redacta el parrafo introductorio (resumen ejecutivo) y la conclusion final del INFORME DE AUDITORIA. Tono ejecutivo, basado en los datos. 4-6 parrafos bien estructurados.",
            "action_plan_action": "Eres un experto en SG-SST colombiano. Redacta UNA sola accion correctiva/preventiva/de mejora especifica, medible, alineada con el hallazgo. Responde con un parrafo o lista corta (maximo 80 palabras). Sin titulos ni explicaciones.",
            "action_plan_resources": "Eres un experto en SG-SST. Lista los RECURSOS necesarios (humanos, tecnicos, economicos, materiales) para ejecutar la accion descrita. Responde solo con la lista, sin encabezado.",
            "action_plan_evidence": "Eres un experto en SG-SST. Define la EVIDENCIA OBJETIVA que demuestra el cumplimiento de la accion (registros, listados de asistencia, fotografias, certificados, actas, etc.). Maximo 50 palabras, lista corta o parrafo unico, sin titulos.",
        }
        chat = LlmChat(
            api_key=api_key,
            session_id=f"audit_ai_{uuid.uuid4().hex[:6]}",
            system_message=system_msgs.get(assist_type, system_msgs["finding"])
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text=context)
        response = await chat.send_message(msg)
        return {"result": response, "type": assist_type}
    except Exception as e:
        logger.error(f"AI audit assist error: {e}")
        return {"result": f"Error: {str(e)}", "type": assist_type}


@api_router.put("/audits/{audit_id}/ai-redaction")
async def save_ai_redaction(audit_id: str, request: Request, user=Depends(require_role("admin", "auditor"))):
    """Save AI-improved text sections for the consolidated report"""
    body = await request.json()
    allowed = {"ai_redacted_summary", "ai_redacted_findings", "ai_redacted_strengths",
               "ai_redacted_recommendations", "ai_redacted_conclusions", "executive_summary",
               "narrative_opening", "narrative_closing", "narrative_report"}
    update = {k: v for k, v in body.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["report_stale"] = False  # Clear stale flag when saving redaction
    await db.audits.update_one({"audit_id": audit_id}, {"$set": update})
    return {"message": "Redaccion guardada", "updated_fields": list(update.keys())}


# ==================== REPORTS MODULE ====================

@api_router.get("/reports/excel")
async def generate_excel_report(report_type: str = "general", user=Depends(get_current_user)):
    import openpyxl
    wb = openpyxl.Workbook()

    # Incidents sheet
    ws_inc = wb.active
    ws_inc.title = "Incidentes"
    ws_inc.append(["ID", "Tipo", "Fecha", "Ubicacion", "Descripcion", "Severidad", "Estado"])
    incidents = await db.incidents.find({}, {"_id": 0}).to_list(500)
    for inc in incidents:
        ws_inc.append([inc.get("incident_id",""), inc.get("incident_type",""), inc.get("date",""), inc.get("location",""), inc.get("description",""), inc.get("severity",""), inc.get("status","")])

    # Hazards sheet
    ws_hzd = wb.create_sheet("Peligros")
    ws_hzd.append(["ID", "Area", "Tipo", "Descripcion", "Probabilidad", "Severidad", "Nivel Riesgo", "Categoria"])
    hazards = await db.hazards.find({}, {"_id": 0}).to_list(500)
    for h in hazards:
        ws_hzd.append([h.get("hazard_id",""), h.get("area",""), h.get("hazard_type",""), h.get("description",""), h.get("probability",0), h.get("severity",0), h.get("risk_level",0), h.get("risk_category","")])

    # Audits sheet
    ws_aud = wb.create_sheet("Auditorias")
    ws_aud.append(["ID", "Titulo", "Tipo", "Fecha", "Auditor", "Estado"])
    audits = await db.audits.find({}, {"_id": 0}).to_list(500)
    for a in audits:
        ws_aud.append([a.get("audit_id",""), a.get("title",""), a.get("audit_type",""), a.get("scheduled_date",""), a.get("auditor",""), a.get("status","")])

    # Trainings sheet
    ws_trn = wb.create_sheet("Capacitaciones")
    ws_trn.append(["ID", "Titulo", "Formador", "Fecha", "Duracion(h)", "Estado"])
    trainings = await db.trainings.find({}, {"_id": 0}).to_list(500)
    for t in trainings:
        ws_trn.append([t.get("training_id",""), t.get("title",""), t.get("trainer",""), t.get("scheduled_date",""), t.get("duration_hours",0), t.get("status","")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename=reporte_sgsst_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@api_router.get("/reports/pdf")
async def generate_pdf_report(report_type: str = "general", user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Reporte SG-SST", styles['Title']))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # KPIs summary
    total_incidents = await db.incidents.count_documents({})
    open_incidents = await db.incidents.count_documents({"status": "open"})
    total_hazards = await db.hazards.count_documents({})
    high_risks = await db.hazards.count_documents({"risk_category": {"$in": ["high", "critical"]}})
    total_audits = await db.audits.count_documents({})
    total_trainings = await db.trainings.count_documents({})

    kpi_data = [
        ["Indicador", "Valor"],
        ["Total Incidentes", str(total_incidents)],
        ["Incidentes Abiertos", str(open_incidents)],
        ["Total Peligros", str(total_hazards)],
        ["Peligros Alto Riesgo", str(high_risks)],
        ["Total Auditorias", str(total_audits)],
        ["Total Capacitaciones", str(total_trainings)],
    ]
    kpi_table = Table(kpi_data, colWidths=[250, 150])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0047AB")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))

    # Incidents table
    elements.append(Paragraph("Incidentes Recientes", styles['Heading2']))
    incidents = await db.incidents.find({}, {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)
    if incidents:
        inc_data = [["ID", "Tipo", "Fecha", "Severidad", "Estado"]]
        for inc in incidents:
            inc_data.append([inc.get("incident_id","")[:12], inc.get("incident_type",""), inc.get("date",""), inc.get("severity",""), inc.get("status","")])
        inc_table = Table(inc_data, colWidths=[80, 100, 80, 80, 80])
        inc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0047AB")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(inc_table)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                            headers={"Content-Disposition": f"attachment; filename=reporte_sgsst_{datetime.now().strftime('%Y%m%d')}.pdf"})

# ==================== AI ANALYSIS ====================

@api_router.post("/ai/analyze")
async def ai_analyze(request: Request, user=Depends(get_current_user)):
    body = await request.json()
    query = body.get("query", "")
    context_type = body.get("context_type", "general")

    context_data = {}
    if context_type in ["general", "dashboard"]:
        context_data["incidents"] = await db.incidents.count_documents({})
        context_data["open_incidents"] = await db.incidents.count_documents({"status": "open"})
        context_data["hazards"] = await db.hazards.count_documents({})
        context_data["high_risks"] = await db.hazards.count_documents({"risk_category": {"$in": ["high", "critical"]}})
        context_data["trainings"] = await db.trainings.count_documents({})
        context_data["audits"] = await db.audits.count_documents({})
        checklist_total = await db.checklist.count_documents({})
        checklist_compliant = await db.checklist.count_documents({"compliant": True})
        context_data["compliance_pct"] = round((checklist_compliant / checklist_total * 100) if checklist_total > 0 else 0, 1)

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")
        chat = LlmChat(
            api_key=api_key,
            session_id=f"sgsst_{user.get('user_id', 'anon')}_{uuid.uuid4().hex[:6]}",
            system_message="Eres un experto en Seguridad y Salud en el Trabajo (SG-SST) en Colombia, especializado en el Decreto 1072 de 2015. Analiza los datos proporcionados y da recomendaciones claras y accionables en espanol."
        ).with_model("openai", "gpt-5.2")
        prompt = f"Datos del sistema SG-SST: {context_data}\n\nConsulta del usuario: {query}"
        msg = UserMessage(text=prompt)
        response = await chat.send_message(msg)
        return {"analysis": response, "context": context_data}
    except Exception as e:
        logger.error(f"AI analysis error: {e}")
        return {"analysis": f"Error al generar analisis: {str(e)}", "context": context_data}


# ==================== AI CHATBOT NORMATIVO ====================

CHATBOT_SYSTEM_MESSAGE = """Eres TraciumBot, un asistente experto en normativa colombiana de Seguridad y Salud en el Trabajo (SG-SST) y Seguridad Vial (PESV). Tu rol es responder consultas de profesionales SST de forma clara, precisa y accionable.

DOMINIOS DE CONOCIMIENTO:
1. Decreto 1072 de 2015 (Libro 2, Parte 2, Titulo 4, Capitulo 6): 16 elementos del SG-SST, obligaciones del empleador, ciclo PHVA (Planear, Hacer, Verificar, Actuar).
2. Resolucion 0312 de 2019: 60 estandares minimos, clasificacion por tamano (<=10, 11-50, >50) y nivel de riesgo (I-V), puntajes Critico (<60%), Moderado (60-85%), Aceptable (>85%).
3. Resolucion 40595 de 2022: Plan Estrategico de Seguridad Vial (PESV), niveles Basico/Estandar/Avanzado, 4 fases (Planificacion, Implementacion, Seguimiento, Mejora).
4. Ley 1562 de 2012, Decreto 1443 de 2014, Resolucion 1401 de 2007 (investigacion de incidentes), GTC 45 (matriz IPER).

REGLAS DE RESPUESTA:
- Responde SIEMPRE en espanol, tono profesional y cercano.
- Cita articulos y resoluciones especificas cuando sea relevante (Ej: "Segun el Art. 2.2.4.6.12 del Decreto 1072...").
- Si la consulta es ambigua, pide un dato minimo (tamano empresa, nivel riesgo, sector).
- Estructura respuestas largas con vinetas o numeracion.
- Manten respuestas concisas (maximo 250 palabras salvo que el usuario pida detalle).
- NO inventes articulos ni numeros de resolucion. Si no estas seguro, dilo.
- Si te preguntan fuera de alcance (temas no SST/PESV), redirige amablemente al tema.
- Para temas criticos (accidentes graves, muerte), recomienda notificacion a ARL dentro de 48h (Art. 2.2.4.1.7 Dec. 1072) y la Direccion Territorial del Ministerio de Trabajo.

Eres util, directo y siempre basado en la normativa vigente colombiana."""


@api_router.post("/ai/chat")
async def ai_chat(request: Request, user=Depends(get_current_user)):
    """Persistent AI chatbot for SG-SST/PESV normative queries with multi-turn context."""
    body = await request.json()
    message = (body.get("message") or "").strip()
    session_id = body.get("session_id") or f"chat_{user.get('user_id', 'anon')}_{uuid.uuid4().hex[:10]}"
    include_context = bool(body.get("include_context", False))

    if not message:
        raise HTTPException(status_code=400, detail="Mensaje vacio")
    if len(message) > 4000:
        raise HTTPException(status_code=400, detail="Mensaje demasiado largo (maximo 4000 caracteres)")

    # Build live company context if requested
    context_block = ""
    if include_context:
        cid = get_company_id(user)
        comp = await db.companies.find_one({"company_id": cid}, {"_id": 0}) or {}
        q = {"company_id": cid}
        total_findings = await db.findings.count_documents({**q, "status": {"$ne": "resolved_by_compliance"}})
        open_findings = await db.findings.count_documents({**q, "status": "open"})
        top_findings = await db.findings.find({**q, "status": "open"}, {"_id": 0}).limit(5).to_list(5)
        total_incidents = await db.incidents.count_documents(q)
        open_incidents = await db.incidents.count_documents({**q, "status": "open"})
        high_risks = await db.hazards.count_documents({**q, "risk_category": {"$in": ["high", "critical"]}})
        total_plans = await db.action_plans.count_documents({**q, "status": {"$in": ["open", "in_progress"]}})
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        overdue = await db.action_plans.count_documents({**q, "status": {"$in": ["open", "in_progress"]}, "due_date": {"$lt": today, "$ne": ""}})
        latest_audit = await db.audits.find(q, {"_id": 0}).sort("created_at", -1).limit(1).to_list(1)
        score_pct = latest_audit[0].get("score_result", {}).get("percentage") if latest_audit and latest_audit[0].get("score_result") else None
        audit_title = latest_audit[0].get("title") if latest_audit else None

        findings_brief = "\n".join([f"  - [{f.get('type','')}] {f.get('title','')} (sev: {f.get('severity','')})" for f in top_findings]) or "  - (sin hallazgos abiertos)"
        context_block = f"""

DATOS EN VIVO DE LA EMPRESA (usalos SOLO si la pregunta lo requiere):
- Empresa: {comp.get('name','N/A')} ({comp.get('workers_count','?')} trabajadores, Riesgo nivel {comp.get('risk_level','?')})
- Hallazgos abiertos: {open_findings} (de {total_findings} activos)
- Top hallazgos abiertos:
{findings_brief}
- Incidentes: {open_incidents} abiertos de {total_incidents}
- Riesgos altos/criticos: {high_risks}
- Planes de accion abiertos: {total_plans} (vencidos: {overdue})
- Ultima auditoria: {audit_title or 'sin auditorias'} - Puntaje: {score_pct if score_pct is not None else 'N/A'}%

INSTRUCCION: Cuando el usuario pregunte sobre su estado actual, cita estos numeros. Para preguntas puramente normativas, ignora este bloque."""

    # Load prior history (last 20 messages for token efficiency)
    history = await db.chat_messages.find(
        {"session_id": session_id, "user_id": user.get("user_id")},
        {"_id": 0}
    ).sort("created_at", 1).limit(40).to_list(40)
    history = history[-20:] if len(history) > 20 else history

    # Save user message
    user_msg = {
        "message_id": uuid.uuid4().hex,
        "session_id": session_id,
        "user_id": user.get("user_id"),
        "role": "user",
        "content": message,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one({**user_msg})

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message=CHATBOT_SYSTEM_MESSAGE + context_block
        ).with_model("openai", "gpt-5.2")

        # Replay history into context
        context_parts = []
        for h in history:
            prefix = "Usuario" if h.get("role") == "user" else "Asistente"
            context_parts.append(f"[{prefix}]: {h.get('content', '')}")
        context_parts.append(f"[Usuario actual]: {message}")
        full_prompt = "\n\n".join(context_parts) if history else message

        msg = UserMessage(text=full_prompt)
        response_text = await chat.send_message(msg)

        assistant_msg = {
            "message_id": uuid.uuid4().hex,
            "session_id": session_id,
            "user_id": user.get("user_id"),
            "role": "assistant",
            "content": response_text,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.chat_messages.insert_one({**assistant_msg})

        return {
            "session_id": session_id,
            "message": response_text,
            "message_id": assistant_msg["message_id"],
            "context_used": include_context,
        }
    except Exception as e:
        logger.error(f"Chatbot error: {e}")
        raise HTTPException(status_code=500, detail=f"Error del asistente: {str(e)}")


@api_router.get("/ai/chat/history")
async def ai_chat_history(session_id: str, user=Depends(get_current_user)):
    """Retrieve chat history for a session (user-scoped)."""
    msgs = await db.chat_messages.find(
        {"session_id": session_id, "user_id": user.get("user_id")},
        {"_id": 0}
    ).sort("created_at", 1).to_list(200)
    return {"session_id": session_id, "messages": msgs}


@api_router.delete("/ai/chat/history")
async def ai_chat_clear(session_id: str, user=Depends(get_current_user)):
    """Clear chat history for a session."""
    result = await db.chat_messages.delete_many({
        "session_id": session_id,
        "user_id": user.get("user_id")
    })
    return {"deleted": result.deleted_count}


# ==================== USERS MANAGEMENT ====================

@api_router.get("/users")
async def get_users(user=Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, request: Request, user=Depends(require_role("admin"))):
    body = await request.json()
    new_role = body.get("role")
    if new_role not in ROLE_HIERARCHY:
        raise HTTPException(status_code=400, detail="Rol invalido")
    update = {"role": new_role}
    if body.get("company_id"):
        update["active_company_id"] = body["company_id"]
    result = await db.users.update_one({"user_id": user_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Rol actualizado"}

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data(user=Depends(get_current_user)):
    """Seed demo data for testing"""
    # Seed checklist (Estandares Minimos)
    standards = [
        {"standard": "1.1.1", "description": "Asignacion de responsable del SG-SST", "compliant": True, "evidence": "Acta de nombramiento", "observations": ""},
        {"standard": "1.1.2", "description": "Responsabilidades en SST", "compliant": True, "evidence": "Manual de funciones", "observations": ""},
        {"standard": "1.1.3", "description": "Asignacion de recursos para el SG-SST", "compliant": False, "evidence": "", "observations": "Pendiente aprobacion presupuestal"},
        {"standard": "1.1.4", "description": "Afiliacion al Sistema de Seguridad Social", "compliant": True, "evidence": "Planillas de pago", "observations": ""},
        {"standard": "2.1.1", "description": "Politica de SST firmada y divulgada", "compliant": True, "evidence": "Politica publicada", "observations": ""},
        {"standard": "2.2.1", "description": "Objetivos del SG-SST definidos", "compliant": False, "evidence": "", "observations": "En revision"},
        {"standard": "3.1.1", "description": "Evaluaciones medicas ocupacionales", "compliant": True, "evidence": "Certificados medicos", "observations": ""},
        {"standard": "3.1.2", "description": "Actividades de promocion y prevencion", "compliant": True, "evidence": "Registro fotografico", "observations": ""},
        {"standard": "4.1.1", "description": "Identificacion de peligros y evaluacion de riesgos", "compliant": False, "evidence": "", "observations": "Matriz desactualizada"},
        {"standard": "4.2.1", "description": "Medidas de prevencion y control implementadas", "compliant": True, "evidence": "Registros de inspeccion", "observations": ""},
    ]
    existing = await db.checklist.count_documents({})
    if existing == 0:
        for s in standards:
            s["item_id"] = f"chk_{uuid.uuid4().hex[:8]}"
            s["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.checklist.insert_many(standards)

    # Seed activities
    activities = [
        {"title": "Revision de politica SST", "description": "Actualizar la politica de SST anual", "responsible": "Responsable SG-SST", "due_date": "2026-03-15", "category": "policy", "priority": "high"},
        {"title": "Capacitacion primeros auxilios", "description": "Formacion en primeros auxilios para brigada", "responsible": "Lider de area", "due_date": "2026-04-01", "category": "training", "priority": "high"},
        {"title": "Inspeccion de extintores", "description": "Inspeccion trimestral de extintores", "responsible": "Responsable SG-SST", "due_date": "2026-03-30", "category": "inspection", "priority": "medium"},
        {"title": "Simulacro de evacuacion", "description": "Simulacro semestral de evacuacion", "responsible": "Brigada de emergencia", "due_date": "2026-06-15", "category": "emergency", "priority": "high"},
    ]
    existing_act = await db.activities.count_documents({})
    if existing_act == 0:
        for a in activities:
            a["activity_id"] = f"act_{uuid.uuid4().hex[:8]}"
            a["status"] = "pending"
            a["completion_percentage"] = 0
            a["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.activities.insert_many(activities)

    # Seed hazards
    hazards = [
        {"area": "Produccion", "hazard_type": "Mecanico", "description": "Atrapamiento por maquinaria sin guarda", "risk_source": "Maquina cortadora", "probability": 4, "severity": 5, "existing_controls": "Senalizacion", "proposed_controls": "Instalar guardas de seguridad"},
        {"area": "Oficinas", "hazard_type": "Ergonomico", "description": "Postura prolongada en silla inadecuada", "risk_source": "Puesto de trabajo", "probability": 3, "severity": 2, "existing_controls": "Pausas activas", "proposed_controls": "Cambiar sillas ergonomicas"},
        {"area": "Bodega", "hazard_type": "Locativo", "description": "Pisos humedos sin senalizacion", "risk_source": "Area de almacenamiento", "probability": 4, "severity": 3, "existing_controls": "Limpieza periodica", "proposed_controls": "Instalar senalizacion permanente y pisos antideslizantes"},
    ]
    existing_hzd = await db.hazards.count_documents({})
    if existing_hzd == 0:
        for h in hazards:
            h["hazard_id"] = f"hzd_{uuid.uuid4().hex[:8]}"
            h["risk_level"] = h["probability"] * h["severity"]
            h["risk_category"] = calc_risk_category(h["risk_level"])
            h["status"] = "active"
            h["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.hazards.insert_many(hazards)

    # Seed incidents
    incidents = [
        {"incident_type": "Accidente", "date": "2026-01-15", "location": "Produccion - Zona A", "description": "Corte en mano derecha por maquina cortadora", "affected_person": "Juan Perez", "severity": "moderate", "immediate_actions": "Primeros auxilios aplicados"},
        {"incident_type": "Incidente", "date": "2026-01-20", "location": "Bodega", "description": "Caida de estanteria por sobrecarga", "affected_person": "", "severity": "minor", "immediate_actions": "Area acordonada"},
    ]
    existing_inc = await db.incidents.count_documents({})
    if existing_inc == 0:
        for i in incidents:
            i["incident_id"] = f"inc_{uuid.uuid4().hex[:8]}"
            i["root_cause"] = ""
            i["corrective_actions"] = ""
            i["status"] = "open"
            i["created_by"] = "Sistema"
            i["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.incidents.insert_many(incidents)

    return {"message": "Seed data created successfully"}

# ==================== COMPANY CONFIGURATION (legacy compatible) ====================

@api_router.get("/company")
async def get_company(user=Depends(get_current_user)):
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not company:
        company = {
            "company_id": "default",
            "name": "Mi Empresa",
            "nit": "",
            "workers_count": 25,
            "risk_level": 2,
            "economic_activity": "",
            "city": "",
            "sedes": ["Sede Principal"],
            "processes": ["Administrativo", "Operativo"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.companies.insert_one(company)
        company.pop("_id", None)
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"active_company_id": "default", "company_ids": ["default"]}})
    return company

@api_router.put("/company")
async def update_company(request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    body.pop("_id", None)
    cid = get_company_id(user)
    if isinstance(body.get("sedes"), str):
        body["sedes"] = [s.strip() for s in body["sedes"].split(",")]
    if isinstance(body.get("processes"), str):
        body["processes"] = [s.strip() for s in body["processes"].split(",")]
    existing = await db.companies.find_one({"company_id": cid})
    if existing:
        await db.companies.update_one({"company_id": cid}, {"$set": body})
    else:
        body["company_id"] = cid
        body["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.companies.insert_one(body)
    result = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    return result

# ==================== STANDARDS BANK (Resolución 0312/2019) ====================

@api_router.get("/standards/bank")
async def get_standards_bank(user=Depends(get_current_user)):
    """Get the complete standards bank from Resolución 0312/2019"""
    return STANDARDS_BANK

@api_router.get("/standards/applicable")
async def get_applicable(user=Depends(get_current_user)):
    """Get applicable standards based on company classification"""
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    workers = company.get("workers_count", 25) if company else 25
    risk = company.get("risk_level", 2) if company else 2
    applicable_codes = get_applicable_standards(workers, risk)
    applicable = [s for s in STANDARDS_BANK if s["code"] in applicable_codes]
    return {
        "company_type": "10_or_less" if workers <= 10 else ("11_to_50" if workers <= 50 else "50_plus"),
        "workers_count": workers,
        "risk_level": risk,
        "total_standards": len(STANDARDS_BANK),
        "applicable_count": len(applicable),
        "total_weight": get_total_weight(applicable_codes),
        "standards": applicable
    }

@api_router.post("/standards/seed")
async def seed_standards(user=Depends(get_current_user)):
    """Seed all standards into DB with compliance status"""
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    workers = company.get("workers_count", 25) if company else 25
    risk = company.get("risk_level", 2) if company else 2
    applicable_codes = get_applicable_standards(workers, risk)

    existing = await db.standards_compliance.count_documents({"company_id": cid})
    if existing > 0:
        return {"message": f"Standards already seeded ({existing} items). Use reset to reload."}

    docs = []
    for s in STANDARDS_BANK:
        doc = {
            "code": s["code"],
            "standard": s["standard"],
            "subestandar": s["subestandar"],
            "description": s["description"],
            "detail": s["detail"],
            "evidence": s["evidence"],
            "phva": s["phva"],
            "chapter": s["chapter"],
            "weight": s["weight"],
            "applicable": s["code"] in applicable_codes,
            "compliant": False,
            "evidence_uploaded": "",
            "evidence_files": [],
            "observations": "",
            "responsible": "",
            "sede": "",
            "process": "",
            "company_id": cid,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": user.get("name", "")
        }
        docs.append(doc)
    await db.standards_compliance.insert_many(docs)
    return {"message": f"Seeded {len(docs)} standards. {len(applicable_codes)} applicable."}

@api_router.post("/standards/reset")
async def reset_standards(user=Depends(require_role("admin", "sgsst_manager"))):
    """Reset and reseed standards based on current company config"""
    cid = get_company_id(user)
    await db.standards_compliance.delete_many({"company_id": cid})
    return await seed_standards(user=user)

@api_router.get("/standards/compliance")
async def get_standards_compliance(user=Depends(get_current_user)):
    """Get all standards with compliance status"""
    cid = get_company_id(user)
    items = await db.standards_compliance.find({"company_id": cid}, {"_id": 0}).to_list(200)
    if not items:
        # Auto-seed if empty
        await seed_standards(user=user)
        items = await db.standards_compliance.find({"company_id": cid}, {"_id": 0}).to_list(200)
    return items

@api_router.put("/standards/compliance/{code}")
async def update_standard_compliance(code: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    cid = get_company_id(user)
    body = await request.json()
    body.pop("_id", None)
    body.pop("code", None)
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    body["updated_by"] = user.get("name", "")
    result = await db.standards_compliance.update_one({"code": code, "company_id": cid}, {"$set": body})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Standard not found")
    updated = await db.standards_compliance.find_one({"code": code, "company_id": cid}, {"_id": 0})
    return updated

@api_router.get("/standards/compliance/summary")
async def get_compliance_summary(user=Depends(get_current_user)):
    """Get compliance summary with multiple views"""
    cid = get_company_id(user)
    items = await db.standards_compliance.find({"company_id": cid}, {"_id": 0}).to_list(200)
    if not items:
        await seed_standards(user=user)
        items = await db.standards_compliance.find({"company_id": cid}, {"_id": 0}).to_list(200)

    applicable = [i for i in items if i.get("applicable", False)]
    compliant = [i for i in applicable if i.get("compliant", False)]
    total_weight = sum(i.get("weight", 0) for i in applicable)
    compliant_weight = sum(i.get("weight", 0) for i in compliant)
    score = round((compliant_weight / total_weight * 100) if total_weight > 0 else 0, 1)

    # By PHVA
    phva_summary = {}
    for phase in ["PLANEAR", "HACER", "VERIFICAR", "ACTUAR"]:
        phase_items = [i for i in applicable if i.get("phva") == phase]
        phase_compliant = [i for i in phase_items if i.get("compliant", False)]
        phase_weight = sum(i.get("weight", 0) for i in phase_items)
        phase_compliant_weight = sum(i.get("weight", 0) for i in phase_compliant)
        phva_summary[phase] = {
            "total": len(phase_items),
            "compliant": len(phase_compliant),
            "total_weight": phase_weight,
            "compliant_weight": phase_compliant_weight,
            "percentage": round((phase_compliant_weight / phase_weight * 100) if phase_weight > 0 else 0, 1)
        }

    # By Chapter/Standard
    chapter_summary = {}
    for item in applicable:
        ch = item.get("standard", "Otro")
        if ch not in chapter_summary:
            chapter_summary[ch] = {"total": 0, "compliant": 0, "weight": 0, "compliant_weight": 0}
        chapter_summary[ch]["total"] += 1
        chapter_summary[ch]["weight"] += item.get("weight", 0)
        if item.get("compliant"):
            chapter_summary[ch]["compliant"] += 1
            chapter_summary[ch]["compliant_weight"] += item.get("weight", 0)
    for ch in chapter_summary:
        w = chapter_summary[ch]["weight"]
        cw = chapter_summary[ch]["compliant_weight"]
        chapter_summary[ch]["percentage"] = round((cw / w * 100) if w > 0 else 0, 1)

    # By Responsible
    responsible_summary = {}
    for item in applicable:
        resp = item.get("responsible", "Sin asignar") or "Sin asignar"
        if resp not in responsible_summary:
            responsible_summary[resp] = {"total": 0, "compliant": 0}
        responsible_summary[resp]["total"] += 1
        if item.get("compliant"):
            responsible_summary[resp]["compliant"] += 1
    for r in responsible_summary:
        t = responsible_summary[r]["total"]
        c = responsible_summary[r]["compliant"]
        responsible_summary[r]["percentage"] = round((c / t * 100) if t > 0 else 0, 1)

    # By Sede
    sede_summary = {}
    for item in applicable:
        sede = item.get("sede", "Sin asignar") or "Sin asignar"
        if sede not in sede_summary:
            sede_summary[sede] = {"total": 0, "compliant": 0}
        sede_summary[sede]["total"] += 1
        if item.get("compliant"):
            sede_summary[sede]["compliant"] += 1
    for s in sede_summary:
        t = sede_summary[s]["total"]
        c = sede_summary[s]["compliant"]
        sede_summary[s]["percentage"] = round((c / t * 100) if t > 0 else 0, 1)

    return {
        "overall": {
            "total_standards": len(items),
            "applicable_count": len(applicable),
            "compliant_count": len(compliant),
            "total_weight": total_weight,
            "compliant_weight": compliant_weight,
            "score": score
        },
        "phva": phva_summary,
        "by_chapter": chapter_summary,
        "by_responsible": responsible_summary,
        "by_sede": sede_summary
    }

# ==================== DECRETO 1072 COMPONENTS ====================

@api_router.get("/decreto1072/components")
async def get_decreto_components(user=Depends(get_current_user)):
    """Get all Decreto 1072 components with compliance mapping"""
    cid = get_company_id(user)
    items = await db.standards_compliance.find({"applicable": True, "company_id": cid}, {"_id": 0}).to_list(200)
    compliance_map = {i["code"]: i.get("compliant", False) for i in items}

    result = []
    for comp in DECRETO_1072_COMPONENTS:
        related = comp["related_standards"]
        related_compliant = sum(1 for c in related if compliance_map.get(c, False))
        related_total = len(related)
        pct = round((related_compliant / related_total * 100) if related_total > 0 else 0, 1)
        result.append({
            **comp,
            "compliance_percentage": pct,
            "compliant_count": related_compliant,
            "total_count": related_total
        })
    return result

# ==================== FILE UPLOAD (Object Storage) ====================

@api_router.post("/files/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    if not can_write(user):
        raise HTTPException(status_code=403, detail="Permiso insuficiente")
    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    cid = get_company_id(user)
    path = f"{APP_NAME}/{cid}/uploads/{uuid.uuid4()}.{ext}"
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (max 20MB)")
    result = put_object(path, data, file.content_type or "application/octet-stream")
    file_doc = {
        "file_id": f"file_{uuid.uuid4().hex[:8]}",
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "company_id": cid,
        "uploaded_by": user.get("name", ""),
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.files.insert_one(file_doc)
    file_doc.pop("_id", None)
    return file_doc

@api_router.get("/files/{file_id}/download")
async def download_file(file_id: str, user=Depends(get_current_user)):
    record = await db.files.find_one({"file_id": file_id, "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    data, content_type = get_object(record["storage_path"])
    return Response(
        content=data,
        media_type=record.get("content_type", content_type),
        headers={"Content-Disposition": f'attachment; filename="{record.get("original_filename", "file")}"'}
    )

@api_router.get("/files")
async def list_files(user=Depends(get_current_user)):
    cid = get_company_id(user)
    files = await db.files.find({"company_id": cid, "is_deleted": False}, {"_id": 0}).to_list(500)
    return files

@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str, user=Depends(get_current_user)):
    if not can_write(user):
        raise HTTPException(status_code=403, detail="Permiso insuficiente")
    result = await db.files.update_one({"file_id": file_id}, {"$set": {"is_deleted": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return {"message": "Eliminado"}

# ==================== MULTI-COMPANY MANAGEMENT ====================

@api_router.get("/companies")
async def list_companies(user=Depends(get_current_user)):
    """List companies the user has access to. Owner/admin see all, others only assigned (strict)."""
    if is_owner(user) or user.get("role") == "admin":
        companies = await db.companies.find({}, {"_id": 0}).to_list(100)
    else:
        user_companies = user.get("company_ids", [])
        # Strict isolation: remove legacy "default" leaks for non-admin/non-owner users
        user_companies = [c for c in user_companies if c and c != "default"]
        if not user_companies:
            return []
        companies = await db.companies.find({"company_id": {"$in": user_companies}}, {"_id": 0}).to_list(100)
    return companies

@api_router.post("/companies")
async def create_company(request: Request, user=Depends(require_role("admin"))):
    body = await request.json()
    company = {
        "company_id": f"comp_{uuid.uuid4().hex[:8]}",
        "name": body.get("name", ""),
        "nit": body.get("nit", ""),
        "workers_count": body.get("workers_count", 25),
        "risk_level": body.get("risk_level", 2),
        "economic_activity": body.get("economic_activity", ""),
        "city": body.get("city", ""),
        "sedes": body.get("sedes", ["Sede Principal"]),
        "processes": body.get("processes", ["Administrativo", "Operativo"]),
        "created_by": user.get("user_id", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.companies.insert_one(company)
    company.pop("_id", None)
    # Add company to user's company list
    user_companies = user.get("company_ids", [])
    user_companies.append(company["company_id"])
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"company_ids": user_companies, "active_company_id": company["company_id"]}})
    return company

@api_router.put("/companies/{company_id}")
async def update_company_multi(company_id: str, request: Request, user=Depends(require_role("admin", "sgsst_manager"))):
    body = await request.json()
    body.pop("_id", None)
    body.pop("company_id", None)
    if isinstance(body.get("sedes"), str):
        body["sedes"] = [s.strip() for s in body["sedes"].split(",")]
    if isinstance(body.get("processes"), str):
        body["processes"] = [s.strip() for s in body["processes"].split(",")]
    result = await db.companies.update_one({"company_id": company_id}, {"$set": body})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    updated = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    return updated

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, user=Depends(require_role("admin"))):
    result = await db.companies.delete_one({"company_id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return {"message": "Empresa eliminada"}

# ---- Company Logo Upload (base64 stored in DB) ----
ALLOWED_LOGO_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp"}
MAX_LOGO_BYTES = 2 * 1024 * 1024  # 2 MB

@api_router.post("/companies/{company_id}/logo")
async def upload_company_logo(company_id: str, file: UploadFile = File(...), user=Depends(require_role("admin", "sgsst_manager"))):
    """Upload (or replace) the company logo. Stored as base64 data URL in company.logo_data_url."""
    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    ct = (file.content_type or "").lower()
    if ct not in ALLOWED_LOGO_TYPES:
        raise HTTPException(status_code=400, detail="Formato no permitido. Use PNG, JPG o WebP")
    raw = await file.read()
    if len(raw) > MAX_LOGO_BYTES:
        raise HTTPException(status_code=400, detail="Logo demasiado grande. Maximo 2 MB")
    if len(raw) < 100:
        raise HTTPException(status_code=400, detail="Archivo vacio o corrupto")

    # Optional: validate image integrity and reject exotic content
    try:
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(raw))
        img.verify()
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es una imagen valida")

    b64 = base64.b64encode(raw).decode("ascii")
    data_url = f"data:{ct};base64,{b64}"
    await db.companies.update_one({"company_id": company_id}, {"$set": {
        "logo_data_url": data_url,
        "logo_content_type": ct,
        "logo_uploaded_at": datetime.now(timezone.utc).isoformat(),
        "logo_uploaded_by": user.get("name", ""),
    }})
    return {"message": "Logo actualizado", "logo_data_url": data_url}

@api_router.delete("/companies/{company_id}/logo")
async def delete_company_logo(company_id: str, user=Depends(require_role("admin", "sgsst_manager"))):
    result = await db.companies.update_one(
        {"company_id": company_id},
        {"$unset": {"logo_data_url": "", "logo_content_type": "", "logo_uploaded_at": "", "logo_uploaded_by": ""}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return {"message": "Logo eliminado"}

@api_router.post("/companies/{company_id}/switch")
async def switch_company(company_id: str, user=Depends(get_current_user)):
    """Switch user's active company. Non-owner users can only switch to assigned companies."""
    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    if not is_owner(user) and user.get("role") != "admin":
        user_companies = user.get("company_ids", [])
        if company_id not in user_companies:
            raise HTTPException(status_code=403, detail="No tienes acceso a esta empresa")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"active_company_id": company_id}})
    return {"message": "Empresa activa cambiada", "company": company}

@api_router.get("/companies/active")
async def get_active_company(user=Depends(get_current_user)):
    cid = get_company_id(user)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not company:
        # Only auto-create "default" for owner/admin. Other roles must have a real assigned company.
        if not (is_owner(user) or user.get("role") == "admin"):
            raise HTTPException(status_code=403, detail="No tienes empresa asignada. Contacta al administrador.")
        company = {
            "company_id": "default",
            "name": "Mi Empresa",
            "nit": "",
            "workers_count": 25,
            "risk_level": 2,
            "economic_activity": "",
            "city": "",
            "sedes": ["Sede Principal"],
            "processes": ["Administrativo", "Operativo"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.companies.insert_one(company)
        company.pop("_id", None)
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"active_company_id": "default", "company_ids": ["default"]}})
    return company


# ==================== EMAIL ALERTS MODULE ====================

def build_alert_email_html(plans, company_name, alert_type="upcoming"):
    """Build HTML email for action plan alerts"""
    if alert_type == "overdue":
        title = "ALERTA: Planes de Accion Vencidos"
        color = "#D90429"
        intro = f"Los siguientes planes de accion de <b>{company_name}</b> se encuentran <b>vencidos</b> y requieren atencion inmediata:"
    elif alert_type == "due_today":
        title = "Planes de Accion que Vencen HOY"
        color = "#F97316"
        intro = f"Los siguientes planes de accion de <b>{company_name}</b> vencen <b>hoy</b>:"
    else:
        title = "Planes de Accion Proximos a Vencer"
        color = "#FFC300"
        intro = f"Los siguientes planes de accion de <b>{company_name}</b> estan proximos a vencer:"
    rows = ""
    for p in plans:
        rows += f'<tr><td style="padding:6px;border:1px solid #E2E8F0;font-size:12px">{p.get("action","")[:120]}</td><td style="padding:6px;border:1px solid #E2E8F0;font-size:12px;text-align:center">{p.get("due_date","N/A")}</td><td style="padding:6px;border:1px solid #E2E8F0;font-size:12px">{p.get("responsible","N/A")}</td><td style="padding:6px;border:1px solid #E2E8F0;font-size:12px;text-align:center">{p.get("status","open")}</td></tr>'
    return f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
<div style="background:{color};padding:16px;border-radius:8px 8px 0 0"><h2 style="color:#fff;margin:0;font-size:18px">TraciumSST - {title}</h2></div>
<div style="padding:20px;background:#fff;border:1px solid #E2E8F0">
<p style="font-size:14px;color:#334155">{intro}</p>
<table style="width:100%;border-collapse:collapse;margin:16px 0"><thead><tr style="background:#1F3C5E"><th style="padding:8px;color:#fff;font-size:12px;text-align:left">Accion</th><th style="padding:8px;color:#fff;font-size:12px">Fecha Limite</th><th style="padding:8px;color:#fff;font-size:12px">Responsable</th><th style="padding:8px;color:#fff;font-size:12px">Estado</th></tr></thead><tbody>{rows}</tbody></table>
<p style="font-size:12px;color:#94A3B8">Este es un mensaje automatico de TraciumSST. Por favor tome las acciones correspondientes.</p>
</div></div>'''

def build_weekly_summary_html(data):
    """Build HTML for weekly summary digest"""
    sections = ""
    for company_name, info in data.items():
        overdue = info.get("overdue", 0)
        upcoming = info.get("upcoming", 0)
        total = info.get("total", 0)
        bar_color = "#D90429" if overdue > 0 else "#FFC300" if upcoming > 0 else "#2A9D8F"
        sections += f'<div style="padding:12px;margin:8px 0;background:#F8F9FA;border-left:4px solid {bar_color};border-radius:4px"><b>{company_name}</b><br/><span style="font-size:12px;color:#475569">Total: {total} planes | <span style="color:#D90429">{overdue} vencidos</span> | <span style="color:#FFC300">{upcoming} proximos</span></span></div>'
    return f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
<div style="background:#1F3C5E;padding:16px;border-radius:8px 8px 0 0"><h2 style="color:#fff;margin:0;font-size:18px">TraciumSST - Resumen Semanal de Planes de Accion</h2></div>
<div style="padding:20px;background:#fff;border:1px solid #E2E8F0">{sections}
<p style="font-size:12px;color:#94A3B8;margin-top:16px">Resumen generado automaticamente por TraciumSST.</p></div></div>'''


@api_router.post("/alerts/send-plan-alerts")
async def send_plan_alerts(request: Request, user=Depends(require_role("admin"))):
    """Manually trigger action plan email alerts"""
    body = await request.json()
    days_before = body.get("days_before", [5, 3, 1, 0])
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_dt = datetime.now(timezone.utc)

    # Get all open action plans with due dates
    plans = await db.action_plans.find({"status": {"$in": ["open", "in_progress"]}, "due_date": {"$ne": ""}}, {"_id": 0}).to_list(500)
    if not plans:
        return {"message": "No hay planes de accion pendientes con fecha limite", "sent": 0}

    # Get responsible SST user emails (sgsst_manager role)
    sgsst_users = await db.users.find({"role": {"$in": ["owner", "admin", "sgsst_manager"]}, "active": {"$ne": False}}, {"_id": 0}).to_list(50)

    overdue = []
    due_today = []
    upcoming = []

    for p in plans:
        try:
            due = datetime.strptime(p["due_date"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            continue
        diff = (due - today_dt).days
        # Get company name
        company = await db.companies.find_one({"company_id": p.get("company_id", "")}, {"_id": 0, "name": 1})
        p["company_name"] = company.get("name", "N/A") if company else "N/A"
        if diff < 0:
            overdue.append(p)
        elif diff == 0:
            due_today.append(p)
        elif diff in days_before or diff <= max(days_before):
            upcoming.append(p)

    sent = 0
    # Group by company for alerts
    for alert_type, alert_plans in [("overdue", overdue), ("due_today", due_today), ("upcoming", upcoming)]:
        if not alert_plans:
            continue
        by_company = {}
        for p in alert_plans:
            cn = p.get("company_name", "N/A")
            by_company.setdefault(cn, []).append(p)
        for cn, company_plans in by_company.items():
            html = build_alert_email_html(company_plans, cn, alert_type)
            subject_map = {"overdue": f"ALERTA: {len(company_plans)} planes vencidos - {cn}", "due_today": f"Planes que vencen HOY - {cn}", "upcoming": f"Planes proximos a vencer - {cn}"}
            subject = subject_map.get(alert_type, "Alerta TraciumSST")
            for u in sgsst_users:
                if u.get("email"):
                    await send_email(u["email"], subject, html)
                    sent += 1

    # Log alert
    await db.alert_logs.insert_one({
        "type": "plan_alerts", "sent": sent,
        "overdue": len(overdue), "due_today": len(due_today), "upcoming": len(upcoming),
        "triggered_by": user.get("name", ""), "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": f"Alertas enviadas: {sent} emails", "sent": sent, "overdue": len(overdue), "due_today": len(due_today), "upcoming": len(upcoming)}


@api_router.post("/alerts/send-weekly-summary")
async def send_weekly_summary(request: Request, user=Depends(require_role("admin"))):
    """Send weekly summary digest to owner/admin"""
    companies = await db.companies.find({}, {"_id": 0}).to_list(100)
    summary_data = {}
    for c in companies:
        cid = c["company_id"]
        total = await db.action_plans.count_documents({"company_id": cid, "status": {"$in": ["open", "in_progress"]}})
        if total == 0:
            continue
        plans = await db.action_plans.find({"company_id": cid, "status": {"$in": ["open", "in_progress"]}, "due_date": {"$ne": ""}}, {"_id": 0}).to_list(200)
        today = datetime.now(timezone.utc)
        overdue = 0
        upcoming = 0
        for p in plans:
            try:
                due = datetime.strptime(p["due_date"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                diff = (due - today).days
                if diff < 0:
                    overdue += 1
                elif diff <= 7:
                    upcoming += 1
            except (ValueError, TypeError):
                continue
        summary_data[c.get("name", cid)] = {"total": total, "overdue": overdue, "upcoming": upcoming}

    if not summary_data:
        return {"message": "No hay planes de accion activos", "sent": 0}

    html = build_weekly_summary_html(summary_data)
    owner_admins = await db.users.find({"role": {"$in": ["owner", "admin"]}, "active": {"$ne": False}}, {"_id": 0}).to_list(20)
    sent = 0
    for u in owner_admins:
        if u.get("email"):
            await send_email(u["email"], "TraciumSST - Resumen Semanal de Planes de Accion", html)
            sent += 1

    await db.alert_logs.insert_one({
        "type": "weekly_summary", "sent": sent, "companies": len(summary_data),
        "triggered_by": user.get("name", ""), "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": f"Resumen semanal enviado a {sent} destinatarios", "sent": sent, "companies": len(summary_data)}


@api_router.get("/alerts/config")
async def get_alert_config(user=Depends(require_role("admin"))):
    """Get alert configuration"""
    config = await db.alert_config.find_one({"type": "plan_alerts"}, {"_id": 0})
    if not config:
        config = {"type": "plan_alerts", "days_before": [5, 3, 1, 0], "enabled": True, "frequency": "on_demand"}
    logs = await db.alert_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(10)
    return {"config": config, "recent_logs": logs}


@api_router.put("/alerts/config")
async def update_alert_config(request: Request, user=Depends(require_role("admin"))):
    """Update alert configuration"""
    body = await request.json()
    allowed = {"days_before", "enabled", "frequency"}
    update = {k: v for k, v in body.items() if k in allowed}
    update["type"] = "plan_alerts"
    update["updated_by"] = user.get("name", "")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.alert_config.update_one({"type": "plan_alerts"}, {"$set": update}, upsert=True)
    return {"message": "Configuracion actualizada"}


# ==================== RBAC INFO ====================

@api_router.get("/rbac/permissions")
async def get_permissions(user=Depends(get_current_user)):
    """Return current user's permissions based on role"""
    role = user.get("role", "collaborator")
    owner = is_owner(user)
    return {
        "role": role,
        "is_owner": owner,
        "can_write": owner or role in WRITE_ROLES,
        "can_audit_write": owner or role in AUDIT_WRITE_ROLES,
        "can_report_incidents": owner or role in INCIDENT_REPORT_ROLES,
        "can_manage_users": owner or role in ("admin", "owner"),
        "can_manage_companies": owner or role in ("admin", "owner"),
        "can_view_all_companies": owner or role in ("admin", "owner"),
        "can_download_reports": owner or role in AUDIT_WRITE_ROLES,
        "can_view_audits": owner or role in AUDIT_WRITE_ROLES,
        "can_edit_audit_items": owner or role in ("admin", "auditor"),
        "can_edit_action_plans": owner or role in AUDIT_WRITE_ROLES,
        "can_use_ai_narrative": owner or role in ("admin", "auditor"),
        "can_view_documents": owner or role in ("admin", "sgsst_manager"),
        "can_view_hazards": owner or role in ("admin", "sgsst_manager", "area_leader"),
        "can_view_training": owner or role in ("admin", "sgsst_manager"),
        "can_view_reports": owner or role in ("admin", "sgsst_manager"),
        "can_view_implementation": owner or role in ("admin", "sgsst_manager"),
    }

# ==================== PROFESSIONAL AUDIT REPORT PDF ====================

AUDIT_STATUS_LABELS = {"planned": "Programada", "assigned": "Asignada", "in_progress": "En Ejecucion", "evidence_review": "Revision Evidencias", "findings_review": "Revision Hallazgos", "action_plan": "Plan de Accion", "follow_up": "Seguimiento", "closed": "Cerrada", "reviewed": "Revisada"}

@api_router.get("/audits/{audit_id}/report/pdf")
async def generate_audit_report_pdf(audit_id: str, user=Depends(get_current_user)):
    await _require_pdf_download_access(audit_id, user)
    """Generate professional audit report PDF with brand identity"""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    audit = await db.audits.find_one({"audit_id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Auditoria no encontrada")
    cid = audit.get("company_id", get_company_id(user))
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    checklist = await db.audit_checklist.find({"audit_id": audit_id}, {"_id": 0}).to_list(200)
    findings = await db.findings.find({"audit_id": audit_id, "status": {"$ne": "resolved_by_compliance"}}, {"_id": 0}).to_list(100)
    action_plans = await db.action_plans.find({"audit_id": audit_id, "status": {"$ne": "closed"}}, {"_id": 0}).to_list(100)
    review = audit.get("management_review")

    # Brand colors from letterhead
    CORAL = colors.HexColor("#F2A292")
    DARK_BLUE = colors.HexColor("#1F3C5E")
    BLUE = colors.HexColor("#0047AB")
    LIGHT_BG = colors.HexColor("#F8F9FA")
    WHITE = colors.white

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.2*inch, bottomMargin=1*inch, leftMargin=0.8*inch, rightMargin=0.8*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18, textColor=DARK_BLUE, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=11, textColor=CORAL, alignment=TA_CENTER, spaceAfter=12)
    heading_style = ParagraphStyle('CustomH2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, textColor=DARK_BLUE, spaceBefore=16, spaceAfter=8)
    subheading_style = ParagraphStyle('CustomH3', parent=styles['Heading3'], fontName='Helvetica-Bold', fontSize=11, textColor=BLUE, spaceBefore=10, spaceAfter=6)
    body_style = ParagraphStyle('CustomBody', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=13, alignment=TA_JUSTIFY)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

    elements = []

    # === COVER PAGE ===
    _cover_logo = _company_logo_flowable(company, max_w=180, max_h=90)
    if _cover_logo:
        elements.append(Spacer(1, 40))
        _cover_logo.hAlign = 'CENTER'
        elements.append(_cover_logo)
        elements.append(Spacer(1, 20))
    else:
        elements.append(Spacer(1, 80))
    elements.append(HRFlowable(width="100%", thickness=3, color=CORAL))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("INFORME FINAL DE AUDITORIA", title_style))
    elements.append(Paragraph("SISTEMA DE GESTION DE SEGURIDAD Y SALUD EN EL TRABAJO", subtitle_style))
    elements.append(Paragraph("SG-SST", ParagraphStyle('Big', parent=title_style, fontSize=28, spaceAfter=20)))
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    elements.append(Spacer(1, 30))

    cover_data = [
        ["Empresa:", company.get("name", "N/A") if company else "N/A"],
        ["NIT:", company.get("nit", "N/A") if company else "N/A"],
        ["Tipo de Auditoria:", "Interna" if audit.get("audit_type") == "internal" else "Externa"],
        ["Auditor:", audit.get("auditor", "N/A")],
        ["Fecha:", audit.get("scheduled_date", "N/A")],
        ["Estado:", AUDIT_STATUS_LABELS.get(audit.get("status", ""), audit.get("status", ""))],
    ]
    cover_table = Table(cover_data, colWidths=[150, 300])
    cover_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), DARK_BLUE),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(cover_table)
    elements.append(Spacer(1, 40))
    elements.append(Paragraph("Stephania Ceballos - Grow Human. Lead Better.", small_style))
    elements.append(Paragraph("Tel: (+57) 321 620 8039 | stephania.ceballos@laofi.onmicrosoft.com", small_style))
    elements.append(Paragraph("Ciudadela Complex, Llanogrande Lote 57-58 Rionegro - Antioquia", small_style))
    elements.append(PageBreak())

    # === 1. INFORMACION GENERAL ===
    elements.append(Paragraph("1. INFORMACION GENERAL", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    elements.append(Spacer(1, 8))

    info_data = [
        ["Tipo de Auditoria", "Interna" if audit.get("audit_type") == "internal" else "Externa"],
        ["Tipo de Informe", "Informe Final"],
        ["Empresa Auditada", company.get("name", "") if company else ""],
        ["NIT", company.get("nit", "") if company else ""],
        ["Trabajadores", str(company.get("workers_count", "")) if company else ""],
        ["Nivel de Riesgo", str(company.get("risk_level", "")) if company else ""],
        ["Proceso Auditado", Paragraph(audit.get("scope", "Todos los procesos del SG-SST"), body_style)],
        ["Objetivo", Paragraph(audit.get("objective", "Verificar cumplimiento del SG-SST"), body_style)],
        ["Alcance", Paragraph(audit.get("scope", ""), body_style)],
        ["Criterios", Paragraph(audit.get("criteria", "Resolucion 0312/2019, Decreto 1072/2015"), body_style)],
        ["Marco Legal", "Decreto 1072 de 2015 - Resolucion 0312 de 2019"],
        ["Auditor Lider", audit.get("auditor", "")],
        ["Auditores Adicionales", ", ".join(audit.get("additional_auditors", [])) or "N/A"],
        ["Miembro COPASST", f"{audit.get('copasst_member', {}).get('name', 'N/A')} ({audit.get('copasst_member', {}).get('role', '')})"],
        ["Fecha Inicio", f"{audit.get('scheduled_date', '')} {audit.get('start_time', '')}"],
        ["Fecha Fin", f"{audit.get('end_date', '')} {audit.get('end_time', '')}"],
        ["Limitaciones", "Ninguna identificada"],
    ]
    info_table = Table(info_data, colWidths=[180, 280])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), LIGHT_BG),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)

    # === 2. DESARROLLO DE LA AUDITORIA ===
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("2. DESARROLLO DE LA AUDITORIA", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))

    # 2.1 Resumen Ejecutivo
    elements.append(Paragraph("2.1 Resumen Ejecutivo", subheading_style))
    # Custom narrative prepend if auditor has filled it via AI
    _narrative_report = audit.get("narrative_report", "").strip() if audit.get("narrative_report") else ""
    if _narrative_report:
        for para in _narrative_report.split("\n"):
            p = para.strip()
            if p:
                elements.append(Paragraph(p, body_style))
                elements.append(Spacer(1, 4))
    total_ck = len(checklist)
    na_count = sum(1 for c in checklist if c.get("result") == "no_aplica")
    evaluable = total_ck - na_count
    cumple = sum(1 for c in checklist if c.get("result") == "cumple")
    no_cumple = sum(1 for c in checklist if c.get("result") == "no_cumple")
    parcial = sum(1 for c in checklist if c.get("result") == "parcial")
    pct = round((cumple / evaluable * 100) if evaluable > 0 else 0, 1)

    # PHVA breakdown
    phva_data = [["Ciclo PHVA", "Estandares", "Cumple", "No Cumple", "% Cumplimiento"]]
    for phase in ["PLANEAR", "HACER", "VERIFICAR", "ACTUAR"]:
        ph_items = [c for c in checklist if c.get("phva") == phase and c.get("result") != "no_aplica"]
        ph_cumple = sum(1 for c in ph_items if c.get("result") == "cumple")
        ph_pct = round((ph_cumple / len(ph_items) * 100) if ph_items else 0, 1)
        phva_data.append([phase, str(len(ph_items)), str(ph_cumple), str(len(ph_items) - ph_cumple), f"{ph_pct}%"])
    phva_data.append(["TOTAL", str(evaluable), str(cumple), str(no_cumple), f"{pct}%"])

    phva_table = Table(phva_data, colWidths=[100, 80, 80, 80, 100])
    phva_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), LIGHT_BG),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(phva_table)
    elements.append(Spacer(1, 8))

    classification = "Critico" if pct < 60 else "Moderadamente aceptable" if pct < 86 else "Aceptable"
    action = "Realizar plan de mejoramiento inmediato" if pct < 60 else "Enviar plan de mejora a ARL" if pct < 86 else "Mantener calificacion e incluir mejoras"
    elements.append(Paragraph(f"<b>Puntaje obtenido:</b> {pct}% | <b>Clasificacion:</b> {classification}", body_style))
    elements.append(Paragraph(f"<b>Accion recomendada:</b> {action}", body_style))

    # 2.2 Metodologia
    elements.append(Paragraph("2.2 Metodologia", subheading_style))
    elements.append(Paragraph("La auditoria se realizo mediante revision documental, entrevistas con responsables de proceso y verificacion de evidencias, conforme a los criterios establecidos en la Resolucion 0312 de 2019 y el Decreto 1072 de 2015. Se evaluo cada estandar minimo aplicable segun la clasificacion de la empresa.", body_style))

    # 2.3 Resultados por estándar
    elements.append(Paragraph("2.3 Resultados por Estandar", subheading_style))
    if checklist:
        cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=7, leading=9)
        cell_bold = ParagraphStyle('cellb', fontName='Helvetica-Bold', fontSize=7, leading=9)
        result_colors = {"Cumple": "#2A9D8F", "No Cumple": "#D90429", "Parcial": "#FFC300", "No Aplica": "#94A3B8", "Pendiente": "#94A3B8"}
        std_data = [["Cod.", "Estandar", "Resultado", "Observaciones del Auditor"]]
        for c in checklist:
            result_text = {"cumple": "Cumple", "no_cumple": "No Cumple", "parcial": "Parcial", "no_aplica": "No Aplica"}.get(c.get("result", ""), "Pendiente")
            rc = result_colors.get(result_text, "#94A3B8")
            obs_text = c.get("observations", "") or ""
            std_data.append([
                Paragraph(c.get("code", ""), cell_bold),
                Paragraph(c.get("description", ""), cell_style),
                Paragraph(f'<font color="{rc}"><b>{result_text}</b></font>', cell_style),
                Paragraph(obs_text, cell_style),
            ])
        std_table = Table(std_data, colWidths=[35, 160, 55, 220])
        std_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DARK_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(std_table)

    # === 2.4 Análisis Art. 2.2.4.6.30 Decreto 1072 ===
    elements.append(PageBreak())
    elements.append(Paragraph("2.4 Analisis del Alcance de la Auditoria - Art. 2.2.4.6.30 Decreto 1072 de 2015", subheading_style))
    art_points = [
        "Cumplimiento de la politica de SST",
        "Resultado de los indicadores de estructura, proceso y resultado",
        "Participacion de los trabajadores",
        "Responsabilidad y rendicion de cuentas",
        "Mecanismos de comunicacion del SG-SST",
        "Planificacion, desarrollo y aplicacion del sistema",
        "Gestion del cambio",
        "Consideracion de la SST en nuevas adquisiciones",
        "Alcance y aplicacion del SG-SST frente a proveedores y contratistas",
        "Supervision y medicion de los resultados",
        "Investigacion de incidentes, accidentes de trabajo y enfermedades laborales",
        "Desarrollo del proceso de auditoria",
        "Evaluacion por parte de la alta direccion",
    ]
    for i, point in enumerate(art_points, 1):
        elements.append(Paragraph(f"<b>{i}. {point}:</b> Se verifico conforme al alcance de la auditoria y la evidencia disponible.", body_style))
        elements.append(Spacer(1, 4))

    # === 3. FORTALEZAS ===
    elements.append(Paragraph("3. FORTALEZAS", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    strengths = [c for c in checklist if c.get("result") == "cumple"]
    if strengths:
        for s in strengths[:10]:
            elements.append(Paragraph(f"- {s.get('code', '')}: {s.get('description', '')}", body_style))
    else:
        elements.append(Paragraph("Se identificaran fortalezas durante la ejecucion completa.", body_style))

    # === 4. HALLAZGOS ===
    elements.append(Paragraph("4. HALLAZGOS DE LA AUDITORIA", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    ai_findings_text = audit.get("ai_redacted_findings", "")
    if ai_findings_text:
        elements.append(Paragraph(ai_findings_text, body_style))
        elements.append(Spacer(1, 6))
    if findings:
        for f in findings:
            ftype = {"no_conformity": "No Conformidad", "observation": "Observacion", "improvement": "Oportunidad de Mejora"}.get(f.get("finding_type", ""), "Hallazgo")
            elements.append(Paragraph(f"<b>[{ftype}]</b> Estandar {f.get('standard_ref', 'N/A')} - Area: {f.get('area', 'N/A')}", ParagraphStyle('FindingHead', parent=body_style, fontName='Helvetica-Bold', textColor=colors.HexColor("#D90429"))))
            elements.append(Paragraph(f.get("description", ""), body_style))
            if f.get("responsible"):
                elements.append(Paragraph(f"Responsable: {f['responsible']} | Fecha: {f.get('due_date', 'N/A')}", ParagraphStyle('FindingMeta', parent=body_style, fontSize=8, textColor=colors.grey)))
            elements.append(Spacer(1, 6))
    elif not ai_findings_text:
        elements.append(Paragraph("No se identificaron hallazgos.", body_style))

    # === 4.1 FORTALEZAS ===
    ai_strengths = audit.get("ai_redacted_strengths", "")
    if ai_strengths:
        elements.append(Paragraph("4.1 FORTALEZAS IDENTIFICADAS", subheading_style))
        elements.append(Paragraph(ai_strengths, body_style))
        elements.append(Spacer(1, 6))

    # === 5. RECOMENDACIONES Y PLAN DE ACCION ===
    elements.append(Paragraph("5. RECOMENDACIONES Y PLAN DE ACCION", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    ai_recs = audit.get("ai_redacted_recommendations", "")
    if ai_recs:
        elements.append(Paragraph(ai_recs, body_style))
        elements.append(Spacer(1, 6))
    if action_plans:
        ap_cell = ParagraphStyle('apcell', fontName='Helvetica', fontSize=8, leading=10)
        plan_data = [["Tipo", "Accion", "Responsable", "Fecha", "Estado"]]
        for ap in action_plans:
            atype = {"corrective": "Correctiva", "preventive": "Preventiva", "improvement": "Mejora"}.get(ap.get("action_type", ""), "Accion")
            status_l = {"open": "Abierto", "in_progress": "En Progreso", "closed": "Cerrado"}.get(ap.get("status", ""), ap.get("status", ""))
            plan_data.append([atype, Paragraph(ap.get("action", ""), ap_cell), ap.get("responsible", "N/A"), ap.get("due_date", "N/A"), status_l])
        plan_table = Table(plan_data, colWidths=[55, 210, 75, 60, 55])
        plan_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DARK_BLUE), ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(plan_table)
    elif not ai_recs:
        elements.append(Paragraph("Se generaran planes de accion basados en los hallazgos.", body_style))

    # === 6. CONCLUSIONES ===
    elements.append(Paragraph("6. CONCLUSIONES", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    ai_conclusions = audit.get("ai_redacted_conclusions", "")
    if ai_conclusions:
        elements.append(Paragraph(ai_conclusions, body_style))
    elif audit.get("executive_summary"):
        elements.append(Paragraph(audit["executive_summary"], body_style))
    else:
        na_text = f" ({na_count} estandares no aplican)" if na_count > 0 else ""
        elements.append(Paragraph(f"El Sistema de Gestion de SST de {company.get('name', 'la empresa') if company else 'la empresa'} obtiene un puntaje de {pct}% ({classification}). Se evaluaron {evaluable} estandares{na_text}, de los cuales {cumple} cumplen, {no_cumple} no cumplen y {parcial} cumplen parcialmente. Se identificaron {len(findings)} hallazgos y {len(action_plans)} planes de accion.", body_style))

    # === 7. FIRMAS ===
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("7. FIRMAS", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    elements.append(Spacer(1, 20))

    # Professional signature block for lead auditor
    sig = AUDITOR_SIGNATURE
    sig_name = audit.get("auditor", sig["name"])
    elements.append(Paragraph("_________________________________", ParagraphStyle('sigline', alignment=1, fontSize=9)))
    elements.append(Paragraph(f"<b>{sig_name}</b>", ParagraphStyle('signame', fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=DARK_BLUE)))
    for line in sig["title_lines"]:
        elements.append(Paragraph(line, ParagraphStyle('sigtitle', fontName='Helvetica', fontSize=8, alignment=1, textColor=colors.HexColor("#475569"), leading=11)))
    elements.append(Spacer(1, 20))

    firma_data = [
        ["________________________", "________________________"],
        ["Reviso", "Aprobo"],
        [audit.get("created_by", "Responsable SST"), review.get("reviewer", "Alta Direccion") if review else "Alta Direccion"],
    ]
    firma_table = Table(firma_data, colWidths=[200, 200])
    firma_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(firma_table)

    # === 8. ANEXOS - SOPORTE IDONEIDAD DEL AUDITOR ===
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("8. ANEXOS - SOPORTE IDONEIDAD DEL AUDITOR", heading_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=CORAL))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("A continuacion se relacionan los documentos que acreditan la idoneidad y competencia profesional del auditor lider para la ejecucion de la presente auditoria al Sistema de Gestion de Seguridad y Salud en el Trabajo (SG-SST):", body_style))
    elements.append(Spacer(1, 8))

    annex_data = [["No.", "Documento", "Descripcion"]]
    for i, annex in enumerate(sig["annexes"], 1):
        annex_data.append([
            str(i),
            Paragraph(f"<b>{annex['title']}</b>", ParagraphStyle('at', fontName='Helvetica-Bold', fontSize=8, leading=10)),
            Paragraph(annex["desc"], ParagraphStyle('ad', fontName='Helvetica', fontSize=8, leading=10)),
        ])
    annex_table = Table(annex_data, colWidths=[30, 170, 270])
    annex_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK_BLUE), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
    ]))
    elements.append(annex_table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<i>Nota: Los documentos originales de soporte reposan en los archivos del proceso de auditoria y pueden ser consultados por los interesados.</i>", ParagraphStyle('note', fontName='Helvetica-Oblique', fontSize=7, textColor=colors.HexColor("#94A3B8"), leading=10)))

    # Footer
    elements.append(Spacer(1, 30))
    elements.append(HRFlowable(width="100%", thickness=1, color=CORAL))
    elements.append(Paragraph(f"{sig_name} | (+57) 321 620 8039 | stephania.ceballos@laofi.onmicrosoft.com", small_style))
    elements.append(Paragraph("Ciudadela Complex, Llanogrande Lote 57-58 Rionegro - Antioquia", small_style))

    doc.build(elements)
    buf.seek(0)
    company_name = company.get("name", "empresa").replace(" ", "_") if company else "empresa"
    filename = f"Informe_Auditoria_SGSST_{company_name}_{audit.get('scheduled_date', '')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    try:
        init_storage()
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    # Seed admin user with email/password
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@traciumsst.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "TraciumSST2026!")
    existing = await db.users.find_one({"email": admin_email}, {"_id": 0})
    if not existing:
        admin_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": admin_id, "email": admin_email, "name": "Administrador TraciumSST",
            "password_hash": hash_password(admin_password), "role": "admin",
            "auth_type": "email", "picture": "", "company_ids": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin user seeded: {admin_email}")
    elif not existing.get("password_hash"):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password), "auth_type": "email"}})
        logger.info(f"Admin user updated with password: {admin_email}")
    # Ensure Stephania is owner
    owner_email = OWNER_EMAIL
    owner_user = await db.users.find_one({"email": owner_email}, {"_id": 0})
    if owner_user and owner_user.get("role") != "owner":
        await db.users.update_one({"email": owner_email}, {"$set": {"role": "owner"}})
        logger.info(f"Owner role set for: {owner_email}")
    # Create indexes
    try:
        await db.users.create_index("email")
        await db.login_attempts.create_index("identifier")
    except Exception as e:
        logger.warning(f"Index creation: {e}")

    # Migration: scrub 'default' leak from non-admin/non-owner user company_ids
    try:
        affected = await db.users.update_many(
            {
                "role": {"$nin": ["admin", "owner"]},
                "email": {"$ne": OWNER_EMAIL},
                "company_ids": "default"
            },
            {"$pull": {"company_ids": "default"}}
        )
        if affected.modified_count:
            logger.info(f"Migration: removed 'default' company_id leak from {affected.modified_count} users")
    except Exception as e:
        logger.warning(f"Company isolation migration failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
